#!/usr/bin/env python3
"""
Everlight Intel Center -- Master extractor.

Reads the FREE RESOURCES folder (spreadsheets + Photos), OCRs every image,
parses each line into a {name, url, domain, description} candidate, runs the
rule-based categorizer, and writes the canonical database in 5 formats:

  database/everlight_resources_master.csv
  database/everlight_resources_master.json
  database/everlight_resources.sqlite
  database/resource_review_queue.csv      # low-confidence rows for human eyes
  guides/everlight_resources_master.html  # branded report (gold-on-dark)

Spec: Free Resources handoff (2026-05-11) + Everlight branded HTML doctrine.
"""

from __future__ import annotations

import csv
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import time
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import yaml
from PIL import Image, ImageOps, ImageFilter

ROOT = Path("/mnt/sdcard/AA_MY_DRIVE/06_DEVELOPMENT/everlight_os/intel_center")
SOURCE = Path("/mnt/sdcard/AA_MY_DRIVE/FREE RESOURCES")

RAW_OCR = ROOT / "extracted" / "raw_ocr"
PARSED = ROOT / "extracted" / "parsed"
DB_DIR = ROOT / "database"
GUIDES = ROOT / "guides"
LOGS = ROOT / "logs"
for p in (RAW_OCR, PARSED, DB_DIR, GUIDES, LOGS):
    p.mkdir(parents=True, exist_ok=True)

# ---- 1. URL / domain regex --------------------------------------------------
# OCR mangles bullets (•, ¢, *, +, «) and sometimes splits domains. Tolerate
# a leading non-word run, allow trailing punctuation, capture everything after
# " - " as description.
DOMAIN_RE = re.compile(
    r"""
    ^[\W_]*                                # any leading bullet noise
    (?P<domain>
        [a-zA-Z0-9][a-zA-Z0-9\-]{0,62}     # label
        (?:\.[a-zA-Z0-9][a-zA-Z0-9\-]{0,62}){1,3}    # 1-3 more labels
    )
    \s*[-–—:]+\s*                          # separator (any dash variant or colon)
    (?P<desc>.+?)\s*$
    """,
    re.VERBOSE,
)

# Common OCR confusions for TLDs we care about
TLD_FIXES = {
    ".com1": ".com", ".coni": ".com", ".coin": ".com",
    ".io1": ".io", ".oig": ".org", ".or9": ".org",
    ".n3t": ".net",
}

VALID_TLDS = {
    "com", "org", "net", "io", "ai", "app", "dev", "co", "us", "uk", "ca",
    "me", "tv", "fm", "xyz", "tech", "tools", "design", "art", "studio",
    "page", "site", "online", "info", "biz", "pro", "media", "video",
    "store", "shop", "live", "today", "news", "world", "global", "agency",
    "club", "fund", "money", "finance", "bank", "earth", "space", "city",
    "land", "link", "blog", "guide", "wiki", "press", "graphics", "games",
    "host", "cloud", "data", "ml", "py", "js", "rb", "go", "edu", "gov",
    "id", "to", "ly", "sh", "is", "in", "de", "fr", "it", "es", "nl",
    "ch", "se", "no", "fi", "be", "at", "pl", "ru", "br", "mx", "ar",
    "au", "nz", "jp", "kr", "cn", "hk", "tw", "sg", "il", "tr", "za",
    "tools", "services", "directory", "watch", "report", "lol",
    "rocks", "ninja", "guru", "expert", "expert",
}


def fix_tld(domain: str) -> str:
    d = domain.lower().strip(".,;:")
    for bad, good in TLD_FIXES.items():
        if d.endswith(bad):
            d = d[: -len(bad)] + good
    return d


def is_plausible_domain(d: str) -> bool:
    if "." not in d:
        return False
    parts = d.split(".")
    if any(len(p) == 0 for p in parts):
        return False
    tld = parts[-1].lower()
    return tld in VALID_TLDS


# ---- 2. OCR a single image --------------------------------------------------
def preprocess(in_path: Path, out_path: Path) -> Path:
    """Upscale + grayscale + autocontrast + threshold for crisper OCR."""
    img = Image.open(in_path)
    # Upscale 2x to help small mobile-screenshot text
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    img = ImageOps.grayscale(img)
    img = ImageOps.autocontrast(img, cutoff=2)
    # Slight sharpen
    img = img.filter(ImageFilter.SHARPEN)
    img.save(out_path, "PNG", optimize=True)
    return out_path


def ocr(img_path: Path) -> str:
    pre_path = img_path.with_suffix(".pre.png")
    try:
        preprocess(img_path, pre_path)
        out = subprocess.run(
            ["tesseract", str(pre_path), "-", "--psm", "6", "-l", "eng"],
            capture_output=True, text=True, timeout=120,
        )
        return out.stdout
    except subprocess.TimeoutExpired:
        return ""
    finally:
        if pre_path.exists():
            pre_path.unlink()


# ---- 3. Parse OCR text into candidate rows ---------------------------------
@dataclass
class Resource:
    id: str = ""
    source_type: str = ""
    source_file: str = ""
    raw_text: str = ""
    name: str = ""
    url: str = ""
    domain: str = ""
    category: str = ""
    subcategory: str = ""
    purpose: str = ""
    department: str = ""
    agent_owner: str = ""
    use_case: str = ""
    data_type: str = ""
    access_type: str = "free"
    cost_level: str = "free"
    risk_level: str = "low"
    legal_notes: str = ""
    setup_steps: str = ""
    example_prompt: str = ""
    tags: list = field(default_factory=list)
    priority_score: int = 3
    verified_status: str = "unverified"
    last_checked: str = ""
    notes: str = ""


def extract_header(text: str) -> str:
    """
    Pull the topic header from a ByteClave-style screenshot.
    Strategy: the first 3 non-empty, non-bullet, non-domain lines in the OCR
    are the brand banner + topic title. We strip "ByteClave" and join the rest.
    """
    candidates = []
    for raw_line in text.splitlines()[:8]:
        line = raw_line.strip()
        if not line or len(line) < 4:
            continue
        if DOMAIN_RE.match(line):
            break
        # Strip leading/trailing dash noise and "ByteClave" branding
        cleaned = re.sub(r"[—\-_=<>]+", " ", line).strip()
        cleaned = re.sub(r"\bByteClave\b", "", cleaned, flags=re.I).strip()
        if cleaned and len(cleaned) > 3:
            candidates.append(cleaned)
        if len(candidates) >= 2:
            break
    return " | ".join(candidates)[:140]


def parse_ocr_block(text: str, source_file: str) -> list[Resource]:
    rows: list[Resource] = []
    header = extract_header(text)
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or len(line) < 8:
            continue
        m = DOMAIN_RE.match(line)
        if not m:
            continue
        domain = fix_tld(m.group("domain"))
        if not is_plausible_domain(domain):
            continue
        desc = m.group("desc").strip().rstrip(".,;:")
        desc = re.sub(r"\s+", " ", desc).strip()
        if len(desc) < 4:
            continue
        rows.append(
            Resource(
                source_type="photo_ocr",
                source_file=source_file,
                raw_text=line,
                name=domain.split(".")[0].replace("-", " ").title(),
                url=f"https://{domain}",
                domain=domain,
                purpose=desc,
                notes=f"photo_topic={header}" if header else "",
            )
        )
    return rows


# ---- 4. Spreadsheet ingestion ----------------------------------------------
def from_spreadsheet(xlsx_path: Path) -> list[Resource]:
    rows: list[Resource] = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        header = [str(c.value or "").strip().lower() for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(r):
                continue
            data = dict(zip(header, [str(v or "").strip() for v in r]))
            # Free Resources xlsx: Website, Category, Description
            site = data.get("website") or data.get("folder/repo") or ""
            cat_hint = data.get("category", "")
            desc = data.get("description", "")
            count_hint = data.get("approx api count", "")
            if not site:
                continue
            domain = fix_tld(site)
            url = f"https://{domain}" if "." in domain else ""
            rows.append(
                Resource(
                    source_type="spreadsheet",
                    source_file=xlsx_path.name,
                    raw_text=" | ".join(filter(None, [site, cat_hint, desc, count_hint])),
                    name=domain.split(".")[0].replace("-", " ").title() if "." in domain else site,
                    url=url,
                    domain=domain,
                    purpose=desc,
                    notes=f"sheet_category={cat_hint}" + (f"; api_count={count_hint}" if count_hint else ""),
                    verified_status="checked",  # came from curated sheet
                    priority_score=4,
                )
            )
    return rows


# ---- 5. Categorizer ---------------------------------------------------------
class Categorizer:
    """
    Word-boundary regex classifier. Sorts hints longest-first so multi-word
    hints win over short ones. Applies TLD overrides only when no hint matches.
    """
    def __init__(self, cfg_path: Path):
        cfg = yaml.safe_load(cfg_path.read_text())
        self.fallback = cfg["fallback"]
        self.tld_overrides = cfg.get("tld_overrides", {})
        self.domain_overrides = cfg.get("domain_overrides", {})  # {category: [tokens]}
        # Build category meta lookup once
        self.cat_meta = {
            cat: {"department": meta["department"], "agent_owner": meta["agent_owner"]}
            for cat, meta in cfg["categories"].items()
        }
        # Compiled (regex, hint, category) sorted longest-first
        self.rules: list[tuple[re.Pattern, str, str]] = []
        for cat, meta in cfg["categories"].items():
            for hint in meta["hints"]:
                # `\b` doesn't trigger between letter and `.`/`-` in Python regex
                # so for hints with dots (sec.gov, ap.org) we use a looser match.
                if any(ch in hint for ch in (".", "/")):
                    pat = re.compile(re.escape(hint.lower()))
                else:
                    pat = re.compile(rf"\b{re.escape(hint.lower())}\b")
                self.rules.append((pat, hint.lower(), cat))
        self.rules.sort(key=lambda x: -len(x[1]))

    def classify(self, r: Resource) -> Resource:
        # Domain-token overrides win first. Match against domain SEGMENTS
        # (split by . and -) -- a segment must start with the token. This
        # avoids amusingplanet matching "plane" or fireship matching "ship".
        dom_lower = r.domain.lower()
        segments = re.split(r"[.\-]", dom_lower)
        for cat, tokens in self.domain_overrides.items():
            for tok in tokens:
                t = tok.lower()
                # If token has a dot in it (e.g. "ap.org"), substring match the full domain
                if "." in t:
                    if t in dom_lower:
                        meta = self.cat_meta.get(cat)
                        if meta:
                            r.category = cat; r.department = meta["department"]
                            r.agent_owner = meta["agent_owner"]; r.subcategory = "domain-override"
                            if r.verified_status != "checked": r.verified_status = "auto-domain"
                            return r
                    continue
                # Otherwise: a domain segment must START WITH the token (length >=3 to avoid noise)
                if len(t) < 3:
                    continue
                if any(seg.startswith(t) for seg in segments):
                    meta = self.cat_meta.get(cat)
                    if meta:
                        r.category = cat; r.department = meta["department"]
                        r.agent_owner = meta["agent_owner"]; r.subcategory = "domain-override"
                        if r.verified_status != "checked": r.verified_status = "auto-domain"
                        return r

        hay = " ".join([r.domain, r.purpose, r.notes, r.raw_text]).lower()
        for pat, hint, cat in self.rules:
            if pat.search(hay):
                meta = self.cat_meta[cat]
                r.category = cat
                r.department = meta["department"]
                r.agent_owner = meta["agent_owner"]
                r.tags = sorted({hint, *r.tags})
                # Spreadsheet rows are "checked" (curated), OCR rows go "auto"
                if r.verified_status != "checked":
                    r.verified_status = "auto"
                return r

        # TLD override before fallback
        for suffix, cat in self.tld_overrides.items():
            if r.domain.endswith(suffix):
                meta = self.cat_meta.get(cat)
                if meta:
                    r.category = cat
                    r.department = meta["department"]
                    r.agent_owner = meta["agent_owner"]
                    r.subcategory = "TLD-routed"
                    if r.verified_status != "checked":
                        r.verified_status = "auto-tld"
                    return r

        r.category = self.fallback["category"]
        r.subcategory = self.fallback["subcategory"]
        r.department = self.fallback["department"]
        r.agent_owner = self.fallback["agent_owner"]
        r.verified_status = self.fallback.get("verified_status", "unverified")
        return r


# ---- 5b. Use-case generator ------------------------------------------------
# Per-category templates. Each gets formatted with: agent, domain, url, purpose.
# This turns every catalogued resource into a usable how-to-use block.
USE_CASE_TEMPLATES = {
    "News & Journalism": {
        "use_case": "{agent} pulls headlines from {domain} as a primary newsroom feed. Scan for breaking flags, geopolitical shifts, and stories that change the operating picture for Everlight Ventures.",
        "setup": "1. Open {url} to inspect the feed format.\n2. Pull live: `intel pull {domain}` (auto-detects RSS).\n3. For ongoing watch: add to `pull-all news` cron.",
    },
    "Weather & Disaster Intel": {
        "use_case": "{agent} watches {domain} for severe-weather, geophysical, or emergency triggers. When a flag fires, escalate to the Disaster Response Desk and check supply-chain exposure.",
        "setup": "1. Open {url} and bookmark the live map / API.\n2. Pull conditions: `intel pull {domain}`.\n3. Add to disaster-watch cron when a named storm is active.",
    },
    "Maps & Geospatial": {
        "use_case": "{agent} uses {domain} for geographic context: parcel data, terrain, satellite imagery, or population overlays. Drop coordinates into the OSINT desk's map layer when investigating a property or event.",
        "setup": "1. Open {url}.\n2. Most map tools require GET-based URLs; capture the share-link for the location of interest.\n3. Save to `Wholesale/Maps_Cache/` if recurring.",
    },
    "Aviation & Maritime": {
        "use_case": "{agent} tracks {domain} for live flight, vessel, or supply-chain movement. Use to confirm physical presence (charter routes, cargo deliveries, anomalies in port traffic).",
        "setup": "1. Open {url}.\n2. Live pull: `intel pull {domain}` -- captures positions if a public API exists.\n3. Cross-reference with logistics-desk shipment tracker.",
    },
    "Space & Science": {
        "use_case": "{agent} pulls {domain} for scientific data, space weather, biological/chemical reference, or research signals. Useful for long-horizon strategy and any health/biotech-adjacent decision.",
        "setup": "1. Open {url} for the dataset.\n2. Many science DBs offer JSON or CSV downloads -- inspect the API.\n3. Cache findings to `cache/science/` for repeat queries.",
    },
    "Economics & Markets": {
        "use_case": "{agent} sources macroeconomic data from {domain}: inflation, employment, GDP, monetary policy. Feed into Bull Archer's daily macro snapshot and Lucrex's strategic decisions.",
        "setup": "1. Open {url}.\n2. Many series have JSON endpoints (e.g., FRED, BLS).\n3. Pull: `intel pull {domain}` for latest release headlines.",
    },
    "Trading & Finance": {
        "use_case": "{agent} pulls {domain} for prices, on-chain flows, options, or portfolio reference. Cross-check against XLM bot signals and broker pipeline stage.",
        "setup": "1. Open {url}.\n2. Pull a snapshot: `intel pull {domain}`.\n3. For automated tracking, wire into `06_DEVELOPMENT/xlm_bot/` if relevant.",
    },
    "OSINT & Investigation": {
        "use_case": "{agent} runs lookups on {domain} for public records, leaked data, archives, or investigative leads. Use only with a documented business purpose; log every query in the compliance audit.",
        "setup": "1. Open {url} and verify access (some sources gate behind free signup).\n2. Pull: `intel pull {domain}`.\n3. Log the investigation reason in `Wholesale/compliance/osint_log.jsonl`.",
    },
    "Cybersecurity": {
        "use_case": "{agent} uses {domain} for threat intel, breach checks, vulnerability lookup, or network reconnaissance on Everlight infrastructure. Authorized targets only.",
        "setup": "1. Open {url}.\n2. Many tools require API keys (free tier usually fine).\n3. Wire findings into the DevOps incident response runbook.",
    },
    "AI & Automation": {
        "use_case": "{agent} reaches for {domain} when building or extending Hive automation: agents, workflows, prompts, models, vector stores. Slot into existing 03_AUTOMATION_CORE/ scripts.",
        "setup": "1. Open {url} for docs.\n2. If it's a model/API, capture endpoint + auth into `06_DEVELOPMENT/everlight_os/secrets/`.\n3. Add a wrapper in `content_tools/` so all agents share one entry point.",
    },
    "APIs & Developer Tools": {
        "use_case": "{agent} uses {domain} to build, test, or deploy. Drop into the engineering toolbox; reference from `49_engineering_assistant.md` runbooks.",
        "setup": "1. Open {url} for docs.\n2. Install / clone as needed.\n3. Add to `06_DEVELOPMENT/` if it becomes part of the stack.",
    },
    "Content Creation": {
        "use_case": "{agent} grabs {domain} for design, video, audio, image, or copy work feeding the Content Engine. Output goes through `02_CONTENT_FACTORY/` and the branded mailer.",
        "setup": "1. Open {url}.\n2. Free tier usually sufficient -- avoid paid until first $10k revenue (Operator Truth doctrine).\n3. Save assets to `02_CONTENT_FACTORY/01_Queue/`.",
    },
    "eCommerce & Product Research": {
        "use_case": "{agent} scouts {domain} for product opportunities, supplier data, or marketplace signals. Outputs feed Pitch Adler's product desk and HIM Loadout affiliate pipeline.",
        "setup": "1. Open {url}.\n2. Capture promising SKUs into `01_BUSINESSES/Everlight_Ventures/HIM_Loadout/scout_log.csv`.\n3. Cross-check with trends data before any inventory commit.",
    },
    "Real Estate & Property": {
        "use_case": "{agent} mines {domain} for parcel data, comps, foreclosures, or rental signals. Feeds the Wholesale pipeline -- Rex Blackwell's deal hunt.",
        "setup": "1. Open {url}.\n2. For automated pulls: see if county records have direct CSV (free).\n3. Push leads into `Wholesale/leads_db.sqlite`.",
    },
    "Logistics & Supply Chain": {
        "use_case": "{agent} watches {domain} for shipping, freight, port congestion, or warehouse signals. Cross-reference with eCommerce demand for arbitrage windows.",
        "setup": "1. Open {url}.\n2. Most logistics tools require enterprise auth; default to free dashboards.\n3. Log macro-disruption events in `intel/logistics_events.jsonl`.",
    },
    "Legal & Compliance": {
        "use_case": "{agent} searches {domain} for case law, regulations, court filings, or compliance guidance. All findings cite the source URL and rule number; never ad-lib legal advice.",
        "setup": "1. Open {url}.\n2. For citation use, copy the canonical reference exactly.\n3. Log compliance touches in `01_BUSINESSES/Everlight_Ventures/Wholesale/compliance/`.",
    },
    "Health & Environment": {
        "use_case": "{agent} uses {domain} for public health, environmental quality, or biotech signals. Useful for both disaster response and product/biotech opportunity scouting.",
        "setup": "1. Open {url}.\n2. Pull: `intel pull {domain}` for latest update.\n3. For ongoing health monitoring, add to disaster-watch cron.",
    },
    "Education & Training": {
        "use_case": "{agent} taps {domain} for learning, research, or reference material. Drop key concepts into Blinko (RAG) so the rest of the Hive can recall them.",
        "setup": "1. Open {url}.\n2. Note key takeaways into `Notes/inbox.org` or push to Blinko via `agentmemory MCP`.\n3. Build a study-track in `05_PERSONAL/` if multi-week.",
    },
    "Self-Hosting & Privacy": {
        "use_case": "{agent} self-hosts or evaluates {domain} as part of the Hive infrastructure. Replaces SaaS where free + open-source is sufficient (Operator Truth: avoid paid SaaS pre-revenue).",
        "setup": "1. Read docs at {url}.\n2. Deploy via `coolify.io` or directly to Oracle E5/Micro.\n3. Add the service file to `03_AUTOMATION_CORE/01_Scripts/oracle_services/`.",
    },
    "Decision Intelligence": {
        "use_case": "{agent} uses {domain} for strategic frameworks, decision trees, scenario analysis, or executive intelligence. Outputs feed Lucrex's command decisions.",
        "setup": "1. Open {url}.\n2. Snapshot key insights into `01_BUSINESSES/Everlight_Ventures/Wealth_OS/`.\n3. For repeat use, build a wrapper in `40_strategic_modeler.md` runbook.",
    },
}


def attach_use_cases(rows: list[Resource]) -> None:
    """Stamp each resource with use_case + setup_steps from its category template."""
    for r in rows:
        tpl = USE_CASE_TEMPLATES.get(r.category)
        if not tpl:
            continue
        ctx = {
            "agent": r.agent_owner or "The owning agent",
            "domain": r.domain or "this source",
            "url": r.url or f"https://{r.domain}",
            "purpose": r.purpose or "(no description)",
        }
        try:
            r.use_case = tpl["use_case"].format(**ctx)
            r.setup_steps = tpl["setup"].format(**ctx)
        except (KeyError, IndexError):
            pass


# ---- 6. Dedupe + finalize --------------------------------------------------
def dedupe(rows: list[Resource]) -> list[Resource]:
    seen: dict[str, Resource] = {}
    for r in rows:
        key = r.domain or r.raw_text[:40]
        if key in seen:
            existing = seen[key]
            # Prefer the one with more description and richer source
            if len(r.purpose) > len(existing.purpose):
                existing.purpose = r.purpose
            if r.source_type == "spreadsheet":
                existing.verified_status = "checked"
                existing.priority_score = max(existing.priority_score, r.priority_score)
            existing.notes = (existing.notes + " | " + r.notes).strip(" |")
        else:
            seen[key] = r
    return list(seen.values())


def assign_ids(rows: list[Resource]) -> None:
    rows.sort(key=lambda r: (r.category, r.domain))
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for i, r in enumerate(rows, start=1):
        r.id = f"EIC-{i:04d}"
        r.last_checked = today


# ---- 7. Writers -------------------------------------------------------------
COLUMNS = [
    "id", "source_type", "source_file", "raw_text", "name", "url", "domain",
    "category", "subcategory", "purpose", "department", "agent_owner",
    "use_case", "data_type", "access_type", "cost_level", "risk_level",
    "legal_notes", "setup_steps", "example_prompt", "tags", "priority_score",
    "verified_status", "last_checked", "notes",
]


def write_csv(rows: list[Resource], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            d = asdict(r)
            d["tags"] = ", ".join(r.tags)
            w.writerow(d)


def write_json(rows: list[Resource], path: Path) -> None:
    path.write_text(json.dumps([asdict(r) for r in rows], indent=2, default=str))


def write_sqlite(rows: list[Resource], path: Path) -> None:
    # Preserve audit table across rebuilds by snapshotting before drop.
    audit_rows = []
    if path.exists():
        try:
            old = sqlite3.connect(path)
            audit_rows = old.execute("SELECT domain, in_use, evidence, checked_at FROM audit").fetchall()
            old.close()
        except sqlite3.OperationalError:
            audit_rows = []
        path.unlink()
    con = sqlite3.connect(path)
    cols_sql = ",\n  ".join(f"{c} TEXT" for c in COLUMNS)
    con.execute(f"CREATE TABLE resources (\n  {cols_sql}\n)")
    con.execute("CREATE INDEX idx_category ON resources(category)")
    con.execute("CREATE INDEX idx_domain ON resources(domain)")
    con.execute("CREATE INDEX idx_agent ON resources(agent_owner)")
    placeholders = ",".join("?" * len(COLUMNS))
    for r in rows:
        d = asdict(r)
        d["tags"] = ", ".join(r.tags)
        con.execute(
            f"INSERT INTO resources VALUES ({placeholders})",
            [str(d[c]) for c in COLUMNS],
        )
    # Restore audit table
    con.execute("CREATE TABLE IF NOT EXISTS audit (domain TEXT PRIMARY KEY, in_use INT, evidence TEXT, checked_at TEXT)")
    for row in audit_rows:
        try:
            con.execute("INSERT OR REPLACE INTO audit VALUES (?,?,?,?)", row)
        except sqlite3.Error:
            pass
    con.commit()
    con.close()


def write_review_queue(rows: list[Resource], path: Path) -> None:
    review = [r for r in rows if r.verified_status == "unverified"
              or (r.subcategory == "Uncategorized")
              or len(r.purpose) < 6]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "domain", "category", "purpose", "raw_text", "source_file"])
        w.writeheader()
        for r in review:
            w.writerow({
                "id": r.id, "domain": r.domain, "category": r.category,
                "purpose": r.purpose, "raw_text": r.raw_text, "source_file": r.source_file,
            })


def write_html(rows: list[Resource], path: Path, stats: dict) -> None:
    """Branded gold-on-dark master index, grouped by category."""
    by_cat: dict[str, list[Resource]] = {}
    for r in rows:
        by_cat.setdefault(r.category, []).append(r)

    cat_blocks = []
    for cat in sorted(by_cat):
        items = sorted(by_cat[cat], key=lambda r: r.domain)
        rows_html = "\n".join(
            f"""<tr>
  <td class="dom"><a href="{r.url}" target="_blank" rel="noopener">{r.domain}</a></td>
  <td>{r.purpose}</td>
  <td class="agent">{r.agent_owner}</td>
  <td class="dept">{r.department}</td>
  <td class="src">{r.source_type}</td>
</tr>""" for r in items
        )
        cat_blocks.append(f"""
<section class="cat">
  <h2>{cat} <span class="ct">({len(items)})</span></h2>
  <table>
    <thead><tr><th>Domain</th><th>Purpose</th><th>Agent</th><th>Department</th><th>Source</th></tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</section>""")

    body = "\n".join(cat_blocks)
    html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Everlight Intel Center -- Master Resource Index</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700;900&family=Inter:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {{
    --gold: #D4A843;
    --gold-soft: #b9913a;
    --dark: #0A0A0A;
    --panel: #15140f;
    --line: #2a2615;
    --text: #E8E8E8;
    --muted: #9b9586;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0; background: var(--dark); color: var(--text);
    font-family: Inter, system-ui, sans-serif; line-height: 1.55;
  }}
  header {{
    border-bottom: 2px solid var(--gold);
    padding: 48px 56px 32px; background:
      radial-gradient(ellipse at top, rgba(212,168,67,0.08), transparent 60%),
      var(--dark);
  }}
  .wordmark {{
    font-family: 'Playfair Display', serif; font-weight: 900;
    color: var(--gold); letter-spacing: 0.15em; font-size: 0.85rem;
    text-transform: uppercase;
  }}
  h1 {{
    font-family: 'Playfair Display', serif; font-weight: 700;
    color: var(--gold); margin: 8px 0 4px; font-size: 2.4rem;
  }}
  header .sub {{ color: var(--muted); font-size: 0.95rem; }}
  .stats {{
    display: flex; flex-wrap: wrap; gap: 28px; margin-top: 24px;
    font-family: 'JetBrains Mono', monospace; font-size: 0.9rem;
  }}
  .stats span b {{ color: var(--gold); margin-left: 6px; }}
  main {{ padding: 32px 56px 80px; max-width: 1200px; margin: 0 auto; }}
  section.cat {{ margin-bottom: 40px; }}
  h2 {{
    font-family: 'Playfair Display', serif; color: var(--gold);
    font-size: 1.4rem; border-bottom: 1px solid var(--line);
    padding-bottom: 8px; margin-bottom: 14px;
  }}
  h2 .ct {{ color: var(--muted); font-family: 'JetBrains Mono', monospace; font-size: 0.85rem; }}
  table {{ width: 100%; border-collapse: collapse; background: var(--panel); border-radius: 6px; overflow: hidden; }}
  th, td {{ text-align: left; padding: 10px 14px; border-bottom: 1px solid var(--line); font-size: 0.92rem; }}
  th {{
    background: rgba(212,168,67,0.06); color: var(--gold);
    font-weight: 500; text-transform: uppercase; letter-spacing: 0.08em; font-size: 0.75rem;
  }}
  tr:last-child td {{ border-bottom: none; }}
  td.dom {{ font-family: 'JetBrains Mono', monospace; font-size: 0.85rem; min-width: 180px; }}
  td.dom a {{ color: var(--gold); text-decoration: none; }}
  td.dom a:hover {{ text-decoration: underline; }}
  td.agent, td.dept {{ color: var(--muted); font-size: 0.82rem; white-space: nowrap; }}
  td.src {{ color: var(--gold-soft); font-family: 'JetBrains Mono', monospace; font-size: 0.75rem; }}
  footer {{
    border-top: 1px solid var(--line); padding: 28px 56px; color: var(--muted);
    font-size: 0.85rem; text-align: center;
  }}
  footer b {{ color: var(--gold); }}
</style>
</head>
<body>
<header>
  <div class="wordmark">Everlight Ventures</div>
  <h1>Intel Center -- Master Resource Index</h1>
  <div class="sub">Free + open-source resources extracted from {stats['source_count']} sources, categorized for Hive dispatch</div>
  <div class="stats">
    <span>Total resources<b>{stats['total']}</b></span>
    <span>Categories<b>{stats['cats']}</b></span>
    <span>From spreadsheets<b>{stats['from_xlsx']}</b></span>
    <span>From OCR<b>{stats['from_ocr']}</b></span>
    <span>Review queue<b>{stats['review']}</b></span>
    <span>Generated<b>{stats['generated']}</b></span>
  </div>
</header>
<main>
{body}
</main>
<footer>
  <b>Lucrex / Everlight Intel Center</b> -- the mind behind the money.
  Built {stats['generated']}. Source: <code>FREE RESOURCES/</code> -- spreadsheets + {stats['photos']} screenshots.
</footer>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")


# ---- 8. Main ---------------------------------------------------------------
def main():
    t0 = time.time()
    skip_ocr = "--no-ocr" in sys.argv
    log = LOGS / f"build_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl"
    log_lines: list[str] = []

    def loglog(msg, **kw):
        entry = {"ts": datetime.now().isoformat(), "msg": msg, **kw}
        log_lines.append(json.dumps(entry))
        print(msg, file=sys.stderr)

    # 1. Spreadsheets
    sheet_paths = [
        SOURCE / "Everlight_Free_Resources_List.xlsx",
        SOURCE / "Photos" / "API_Mega_List_Repository_Index.xlsx",
    ]
    rows: list[Resource] = []
    for p in sheet_paths:
        if p.exists():
            sub = from_spreadsheet(p)
            loglog("spreadsheet_loaded", file=p.name, rows=len(sub))
            rows.extend(sub)
    from_xlsx = len(rows)

    # 2. Photos -- OCR on-demand or load cached parsed candidates
    photos = sorted([p for p in (SOURCE / "Photos").iterdir() if p.suffix.lower() in (".jpg", ".jpeg", ".png")])
    for i, p in enumerate(photos, 1):
        cache = PARSED / f"{p.stem}.json"
        if skip_ocr and cache.exists():
            cached = json.loads(cache.read_text())
            sub = [Resource(**c) for c in cached]
            loglog("ocr_cached", file=p.name, parsed=len(sub))
        else:
            loglog("ocr_start", file=p.name, idx=i, total=len(photos))
            text = ocr(p)
            (RAW_OCR / f"{p.stem}.txt").write_text(text)
            sub = parse_ocr_block(text, p.name)
            cache.write_text(json.dumps([asdict(r) for r in sub], indent=2))
            loglog("ocr_done", file=p.name, parsed=len(sub))
        rows.extend(sub)
    from_ocr = len(rows) - from_xlsx

    # 3. Categorize + dedupe + assign IDs
    cat = Categorizer(ROOT / "config" / "categories.yaml")
    for r in rows:
        cat.classify(r)
    rows = dedupe(rows)
    assign_ids(rows)
    attach_use_cases(rows)

    # 4. Write outputs
    write_csv(rows, DB_DIR / "everlight_resources_master.csv")
    write_json(rows, DB_DIR / "everlight_resources_master.json")
    write_sqlite(rows, DB_DIR / "everlight_resources.sqlite")
    write_review_queue(rows, DB_DIR / "resource_review_queue.csv")

    # 5. Branded HTML report
    cats = sorted({r.category for r in rows})
    review_count = sum(1 for r in rows if r.verified_status == "unverified")
    stats = {
        "total": len(rows),
        "cats": len(cats),
        "from_xlsx": from_xlsx,
        "from_ocr": from_ocr,
        "review": review_count,
        "source_count": len(sheet_paths) + len(photos),
        "photos": len(photos),
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M PT"),
    }
    write_html(rows, GUIDES / "everlight_resources_master.html", stats)

    # 6. Logs
    log.write_text("\n".join(log_lines))
    elapsed = time.time() - t0
    print(json.dumps({
        "status": "ok", "total": len(rows), "from_xlsx": from_xlsx,
        "from_ocr": from_ocr, "categories": len(cats),
        "review_queue": review_count, "elapsed_sec": round(elapsed, 1),
        "outputs": {
            "csv": str(DB_DIR / "everlight_resources_master.csv"),
            "json": str(DB_DIR / "everlight_resources_master.json"),
            "sqlite": str(DB_DIR / "everlight_resources.sqlite"),
            "review": str(DB_DIR / "resource_review_queue.csv"),
            "html": str(GUIDES / "everlight_resources_master.html"),
        },
    }, indent=2))


if __name__ == "__main__":
    main()
