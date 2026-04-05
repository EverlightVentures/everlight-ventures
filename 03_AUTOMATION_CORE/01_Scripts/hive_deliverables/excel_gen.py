"""
Excel Generator - Branded Everlight Ventures spreadsheets.
Uses openpyxl for professional Excel generation.
"""
import os
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path(os.environ.get("DELIVERABLES_DIR", "/tmp/hive_deliverables"))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def generate_excel(title: str, sheets: dict[str, list[dict]],
                   author: str = "Everlight Ventures") -> str:
    """
    Generate a branded Excel workbook.

    Args:
        title: Workbook title (used in filename)
        sheets: Dict of sheet_name -> list of row dicts
        author: Author metadata

    Returns:
        Path to generated .xlsx file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return _fallback_csv(title, sheets)

    wb = Workbook()
    wb.properties.creator = author
    wb.properties.title = title

    # Brand styles
    gold_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="333333")
    title_font = Font(name="Arial", size=16, bold=True, color="1A1A1A")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    first_sheet = True
    for sheet_name, rows in sheets.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        if not rows:
            continue

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
        ws.cell(row=1, column=1, value=f"{title} - {sheet_name}").font = title_font
        ws.row_dimensions[1].height = 30

        # Date row
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M PT')}").font = Font(
            name="Arial", size=9, color="888888"
        )

        # Headers (row 4)
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        alt_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.font = body_font
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_data in rows[:50]:
                val = str(row_data.get(header, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_len + 4, 40)

    filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    return str(filepath)


def _fallback_csv(title, sheets):
    """Fallback when openpyxl is not installed."""
    import csv
    filename = f"{title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    filepath = OUTPUT_DIR / filename
    with open(filepath, "w", newline="") as f:
        for sheet_name, rows in sheets.items():
            if not rows:
                continue
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            f.write(f"# {sheet_name}\n")
            writer.writeheader()
            writer.writerows(rows)
            f.write("\n")
    return str(filepath)
