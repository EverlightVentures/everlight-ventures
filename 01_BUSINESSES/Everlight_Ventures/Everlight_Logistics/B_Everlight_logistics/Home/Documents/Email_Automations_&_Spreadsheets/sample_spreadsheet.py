import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set the headers
headers = ["Name", "Email", "Phone", "Membership", "start_date", 
           "billing_details", "email_automation", "receipt"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Define the email templates for the dropdown (data validation)
email_templates = [
    "Who are we",
    "Welcome",
    "Basic receipt",
    "Monthly motivational",
    "Anniversary",
    "Sample product 1",
    "Sample product 2",
    "Sample product 3"
]

# Create a data validation (dropdown) for the "Email Automation" column
dv = DataValidation(type="list", formula1='"' + ",".join(email_templates) + '"', showDropDown=True)
dv.add(ws["G2:G1048576"])  # Apply to column G from row 2 to the end
ws.add_data_validation(dv)

# Save the workbook to a file
file_path = "/mnt/data/client_email_automation.xlsx"
wb.save(file_path)

file_path
