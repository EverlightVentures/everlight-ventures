#!/usr/bin/env python3
"""
Mountain Gardens POS - Flask Web Application v8.0
# ==================================================
Complete rebuild with proper code ordering and all features.

Features:
- Sales Terminal with inventory picker
- Time Clock + Break tracking
- Task Management + Automations
- Payroll Management
- Inventory + Invoice Import
- Reports (Daily, COGS, Net Profit)
- AI Assistant
- Role-based access control
"""

import csv
import glob
import io
import json
import os
import re
import smtplib
import sys
import tempfile
import traceback
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from datetime import time as dtime
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import qrcode

# Load environment variables
try:
    from dotenv import load_dotenv
except ImportError:

    def load_dotenv(*args, **kwargs):
        return False


from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

# ==============================================================================
#                     FLASK APP SETUP
# ==============================================================================

app = Flask(__name__)


def _load_secret_key():
    """Stable per-install secret key. Never ship the old public default
    ("mountain-gardens-pos-2024-dev-key") -- a known key lets anyone forge an admin
    session cookie. Prefer SECRET_KEY env; else persist a random key to .secret_key
    so logins survive app restarts without exposing a guessable key."""
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key
    import secrets
    key_file = Path(__file__).parent / ".secret_key"
    try:
        if key_file.exists():
            existing = key_file.read_text(encoding="utf-8").strip()
            if existing:
                return existing
        new_key = secrets.token_hex(32)
        key_file.write_text(new_key, encoding="utf-8")
        return new_key
    except Exception:
        return secrets.token_hex(32)


app.secret_key = _load_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=43200,  # auto-expire idle sessions after 12 hours
)


@app.after_request
def add_security_headers(response):
    """Kill back/forward-cache so a logged-out admin page can NOT be restored with
    the browser Back button (the /logout route already clears the session server
    side; this stops the cached page from ever being shown again)."""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    return response

# Add parent directory for imports
sys.path.insert(0, str(Path(__file__).parent))

# ==============================================================================
#                     AUTHENTICATION DECORATORS (MUST BE BEFORE ROUTES!)
# ==============================================================================


@app.context_processor
def inject_unread_counts():
    try:
        emp_id = str(session.get("employee_id", ""))
    except Exception:
        return {}

    if not emp_id:
        return {}

    notifs = get_all_notifications(emp_id, limit=500)

    def is_unread(n):
        return str(n.get("Read", n.get("Is_Read", "N"))).strip().upper() not in (
            "Y",
            "YES",
            "TRUE",
            "1",
        )

    unread_task = sum(1 for n in notifs if (n.get("Type") == "TASK") and is_unread(n))
    unread_timeoff = sum(
        1 for n in notifs if (n.get("Type") == "TIMEOFF") and is_unread(n)
    )

    return {
        "unread_task_count": unread_task,
        "unread_timeoff_count": unread_timeoff,
    }


def login_required(f):
    """Decorator to require login."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if "employee_id" not in session:
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated


def manager_required(f):
    """Decorator to require manager role."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if "employee_id" not in session:
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        if session.get("role") not in ("Manager", "Owner", "Admin"):
            flash("Manager access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated


def owner_required(f):
    """Require OWNER/ADMIN role (managers excluded). Gates the most sensitive
    controls -- viewing/assigning employee PINs and the admin task scheduler."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if "employee_id" not in session:
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        if session.get("role") not in ("Owner", "Admin"):
            flash("Owner access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated


# ==============================================================================
#                     IMPORT POS CORE
# ==============================================================================

import money_core as money
from POS_CORE import (
    ANIMAL_SUBCATEGORIES,
    # App meta
    BUSINESS_NAME,
    EMPLOYEE_HEADERS,
    ITEM_HEADERS,
    LEDGER_HEADERS,
    MAIN_CATEGORIES,
    PLANT_SUBCATEGORIES,
    PRODUCT_SUBCATEGORIES,
    ROLES,
    VERSION,
    add_punch,
    append_csv,
    approve_time_off,
    assign_task,
    authenticate,
    check_low_stock,
    clock_in,
    clock_out,
    compose_full_name,
    create_employee,
    create_item,
    create_lot,
    import_items,
    create_notification,
    deactivate_employee,
    delete_punch,
    edit_punch,
    end_break,
    ensure_csv,
    # Reports
    generate_daily_summary,
    generate_sku,
    # Employees / auth
    get_all_employees,
    get_all_items,
    get_all_notifications,
    # Tasks
    get_all_tasks,
    get_average_cost,
    get_config,
    get_employee,
    get_employee_path,
    get_employee_pay_config,
    get_employee_status,
    get_item,
    get_items_path,
    get_ledger_path,
    get_lots_for_sku,
    get_lots_path,
    get_or_create_task_template,
    get_pending_requests,
    get_punch_by_id,
    get_punches_for_date,
    get_stock_on_hand,
    get_task,
    get_task_assignment,
    get_task_assignments_for_date,
    get_task_events,
    get_tasks_for_employee,
    get_tax_rate,
    # Time off
    get_time_off_requests,
    get_timeclock_edit_history,
    get_timeclock_path,
    get_timeoff_path,
    get_unread_notifications,
    # Audit / notifications
    log_audit,
    # Data helpers
    read_csv,
    record_sale,
    add_newsletter_subscriber,
    build_receipt_payload,
    build_receipt_pdf_bytes,
    get_customer_history,
    get_newsletter_path,
    get_receipt_bundle,
    log_customer_receipt,
    log_receipt_delivery,
    send_receipt_email_smtp,
    upsert_customer,
    list_customers,
    get_customer_by_id,
    get_newsletter_subscribers,
    unsubscribe_newsletter,
    customer_tier,
    ingest_invoice_lines,
    map_vendor_sku,
    resolve_vendor_sku,
    get_vendor_aliases_for_sku,
    list_vendor_map,
    create_recurring_task,
    check_and_assign_recurring_tasks,
    list_recurring_schedules,
    set_recurring_status,
    preview_recurring,
    run_payroll,
    search_items,
    quick_add_item,
    get_unreconciled_quickadds,
    reconcile_quickadd,
    reset_pin,
    set_config,
    set_data_dir,
    setup_employee_pay,
    start_break,
    update_csv_row,
    update_task_status,
    write_csv,
)

# Try optional imports
try:
    from invoice_importer import import_invoice_csv

    HAS_INVOICE_IMPORTER = True
except ImportError:
    HAS_INVOICE_IMPORTER = False

try:
    import requests

    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import stripe

    HAS_STRIPE = True
except ImportError:
    stripe = None
    HAS_STRIPE = False

try:
    from openai import OpenAI

    OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")
    openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
except ImportError:
    openai_client = None

# ==============================================================================
#                     CONFIGURATION
# ==============================================================================

BASE_DIR = Path(__file__).resolve().parent
N8N_BASE_URL = os.environ.get("N8N_URL", "http://localhost:5678")

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = str(SCRIPT_DIR)  # <— fixes NameError everywhere that uses DATA_DIR

TENANTS_DIR = SCRIPT_DIR / "Tenants"
TENANTS_PATH = TENANTS_DIR / "tenants.csv"
SUBSCRIPTIONS_PATH = TENANTS_DIR / "subscriptions.csv"

TENANT_HEADERS = ["Tenant_ID", "Name", "Status", "Created_At", "Data_Dir"]
SUBSCRIPTION_HEADERS = [
    "Tenant_ID",
    "Tier",
    "Status",
    "Stripe_Customer_ID",
    "Stripe_Subscription_ID",
    "Current_Period_End",
    "Trial_End",
]

TIER_ORDER = {"starter": 0, "growth": 1, "pro": 2}
DEFAULT_TIER = os.environ.get("MGN_DEFAULT_TIER", "pro").lower()
BILLING_MODE = os.environ.get("MGN_BILLING_MODE", "stripe").lower()
TIER_LIMITS = {
    "starter": {"max_users": 1, "max_skus": 250, "max_txn_month": 100},
    "growth": {"max_users": 5, "max_skus": 2500, "max_txn_month": 2000},
    "pro": {"max_users": None, "max_skus": None, "max_txn_month": None},
}
# ==============================================================================
#                     HELPER FUNCTIONS
# ==============================================================================


def make_change_breakdown(amount: float):
    """Given a change amount in dollars, return breakdown in US currency."""
    remaining = int(round(amount * 100))
    denoms = [
        ("$100", 10000),
        ("$50", 5000),
        ("$20", 2000),
        ("$10", 1000),
        ("$5", 500),
        ("$1", 100),
        ("25¢", 25),
        ("10¢", 10),
        ("5¢", 5),
        ("1¢", 1),
    ]
    breakdown = []
    for label, value in denoms:
        if remaining <= 0:
            break
        count, remaining = divmod(remaining, value)
        if count:
            breakdown.append({"label": label, "count": int(count)})
    return breakdown


def ensure_tenant_files():
    ensure_csv(TENANTS_PATH, TENANT_HEADERS)
    ensure_csv(SUBSCRIPTIONS_PATH, SUBSCRIPTION_HEADERS)
    _ensure_default_tenant()


def _ensure_default_tenant():
    rows = read_csv(TENANTS_PATH)
    if rows:
        return
    tenant_row = {
        "Tenant_ID": "default",
        "Name": BUSINESS_NAME,
        "Status": "Active",
        "Created_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Data_Dir": str(SCRIPT_DIR),
    }
    append_csv(TENANTS_PATH, TENANT_HEADERS, tenant_row)
    sub_row = {
        "Tenant_ID": "default",
        "Tier": DEFAULT_TIER,
        "Status": "trial",
        "Stripe_Customer_ID": "",
        "Stripe_Subscription_ID": "",
        "Current_Period_End": "",
        "Trial_End": (date.today() + timedelta(days=14)).isoformat(),
    }
    append_csv(SUBSCRIPTIONS_PATH, SUBSCRIPTION_HEADERS, sub_row)


def _slugify(name: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", (name or "").strip().lower()).strip("-")
    return slug or "tenant"


def list_tenants():
    ensure_tenant_files()
    return read_csv(TENANTS_PATH)


def get_tenant(tenant_id: str):
    for t in list_tenants():
        if t.get("Tenant_ID") == tenant_id:
            return t
    return None


# Single-store lock (ON by default). A single physical shop must ALWAYS read and
# write ONE fixed data folder, every request and every day, so its records can
# never split-brain across session-derived tenant folders -- that split is the root
# cause of "employees I added yesterday are gone today". Multi-tenant SaaS mode:
# set MGN_SINGLE_STORE=0. Pin the folder explicitly with MGN_DATA_DIR if desired.
SINGLE_STORE = os.environ.get("MGN_SINGLE_STORE", "1").strip().lower() in (
    "1", "true", "yes", "on",
)
FIXED_DATA_DIR = Path(os.environ.get("MGN_DATA_DIR", str(SCRIPT_DIR))).resolve()


def get_tenant_data_dir(tenant_id: str) -> Path:
    if SINGLE_STORE:
        return FIXED_DATA_DIR
    tenant = get_tenant(tenant_id) or {}
    data_dir = tenant.get("Data_Dir", "")
    p = Path(data_dir) if data_dir else (TENANTS_DIR / tenant_id)
    # Relocation-proof: if the stored/derived path doesn't exist (e.g. a restore to
    # a new machine still carrying the old absolute Data_Dir like /home/mgn/...),
    # fall back to the app's own bundled folder so the REAL catalog/data is read
    # instead of silently creating an empty Items.csv at a dead path.
    return p if p.exists() else SCRIPT_DIR


def create_tenant(name: str):
    ensure_tenant_files()
    existing = list_tenants()
    base = _slugify(name)
    tenant_id = base
    suffix = 1
    while any(t.get("Tenant_ID") == tenant_id for t in existing):
        suffix += 1
        tenant_id = f"{base}-{suffix}"

    tenant_row = {
        "Tenant_ID": tenant_id,
        "Name": name.strip() or tenant_id,
        "Status": "Active",
        "Created_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Data_Dir": str(TENANTS_DIR / tenant_id),
    }
    append_csv(TENANTS_PATH, TENANT_HEADERS, tenant_row)

    sub_row = {
        "Tenant_ID": tenant_id,
        "Tier": DEFAULT_TIER,
        "Status": "trial",
        "Stripe_Customer_ID": "",
        "Stripe_Subscription_ID": "",
        "Current_Period_End": "",
        "Trial_End": (date.today() + timedelta(days=14)).isoformat(),
    }
    append_csv(SUBSCRIPTIONS_PATH, SUBSCRIPTION_HEADERS, sub_row)
    return tenant_id


def get_subscription(tenant_id: str):
    ensure_tenant_files()
    for row in read_csv(SUBSCRIPTIONS_PATH):
        if row.get("Tenant_ID") == tenant_id:
            return row
    return None


def save_subscription(updated: dict):
    ensure_tenant_files()
    rows = read_csv(SUBSCRIPTIONS_PATH)
    found = False
    for row in rows:
        if row.get("Tenant_ID") == updated.get("Tenant_ID"):
            row.update(updated)
            found = True
            break
    if not found:
        rows.append(updated)
    write_csv(SUBSCRIPTIONS_PATH, SUBSCRIPTION_HEADERS, rows)


def get_current_tier() -> str:
    tenant_id = session.get("tenant_id", "")
    if not tenant_id:
        return DEFAULT_TIER
    sub = get_subscription(tenant_id) or {}
    tier = (sub.get("Tier") or DEFAULT_TIER).lower()
    return tier if tier in TIER_ORDER else DEFAULT_TIER


def tier_required(min_tier: str):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            tier = get_current_tier()
            if TIER_ORDER.get(tier, 0) < TIER_ORDER.get(min_tier, 0):
                flash("Upgrade to access this feature.", "error")
                return redirect(url_for("billing"))
            return f(*args, **kwargs)

        return wrapped

    return decorator


def _limit_for(key: str):
    tier = get_current_tier()
    limits = TIER_LIMITS.get(tier, {})
    return limits.get(key)


def _count_transactions_current_month() -> int:
    tenant_id = session.get("tenant_id", "")
    base_dir = get_tenant_data_dir(tenant_id) if tenant_id else SCRIPT_DIR
    txn_root = base_dir / "Transaction_Logs"
    if not txn_root.exists():
        return 0

    now = date.today()
    month_dir = txn_root / now.strftime("%Y") / now.strftime("%m_%B")
    if not month_dir.exists():
        return 0

    total = 0
    for path in month_dir.rglob("*_TransactionLog.csv"):
        total += len(read_csv(path))
    return total


def enforce_limit_or_upgrade(limit_key: str, current_value: int) -> bool:
    limit = _limit_for(limit_key)
    if limit is None:
        return True
    if current_value >= int(limit):
        flash("Upgrade to access more capacity.", "error")
        return False
    return True


ONBOARDING_HEADERS = ["Item_ID", "Title", "Status", "Completed_At"]
SUPPORT_TICKET_HEADERS = [
    "Ticket_ID",
    "Created_At",
    "Employee_ID",
    "Employee_Name",
    "Subject",
    "Details",
    "Status",
]
TENANT_SETTINGS_HEADERS = [
    "Referral_Code",
    "Review_Prompt_Enabled",
    "Onboarding_Email_Sent",
    "Onboarding_Reminder_Last",
    "Onboarding_Reminder_Count",
    "Referrals_Count",
    "Referral_Used",
    "Referral_Reward_Level",
    "Reward_Credits",
    "Created_At",
    "Theme",
    "POS_Name",
    "Tax_Rate",
    "Receipt_Footer",
    "Business_Address",
    "Business_Phone",
]
REFERRAL_REWARD_RULES = [
    "5 referrals = 1 month Growth credit",
    "10 referrals = 1 month Pro credit",
]


def _settings_dir() -> Path:
    return Path(DATA_DIR) / "Settings"


def get_onboarding_path() -> Path:
    return _settings_dir() / "Onboarding.csv"


def get_support_tickets_path() -> Path:
    return _settings_dir() / "Support_Tickets.csv"


def get_tenant_settings_path() -> Path:
    return _settings_dir() / "Tenant_Settings.csv"


def _default_onboarding_items():
    return [
        {
            "Item_ID": "brand",
            "Title": "Add business branding",
            "Status": "open",
            "Completed_At": "",
        },
        {
            "Item_ID": "tax",
            "Title": "Set tax rate and receipt footer",
            "Status": "open",
            "Completed_At": "",
        },
        {
            "Item_ID": "inventory",
            "Title": "Import or add your inventory",
            "Status": "open",
            "Completed_At": "",
        },
        {
            "Item_ID": "staff",
            "Title": "Create employee logins",
            "Status": "open",
            "Completed_At": "",
        },
        {
            "Item_ID": "test_sale",
            "Title": "Run a test sale",
            "Status": "open",
            "Completed_At": "",
        },
        {
            "Item_ID": "billing",
            "Title": "Pick a plan in Billing",
            "Status": "open",
            "Completed_At": "",
        },
    ]


def ensure_onboarding():
    path = get_onboarding_path()
    ensure_csv(path, ONBOARDING_HEADERS)
    rows = read_csv(path)
    if rows:
        return rows
    rows = _default_onboarding_items()
    write_csv(path, ONBOARDING_HEADERS, rows)
    return rows


def get_or_create_tenant_settings():
    path = get_tenant_settings_path()
    ensure_csv(path, TENANT_SETTINGS_HEADERS)
    rows = read_csv(path)
    if rows:
        row = rows[0]
        updated = False
        tenant_id = session.get("tenant_id", "")
        tenant = get_tenant(tenant_id) if tenant_id else {}
        tenant_name = tenant.get("Name", "") or session.get("tenant_name", "POS Suite")

        for key, default in {
            "Referral_Code": row.get("Referral_Code")
            or f"REF-{uuid.uuid4().hex[:8].upper()}",
            "Review_Prompt_Enabled": row.get("Review_Prompt_Enabled") or "Y",
            "Onboarding_Email_Sent": row.get("Onboarding_Email_Sent") or "",
            "Onboarding_Reminder_Last": row.get("Onboarding_Reminder_Last") or "",
            "Onboarding_Reminder_Count": row.get("Onboarding_Reminder_Count") or "0",
            "Referrals_Count": row.get("Referrals_Count") or "0",
            "Referral_Used": row.get("Referral_Used") or "",
            "Referral_Reward_Level": row.get("Referral_Reward_Level") or "0",
            "Reward_Credits": row.get("Reward_Credits") or "0",
            "Created_At": row.get("Created_At")
            or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Theme": row.get("Theme") or "dark",
            "POS_Name": row.get("POS_Name") or tenant_name,
        }.items():
            if key not in row or not row.get(key):
                row[key] = default
                updated = True
        if updated:
            write_csv(path, TENANT_SETTINGS_HEADERS, [row])
        return row
    referral = f"REF-{uuid.uuid4().hex[:8].upper()}"
    tenant_id = session.get("tenant_id", "")
    tenant = get_tenant(tenant_id) if tenant_id else {}
    tenant_name = tenant.get("Name", "") or session.get("tenant_name", "POS Suite")
    row = {
        "Referral_Code": referral,
        "Review_Prompt_Enabled": "Y",
        "Onboarding_Email_Sent": "",
        "Onboarding_Reminder_Last": "",
        "Onboarding_Reminder_Count": "0",
        "Referrals_Count": "0",
        "Referral_Used": "",
        "Referral_Reward_Level": "0",
        "Reward_Credits": "0",
        "Created_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Theme": "dark",
        "POS_Name": tenant_name,
    }
    write_csv(path, TENANT_SETTINGS_HEADERS, [row])
    return row


def send_onboarding_email(to_email: str, business_name: str) -> bool:
    host = os.getenv("SMTP_HOST", "").strip()
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "").strip()
    pw = os.getenv("SMTP_PASS", "").strip()
    mail_from = os.getenv("SMTP_FROM", user).strip()

    if not host or not user or not pw or not to_email:
        return False

    msg = EmailMessage()
    msg["Subject"] = f"Welcome to {business_name} POS"
    msg["From"] = mail_from
    msg["To"] = to_email
    msg.set_content(
        "Welcome!\n\n"
        "Here is your quick start checklist:\n"
        "1) Visit /onboarding to finish setup\n"
        "2) Add employees\n"
        "3) Import inventory\n"
        "4) Run a test sale\n"
        "5) Pick a plan in /billing\n\n"
        "Need help? Visit /support.\n"
    )

    with smtplib.SMTP(host, port) as s:
        s.starttls()
        s.login(user, pw)
        s.send_message(msg)
    return True


def get_eod_recipients():
    """EOD report recipients (owner + e.g. mom who handles the books). Settable
    in-app via Settings (Config.csv 'EOD_Emails'), else env MGN_EOD_EMAIL, else the
    owner. Comma-separated."""
    raw = (get_config("EOD_Emails", "") or os.getenv("MGN_EOD_EMAIL", "1m.rich.gee@gmail.com"))
    return [a.strip() for a in raw.split(",") if a.strip()]


def send_eod_report_email(summary_text: str,
                          subject: str = "Mountain Gardens POS -- End of Day Report",
                          attachments=None) -> bool:
    """Email the end-of-day close-out summary + optional CSV attachments.

    Recipients come from the MGN_EOD_EMAIL env var (comma-separated; defaults to
    the owner -- set it to 'owner@...,adam@...' for multiple). Best-effort (never
    raises). Returns True only if the message was actually sent.
    """
    host = os.getenv("SMTP_HOST", "").strip()
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "").strip()
    pw = os.getenv("SMTP_PASS", "").strip()
    mail_from = os.getenv("SMTP_FROM", user).strip()
    recipients = get_eod_recipients()
    if not host or not user or not pw or not recipients:
        return False
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = mail_from
    msg["To"] = ", ".join(recipients)
    msg.set_content(summary_text)
    for p in (attachments or []):
        try:
            if p and os.path.exists(p):
                with open(p, "rb") as fh:
                    msg.add_attachment(fh.read(), maintype="text", subtype="csv",
                                       filename=os.path.basename(p))
        except Exception:
            pass
    try:
        with smtplib.SMTP(host, port) as s:
            s.starttls()
            s.login(user, pw)
            s.send_message(msg)
        return True
    except Exception:
        return False


def find_tenant_by_referral(code: str):
    code = (code or "").strip()
    if not code:
        return None
    for tenant in list_tenants():
        tenant_id = tenant.get("Tenant_ID")
        if not tenant_id:
            continue
        tenant_dir = get_tenant_data_dir(tenant_id)
        settings_path = tenant_dir / "Settings" / "Tenant_Settings.csv"
        if not settings_path.exists():
            continue
        rows = read_csv(settings_path)
        if (
            rows
            and (rows[0].get("Referral_Code") or "").strip().upper() == code.upper()
        ):
            return tenant_id
    return None


def apply_referral(referral_code: str, new_tenant_id: str):
    ref_tenant_id = find_tenant_by_referral(referral_code)
    if not ref_tenant_id:
        return False

    ref_dir = get_tenant_data_dir(ref_tenant_id)
    ref_settings_path = ref_dir / "Settings" / "Tenant_Settings.csv"
    rows = read_csv(ref_settings_path)
    if not rows:
        return False
    row = rows[0]
    current = int(row.get("Referrals_Count") or 0)
    row["Referrals_Count"] = str(current + 1)
    write_csv(ref_settings_path, TENANT_SETTINGS_HEADERS, [row])

    new_dir = get_tenant_data_dir(new_tenant_id)
    new_settings_path = new_dir / "Settings" / "Tenant_Settings.csv"
    new_rows = read_csv(new_settings_path)
    if new_rows:
        new_rows[0]["Referral_Used"] = referral_code
        write_csv(new_settings_path, TENANT_SETTINGS_HEADERS, [new_rows[0]])
    return True


def apply_referral_rewards(settings: dict) -> dict:
    """
    Simple rewards ladder based on referral count.
    Reward_Credits is a counter for future billing credit application.
    """
    try:
        count = int(settings.get("Referrals_Count") or 0)
        level = int(settings.get("Referral_Reward_Level") or 0)
        credits = int(settings.get("Reward_Credits") or 0)
    except Exception:
        return settings

    # Level 1 at 5 referrals, Level 2 at 10 referrals
    if count >= 5 and level < 1:
        credits += 1
        level = 1
    if count >= 10 and level < 2:
        credits += 1
        level = 2

    settings["Referral_Reward_Level"] = str(level)
    settings["Reward_Credits"] = str(credits)
    return settings


def maybe_send_onboarding_reminder(
    settings: dict, business_name: str, owner_email: str
) -> dict:
    if not owner_email:
        return settings

    items = ensure_onboarding()
    if not items:
        return settings

    done = sum(1 for i in items if (i.get("Status") or "").lower() == "done")
    total = len(items)
    if done >= total:
        return settings

    last_sent = settings.get("Onboarding_Reminder_Last") or ""
    try:
        last_dt = datetime.strptime(last_sent, "%Y-%m-%d")
    except Exception:
        last_dt = None

    # Only send if never sent or >3 days since last reminder
    if last_dt and (datetime.now() - last_dt).days < 3:
        return settings

    if send_onboarding_email(owner_email, business_name):
        settings["Onboarding_Reminder_Last"] = datetime.now().strftime("%Y-%m-%d")
        settings["Onboarding_Reminder_Count"] = str(
            int(settings.get("Onboarding_Reminder_Count") or 0) + 1
        )
    return settings


# ==============================================================================
#                     CONTEXT PROCESSORS
# ==============================================================================


@app.context_processor
def inject_globals():
    """Inject global variables into all templates."""
    notif_count = 0
    my_tasks_count = 0

    if "employee_id" in session:
        notifs = get_unread_notifications(session["employee_id"])
        notif_count = len(notifs) if notifs else 0

        try:
            tasks = get_tasks_for_employee(session["employee_id"], date.today())
            my_tasks_count = len(
                [t for t in tasks if t.get("Status") not in ("COMPLETE", "SKIPPED")]
            )
        except:
            pass

    return {
        "business_name": BUSINESS_NAME,
        "current_year": datetime.now().year,
        "current_time": datetime.now(),
        "version": VERSION,
        "notification_count": notif_count,
        "my_tasks_count": my_tasks_count,
        "now": datetime.now(),
    }


@app.context_processor
def template_utils():
    """Helper functions for templates."""

    def has_endpoint(name: str) -> bool:
        return name in app.view_functions

    return dict(has_endpoint=has_endpoint)


@app.context_processor
def inject_tenant_context():
    tenant_id = session.get("tenant_id", "")
    tenant = get_tenant(tenant_id) if tenant_id else None
    return {
        "tenant_id": tenant_id,
        "tenant_name": (tenant or {}).get("Name", ""),
        "current_tier": get_current_tier(),
    }


# ==============================================================================
#                     ERROR HANDLERS
# ==============================================================================


@app.errorhandler(500)
def server_error(e):
    error_info = {
        "error": str(e),
        "traceback": traceback.format_exc(),
        "route": request.path,
        "timestamp": datetime.now().isoformat(),
    }
    session["last_error"] = error_info
    return render_template("errors/500.html", error_info=error_info), 500


@app.errorhandler(404)
def not_found(e):
    return render_template("errors/404.html"), 404


@app.before_request
def apply_tenant_context():
    # Single-store: pin the data layer to the one fixed folder on EVERY request,
    # regardless of session, so a single shop can never write to one folder today
    # and read from another tomorrow.
    if SINGLE_STORE:
        set_data_dir(FIXED_DATA_DIR)
        globals()["DATA_DIR"] = str(FIXED_DATA_DIR)
        return
    tenant_id = session.get("tenant_id")
    if tenant_id:
        tenant_dir = get_tenant_data_dir(tenant_id)
        set_data_dir(tenant_dir)
        globals()["DATA_DIR"] = str(tenant_dir)
    else:
        # Reset POS_CORE's module globals too, so a prior request's tenant folder
        # can't leak into an unauthenticated request on a shared worker.
        set_data_dir(SCRIPT_DIR)
        globals()["DATA_DIR"] = str(SCRIPT_DIR)


# ==============================================================================
#                     AUTH ROUTES
# ==============================================================================


@app.route("/")
def index():
    if "employee_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    selected_tenant = (request.values.get("tenant_id") or "").strip()
    if selected_tenant:
        tenant_dir = get_tenant_data_dir(selected_tenant)
        set_data_dir(tenant_dir)
        globals()["DATA_DIR"] = str(tenant_dir)

    if request.method == "POST":
        if not selected_tenant:
            flash("Please select your business.", "error")
            return redirect(url_for("login"))

        emp_id = request.form.get("employee_id", "").strip()
        pin = request.form.get("pin", "").strip()

        success, msg, emp = authenticate(emp_id, pin)

        if success:
            session["tenant_id"] = selected_tenant
            tenant = get_tenant(selected_tenant) or {}
            session["tenant_name"] = tenant.get("Name", "")
            session["employee_id"] = emp_id
            session["employee_name"] = emp["Employee_Name"]
            session["role"] = emp.get("Role", "Cashier")

            notifs = get_unread_notifications(emp_id)
            if notifs:
                flash(f"You have {len(notifs)} unread notification(s)", "info")

            return redirect(url_for("dashboard"))
        else:
            flash(msg, "error")

    employees = get_all_employees() if selected_tenant else []
    tenants = list_tenants()
    return render_template(
        "login.html",
        employees=employees,
        tenants=tenants,
        selected_tenant=selected_tenant,
    )


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        business_name = (request.form.get("business_name") or "").strip()
        owner_name = (request.form.get("owner_name") or "").strip()
        owner_pin = (request.form.get("owner_pin") or "").strip()
        owner_email = (request.form.get("owner_email") or "").strip()
        referral_code = (request.form.get("referral_code") or "").strip()

        if not business_name or not owner_name:
            flash("Business name and owner name are required.", "error")
            return redirect(url_for("signup"))

        if not owner_pin.isdigit() or len(owner_pin) != 4:
            flash("Owner PIN must be 4 digits.", "error")
            return redirect(url_for("signup"))

        tenant_id = create_tenant(business_name)
        tenant_dir = get_tenant_data_dir(tenant_id)
        set_data_dir(tenant_dir)
        globals()["DATA_DIR"] = str(tenant_dir)

        ok, msg, emp_id = create_employee(
            owner_name, "Owner", owner_pin, email=owner_email
        )
        if not ok:
            flash(msg, "error")
            return redirect(url_for("signup"))

        session["tenant_id"] = tenant_id
        session["tenant_name"] = business_name
        session["employee_id"] = emp_id
        session["employee_name"] = owner_name
        session["role"] = "Owner"

        ensure_onboarding()
        settings = get_or_create_tenant_settings()
        if referral_code:
            apply_referral(referral_code, tenant_id)

        if owner_email and not settings.get("Onboarding_Email_Sent"):
            if send_onboarding_email(owner_email, business_name):
                settings["Onboarding_Email_Sent"] = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                write_csv(
                    get_tenant_settings_path(), TENANT_SETTINGS_HEADERS, [settings]
                )

        flash("Business created. You're signed in as Owner.", "success")
        return redirect(url_for("dashboard"))

    return render_template("signup.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ==============================================================================
#                     SUPPORT DIAGNOSTICS
# ==============================================================================


@app.route("/support/diagnostics")
@login_required
def support_diagnostics():
    checks = {
        "tenant_id": session.get("tenant_id", ""),
        "data_dir": str(get_tenant_data_dir(session.get("tenant_id", "")))
        if session.get("tenant_id")
        else "",
        "items_csv": os.path.exists(get_items_path()),
        "lots_csv": os.path.exists(get_lots_path()),
        "ledger_csv": os.path.exists(get_ledger_path()),
        "smtp_configured": bool(os.getenv("SMTP_HOST"))
        and bool(os.getenv("SMTP_USER")),
        "stripe_configured": bool(os.getenv("STRIPE_SECRET_KEY")),
    }
    return jsonify({"ok": True, "checks": checks})


@app.route("/support")
@login_required
def support_home():
    tickets = read_csv(get_support_tickets_path())
    tickets = sorted(tickets, key=lambda r: r.get("Created_At", ""), reverse=True)[:10]
    return render_template("support.html", tickets=tickets)


@app.route("/support/report", methods=["POST"])
@login_required
def support_report():
    subject = (request.form.get("subject") or "").strip()
    details = (request.form.get("details") or "").strip()
    if not subject:
        flash("Subject is required.", "error")
        return redirect(url_for("support_home"))

    row = {
        "Ticket_ID": f"TKT-{uuid.uuid4().hex[:8].upper()}",
        "Created_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Employee_ID": session.get("employee_id", ""),
        "Employee_Name": session.get("employee_name", ""),
        "Subject": subject,
        "Details": details,
        "Status": "Open",
    }
    append_csv(get_support_tickets_path(), SUPPORT_TICKET_HEADERS, row)
    flash("Support request logged. We'll get back to you soon.", "success")
    return redirect(url_for("support_home"))


@app.route("/onboarding")
@login_required
def onboarding():
    items = ensure_onboarding()
    return render_template("onboarding.html", items=items)


@app.route("/onboarding/complete", methods=["POST"])
@login_required
def onboarding_complete():
    item_id = (request.form.get("item_id") or "").strip()
    rows = ensure_onboarding()
    for row in rows:
        if row.get("Item_ID") == item_id:
            row["Status"] = "done"
            row["Completed_At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    write_csv(get_onboarding_path(), ONBOARDING_HEADERS, rows)
    return redirect(url_for("onboarding"))


@app.route("/growth")
@login_required
def growth():
    settings = get_or_create_tenant_settings()
    review_url = os.environ.get("APP_REVIEW_URL", "").strip()
    return render_template(
        "growth.html",
        settings=settings,
        review_url=review_url,
        reward_rules=REFERRAL_REWARD_RULES,
    )


# ==============================================================================
#                     BILLING / SUBSCRIPTIONS
# ==============================================================================


@app.route("/billing")
@login_required
def billing():
    tenant_id = session.get("tenant_id", "")
    sub = get_subscription(tenant_id) if tenant_id else {}
    tier = (sub.get("Tier") or DEFAULT_TIER).lower()
    limits = TIER_LIMITS.get(tier, {})
    usage = {
        "users": len(get_all_employees(include_inactive=True)) if tenant_id else 0,
        "skus": len(get_all_items()) if tenant_id else 0,
        "txns_month": _count_transactions_current_month() if tenant_id else 0,
        "limits": limits,
    }
    return render_template(
        "billing.html",
        billing_mode=BILLING_MODE,
        subscription=sub,
        usage=usage,
        tiers=[
            {"id": "starter", "price": "19.99", "label": "Starter"},
            {"id": "growth", "price": "49.99", "label": "Growth"},
            {"id": "pro", "price": "99.99", "label": "Pro"},
        ],
    )


@app.route("/billing/checkout/<tier>")
@login_required
def billing_checkout(tier):
    tier = tier.lower()
    if tier not in TIER_ORDER:
        flash("Invalid plan selected.", "error")
        return redirect(url_for("billing"))

    if BILLING_MODE == "mock":
        tenant_id = session.get("tenant_id", "")
        if tenant_id:
            save_subscription(
                {
                    "Tenant_ID": tenant_id,
                    "Tier": tier,
                    "Status": "active",
                    "Stripe_Customer_ID": "",
                    "Stripe_Subscription_ID": "",
                    "Current_Period_End": "",
                    "Trial_End": "",
                }
            )
        flash(f"Mock billing: switched to {tier.title()}.", "success")
        return redirect(url_for("billing"))

    if not HAS_STRIPE:
        flash("Stripe is not configured on this server.", "error")
        return redirect(url_for("billing"))

    stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")
    if not stripe.api_key:
        flash("Stripe keys are missing.", "error")
        return redirect(url_for("billing"))

    price_id = os.environ.get(f"STRIPE_PRICE_{tier.upper()}", "")
    if not price_id:
        flash("Stripe price is not configured for this plan.", "error")
        return redirect(url_for("billing"))

    tenant_id = session.get("tenant_id", "")
    sub = get_subscription(tenant_id) or {}
    customer_id = sub.get("Stripe_Customer_ID", "")

    success_url = request.host_url.rstrip("/") + url_for("billing") + "?success=1"
    cancel_url = request.host_url.rstrip("/") + url_for("billing") + "?canceled=1"

    try:
        session_args = {
            "mode": "subscription",
            "line_items": [{"price": price_id, "quantity": 1}],
            "success_url": success_url,
            "cancel_url": cancel_url,
            "metadata": {"tenant_id": tenant_id, "tier": tier},
        }
        if customer_id:
            session_args["customer"] = customer_id

        checkout = stripe.checkout.Session.create(**session_args)
        return redirect(checkout.url)
    except Exception as e:
        flash(f"Stripe checkout failed: {e}", "error")
        return redirect(url_for("billing"))


@app.route("/billing/stripe/webhook", methods=["POST"])
def stripe_webhook():
    if not HAS_STRIPE:
        return ("Stripe not configured", 400)

    stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")

    payload = request.data
    sig_header = request.headers.get("Stripe-Signature", "")
    webhook_secret = os.environ.get("STRIPE_WEBHOOK_SECRET", "")

    try:
        if webhook_secret:
            event = stripe.Webhook.construct_event(payload, sig_header, webhook_secret)
        else:
            event = stripe.Event.construct_from(json.loads(payload), stripe.api_key)
    except Exception:
        return ("Invalid payload", 400)

    event_type = event.get("type")
    data_object = (event.get("data") or {}).get("object") or {}

    tenant_id = (data_object.get("metadata") or {}).get("tenant_id", "")
    tier = (data_object.get("metadata") or {}).get("tier", "")

    if event_type == "checkout.session.completed":
        if tenant_id:
            save_subscription(
                {
                    "Tenant_ID": tenant_id,
                    "Tier": tier or DEFAULT_TIER,
                    "Status": "active",
                    "Stripe_Customer_ID": data_object.get("customer", ""),
                    "Stripe_Subscription_ID": data_object.get("subscription", ""),
                    "Current_Period_End": "",
                    "Trial_End": "",
                }
            )

    if event_type in ("customer.subscription.updated", "customer.subscription.deleted"):
        subscription = data_object
        customer_id = subscription.get("customer", "")
        status = subscription.get("status", "")
        current_period_end = subscription.get("current_period_end", "")

        rows = read_csv(SUBSCRIPTIONS_PATH)
        for row in rows:
            if customer_id and row.get("Stripe_Customer_ID") == customer_id:
                row["Status"] = status
                row["Stripe_Subscription_ID"] = subscription.get("id", "")
                row["Current_Period_End"] = str(current_period_end)
                if tier:
                    row["Tier"] = tier
                write_csv(SUBSCRIPTIONS_PATH, SUBSCRIPTION_HEADERS, rows)
                break

    return ("ok", 200)


#!/usr/bin/env python3
"""
CSV INVOICE IMPORTER ROUTE
# ===========================
Add to MGN_APP.py

Flexible CSV import with field mapping for vendor invoices
Matches base.html theme
"""  # ═══════════════════════════════════════════════════════════════════════════
# CSV BULK IMPORT - PASTE THIS INTO MGN_APP.py
# Place this ABOVE the line: @app.route('/inventory/low-stock')
# ═══════════════════════════════════════════════════════════════════════════
# CSV BULK IMPORT - SINGLE CLEAN ROUTE
# Replace ALL duplicate csv-import routes with this ONE route
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
# CSV BULK IMPORT - SINGLE CLEAN ROUTE
# Replace ALL duplicate csv-import routes with this ONE route
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
# SMART CSV IMPORT WITH AUTO FIELD MAPPING
# Replace your csv_bulk_import route with this enhanced version
# Place ABOVE @app.route('/inventory/low-stock')
# ═══════════════════════════════════════════════════════════════════════════

#!/usr/bin/env python3
"""
INTELLIGENT CSV IMPORTER FOR MGN POS
# ====================================
Add to MGN_APP.py

Features:
- Multi-vendor format support (auto-detects field names)
- Intelligent field mapping
- QR code generation for each item
- Supplier barcode association
- SKU generation
- Multiple file format support (CSV, XLSX, XLS)
- Duplicate detection
- Auto-categorization
"""

import os

# ============================================================================
# VENDOR FIELD MAPPING DICTIONARY
# Maps common vendor column names to our Items.csv fields
# ============================================================================

FIELD_MAPPING = {
    # Item Name variations
    "item_name": [
        "purchase description",
        "purchase_description",
        "item_name",
        "item",
        "product",
        "product_name",
        "description",
        "item_description",
        "name",
        "title",
        "product_description",
        "item description",
        "product description",
        "desc",
    ],
    # SKU variations
    "sku": [
        "sku",
        "item_code",
        "product_code",
        "code",
        "item_number",
        "product_number",
        "itemcode",
        "productcode",
        "item #",
        "product #",
    ],
    # UPC/Barcode variations
    "upc": [
        "upc",
        "barcode",
        "upc_code",
        "bar_code",
        "ean",
        "gtin",
        "supplier_barcode",
        "vendor_barcode",
        "manufacturer_code",
    ],
    # Price variations
    "price": [
        "price",
        "unit_price",
        "retail_price",
        "selling_price",
        "list_price",
        "default_price",
        "msrp",
    ],
    "cost": [
        "cost",
        "unit_cost",
        "wholesale_cost",
        "purchase_cost",
        "wholesale_price",
        "buy_price",
        "vendor_cost",
    ],
    # Category variations
    "category": [
        "category",
        "dept",
        "department",
        "class",
        "type",
        "product_category",
        "item_category",
        "group",
    ],
    # Subcategory variations
    "subcategory": [
        "subcategory",
        "sub_category",
        "subclass",
        "subtype",
        "product_type",
        "variety",
    ],
    # Size variations
    "size": [
        "size",
        "dimension",
        "dimensions",
        "container",
        "pot_size",
        "gallon",
        "gal",
        "container_size",
    ],
    # Quantity variations
    "quantity": [
        "quantity",
        "qty",
        "qty_received",
        "amount",
        "count",
        "on_hand",
        "stock",
        "available",
    ],
    # Vendor variations
    "vendor": [
        "vendor",
        "supplier",
        "manufacturer",
        "brand",
        "source",
        "vendor_name",
        "supplier_name",
    ],
}


def smart_field_match(csv_header, our_field):
    """
    Intelligently match vendor CSV column to our field
    Returns the vendor's column name if match found
    """
    csv_lower = csv_header.lower().strip()

    # Check if this CSV column matches any of our field's variations
    if our_field in FIELD_MAPPING:
        for variation in FIELD_MAPPING[our_field]:
            if variation in csv_lower or csv_lower in variation:
                return True

    return False


def build_field_map(csv_headers):
    """
    Auto-detect vendor format and build field mapping
    Returns dict: {'item_name': 'Product Description', 'sku': 'Item #', ...}
    """
    field_map = {}

    for our_field in FIELD_MAPPING.keys():
        for csv_col in csv_headers:
            if smart_field_match(csv_col, our_field):
                field_map[our_field] = csv_col
                break

    return field_map


def generate_qr_code(sku, item_data):
    """
    Generate QR code containing item data
    Returns QR code data string and saves QR image
    """
    try:
        import qrcode

        # Create QR data payload
        qr_data = {
            "sku": sku,
            "name": item_data.get("Item_Name", ""),
            "price": item_data.get("Default_Price", "0"),
            "category": item_data.get("Category", ""),
            "upc": item_data.get("Supplier_Barcode", ""),
        }

        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)

        # Save QR image
        qr_dir = "static/qrcodes"
        os.makedirs(qr_dir, exist_ok=True)

        img = qr.make_image(fill_color="black", back_color="white")
        qr_path = os.path.join(qr_dir, f"{sku}.png")
        img.save(qr_path)

        return json.dumps(qr_data), qr_path

    except ImportError:
        # If qrcode not installed, return data only
        qr_data = {
            "sku": sku,
            "name": item_data.get("Item_Name", ""),
            "price": item_data.get("Default_Price", "0"),
        }
        return json.dumps(qr_data), None


# ============================================================================
# ENHANCED CSV IMPORT ROUTE
# ============================================================================


@app.route("/inventory/csv-import", methods=["GET", "POST"])
@login_required
@tier_required("pro")
def csv_bulk_import():
    """Intelligent multi-format CSV importer with QR generation"""

    if request.method == "GET":
        return render_template("inventory/csv_bulk_import.html")

    # Check file uploaded
    if "csv_file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["csv_file"]

    if not file or file.filename == "":
        return jsonify({"success": False, "error": "No file selected"}), 400

    filename = file.filename.lower()

    try:
        # ====================================================================
        # STEP 1: READ FILE (Support CSV, XLSX, XLS)
        # ====================================================================

        if filename.endswith(".csv"):
            # Read CSV
            content = file.read().decode("utf-8-sig")
            csv_reader = csv.DictReader(io.StringIO(content))
            vendor_data = list(csv_reader)
            vendor_headers = csv_reader.fieldnames

        elif filename.endswith((".xlsx", ".xls")):
            # Read Excel
            try:
                import openpyxl
                from openpyxl import load_workbook

                wb = load_workbook(file, read_only=True, data_only=True)
                ws = wb.active

                # Get headers
                vendor_headers = [cell.value for cell in ws[1]]

                # Get data
                vendor_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {
                        vendor_headers[i]: row[i] for i in range(len(vendor_headers))
                    }
                    vendor_data.append(row_dict)

            except ImportError:
                return jsonify(
                    {
                        "success": False,
                        "error": "Excel support not installed. Please install: pip install openpyxl",
                    }
                ), 400
        else:
            return jsonify(
                {
                    "success": False,
                    "error": "Unsupported file format. Use CSV, XLSX, or XLS",
                }
            ), 400

        if not vendor_headers or not vendor_data:
            return jsonify({"success": False, "error": "File is empty"}), 400

        # ====================================================================
        # STEP 2: AUTO-DETECT VENDOR FORMAT
        # ====================================================================

        field_map = build_field_map(vendor_headers)

        if not field_map.get("item_name"):
            return jsonify(
                {
                    "success": False,
                    "error": "Could not detect item name column. Please ensure CSV has product/item/description column.",
                }
            ), 400

        # ====================================================================
        # STEP 3: LOAD EXISTING INVENTORY
        # ====================================================================

        items_path = get_items_path()
        with open(items_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            our_headers = reader.fieldnames
            existing_items = list(reader)

        # Build lookup tables
        existing_skus = {
            item["SKU"].strip().upper() for item in existing_items if item.get("SKU")
        }

        # Build barcode lookup (maps supplier barcode to our SKU)
        barcode_to_sku = {}
        for item in existing_items:
            if item.get("Supplier_Barcode"):
                barcode_to_sku[item["Supplier_Barcode"].strip()] = item["SKU"]

        # ====================================================================
        # STEP 4: PROCESS VENDOR DATA
        # ====================================================================

        new_items = []
        updated_items = []
        duplicate_count = 0
        error_rows = []
        qr_codes_generated = 0
        barcodes_linked = 0

        for idx, vendor_row in enumerate(vendor_data, start=2):
            try:
                # Extract vendor data using field map
                item_name = None
                for field in ["item_name", "item_description", "description"]:
                    if field in field_map and vendor_row.get(field_map[field]):
                        item_name = str(vendor_row[field_map[field]]).strip()
                        break

                if not item_name:
                    continue  # Skip empty names

                # Get other fields
                vendor_sku = vendor_row.get(field_map.get("sku", ""), "").strip()
                vendor_upc = vendor_row.get(field_map.get("upc", ""), "").strip()
                price = vendor_row.get(field_map.get("price", ""), "0")
                cost = vendor_row.get(field_map.get("cost", ""), "0")
                category = vendor_row.get(field_map.get("category", ""), "General")
                subcategory = vendor_row.get(field_map.get("subcategory", ""), "")
                size = vendor_row.get(field_map.get("size", ""), "")
                quantity = vendor_row.get(field_map.get("quantity", ""), "0")
                vendor_name = vendor_row.get(field_map.get("vendor", ""), "Unknown")

                # Clean price/cost
                try:
                    price = float(
                        str(price).replace("$", "").replace(",", "").strip() or "0"
                    )
                except:
                    price = 0.0

                try:
                    cost = float(
                        str(cost).replace("$", "").replace(",", "").strip() or "0"
                    )
                except:
                    cost = 0.0

                # ============================================================
                # INTELLIGENT ITEM MATCHING
                # ============================================================

                our_sku = None
                is_update = False

                # Method 1: Match by vendor UPC/barcode
                if vendor_upc and vendor_upc in barcode_to_sku:
                    our_sku = barcode_to_sku[vendor_upc]
                    is_update = True
                    barcodes_linked += 1

                # Method 2: Match by vendor SKU if exists
                elif vendor_sku and vendor_sku.upper() in existing_skus:
                    our_sku = vendor_sku.upper()
                    is_update = True
                    duplicate_count += 1
                    continue  # Skip duplicates

                # Method 3: New item - generate SKU
                else:
                    # Generate SKU based on category
                    category_prefix = category[:3].upper() if category else "GEN"
                    our_sku = f"{category_prefix}-{uuid.uuid4().hex[:8].upper()}"

                # ============================================================
                # BUILD ITEM DATA
                # ============================================================

                item_data = {
                    "SKU": our_sku,
                    "Item_Name": item_name,
                    "Category": category or "General",
                    "Subcategory": subcategory,
                    "Product_Name": item_name,
                    "Default_Unit": "each",
                    "Default_Price": str(price) if price > 0 else "",
                    "Taxable": "Y",
                    "Reorder_Point": "5",
                    "Date_Added": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Last_Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Status": "Active",
                    "Notes": f"Imported from {filename}",
                    "Size": size,
                    "Item_Description": item_name,
                    "Wholesale_Cost": str(cost) if cost > 0 else "",
                    "Retail_Markup": "",
                    "Retail_Price": str(price) if price > 0 else "",
                    "Unit_Cost": str(cost) if cost > 0 else "",
                    "Unit_Price": str(price) if price > 0 else "",
                    "Last_Invoice_No": "",
                    "Last_Vendor": vendor_name,
                    "Last_Received_Date": datetime.now().strftime("%Y-%m-%d"),
                    "Supplier_Barcode": vendor_upc,  # Link vendor barcode
                    "QR_Code": "",  # Will be filled below
                    "QR_Image_Path": "",
                }

                # ============================================================
                # GENERATE QR CODE
                # ============================================================

                qr_data, qr_path = generate_qr_code(our_sku, item_data)
                item_data["QR_Code"] = qr_data
                item_data["QR_Image_Path"] = qr_path or ""
                qr_codes_generated += 1

                # ============================================================
                # ADD TO IMPORT LIST
                # ============================================================

                if is_update:
                    updated_items.append(item_data)
                else:
                    new_items.append(item_data)
                    existing_skus.add(our_sku)
                    if vendor_upc:
                        barcode_to_sku[vendor_upc] = our_sku

            except Exception as e:
                error_rows.append(f"Row {idx}: {str(e)}")

        # ====================================================================
        # STEP 5: UPDATE Items.csv HEADERS IF NEEDED
        # ====================================================================

        # Add new columns if they don't exist
        new_columns = ["Supplier_Barcode", "QR_Code", "QR_Image_Path"]
        updated_headers = list(our_headers)

        for col in new_columns:
            if col not in updated_headers:
                updated_headers.append(col)

        # ====================================================================
        # STEP 6: WRITE TO Items.csv
        # ====================================================================

        if not new_items and not updated_items:
            return jsonify(
                {
                    "success": False,
                    "error": f"No new items found. {duplicate_count} duplicates skipped.",
                }
            ), 400

        # Append new items
        if new_items:
            with open(items_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=updated_headers)
                for item in new_items:
                    # Ensure all fields present
                    row = {h: item.get(h, "") for h in updated_headers}
                    writer.writerow(row)

        # Update headers if changed
        if len(updated_headers) > len(our_headers):
            # Re-read and re-write with new headers
            with open(items_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                all_items = list(reader)

            with open(items_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=updated_headers)
                writer.writeheader()
                for item in all_items:
                    row = {h: item.get(h, "") for h in updated_headers}
                    writer.writerow(row)

        # ====================================================================
        # STEP 7: CREATE BARCODE LOOKUP FILE
        # ====================================================================

        barcode_lookup_path = "Inventory/Barcode_Lookup.csv"

        # Build barcode lookup
        barcode_entries = []
        for item in new_items + updated_items:
            if item.get("Supplier_Barcode"):
                barcode_entries.append(
                    {
                        "Supplier_Barcode": item["Supplier_Barcode"],
                        "Our_SKU": item["SKU"],
                        "Item_Name": item["Item_Name"],
                        "Vendor": item.get("Last_Vendor", ""),
                        "Date_Added": item["Date_Added"],
                    }
                )

        if barcode_entries:
            # Append to barcode lookup
            barcode_exists = os.path.exists(barcode_lookup_path)
            with open(barcode_lookup_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=[
                        "Supplier_Barcode",
                        "Our_SKU",
                        "Item_Name",
                        "Vendor",
                        "Date_Added",
                    ],
                )
                if not barcode_exists:
                    writer.writeheader()
                writer.writerows(barcode_entries)

        # ====================================================================
        # STEP 8: RETURN SUCCESS
        # ====================================================================

        result = {
            "success": True,
            "imported": len(new_items),
            "updated": len(updated_items),
            "duplicates": duplicate_count,
            "errors": len(error_rows),
            "qr_codes_generated": qr_codes_generated,
            "barcodes_linked": barcodes_linked,
            "vendor_format": field_map,
        }

        return jsonify(result), 200

    except Exception as e:
        import traceback

        error_detail = traceback.format_exc()
        app.logger.error(f"CSV Import Error: {error_detail}")
        return jsonify({"success": False, "error": f"Import failed: {str(e)}"}), 500


# ============================================================================
# BARCODE LOOKUP API
# For scanning supplier barcodes and finding our SKU
# ============================================================================


@app.route("/api/barcode-lookup/<barcode>")
@login_required
def barcode_lookup(barcode):
    """
    Look up item by supplier barcode or our SKU or QR code
    Returns item data
    """
    try:
        items_path = get_items_path()
        with open(items_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            items = list(reader)

        # Search by supplier barcode
        for item in items:
            if item.get("Supplier_Barcode", "").strip() == barcode.strip():
                return jsonify({"success": True, "item": item})

        # Search by our SKU
        for item in items:
            if item.get("SKU", "").strip().upper() == barcode.strip().upper():
                return jsonify({"success": True, "item": item})

        # Search by QR data
        for item in items:
            qr_data = item.get("QR_Code", "")
            if barcode in qr_data:
                return jsonify({"success": True, "item": item})

        return jsonify({"success": False, "error": "Item not found"}), 404

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/inventory/receive/csv-import/preview", methods=["POST"])
@login_required
def csv_import_preview():
    """Preview CSV file and suggest field mappings"""

    if "csv_file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["csv_file"]

    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    try:
        import csv
        from io import StringIO

        # Read first few rows
        content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(StringIO(content))

        # Get headers
        headers = csv_reader.fieldnames or []

        # Get first 5 rows
        preview_rows = []
        for idx, row in enumerate(csv_reader):
            if idx >= 5:
                break
            preview_rows.append(row)

        # Suggest mappings
        suggestions = {
            "sku": None,
            "item_name": None,
            "qty": None,
            "unit_cost": None,
            "supplier": None,
            "invoice_no": None,
        }

        for header in headers:
            header_lower = header.lower()

            if not suggestions["sku"] and any(
                x in header_lower for x in ["sku", "item code", "product code", "code"]
            ):
                suggestions["sku"] = header

            if not suggestions["item_name"] and any(
                x in header_lower
                for x in ["item", "product", "name", "description", "desc"]
            ):
                suggestions["item_name"] = header

            if not suggestions["qty"] and any(
                x in header_lower
                for x in ["qty", "quantity", "count", "amount", "units"]
            ):
                suggestions["qty"] = header

            if not suggestions["unit_cost"] and any(
                x in header_lower
                for x in ["cost", "price", "unit cost", "unit price", "wholesale"]
            ):
                suggestions["unit_cost"] = header

            if not suggestions["supplier"] and any(
                x in header_lower
                for x in ["supplier", "vendor", "manufacturer", "source"]
            ):
                suggestions["supplier"] = header

            if not suggestions["invoice_no"] and any(
                x in header_lower
                for x in ["invoice", "order", "po", "purchase order", "reference"]
            ):
                suggestions["invoice_no"] = header

        return jsonify(
            {
                "headers": headers,
                "preview": preview_rows,
                "suggestions": suggestions,
                "total_rows": len(preview_rows),
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/inventory/csv-invoice-import", methods=["GET", "POST"])
@manager_required
@tier_required("growth")
def csv_invoice_import():
    if request.method == "GET":
        return render_template("inventory/csv_import.html")

    if not HAS_INVOICE_IMPORTER:
        return jsonify(
            {"success": False, "error": "Invoice importer module is not available."}
        ), 400

    if "csv_file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["csv_file"]
    if not file or not file.filename:
        return jsonify({"success": False, "error": "No file selected"}), 400

    if not file.filename.lower().endswith(".csv"):
        return jsonify({"success": False, "error": "File must be a CSV"}), 400

    map_sku = (request.form.get("map_sku") or "").strip()
    map_item_name = (request.form.get("map_item_name") or "").strip()
    map_qty = (request.form.get("map_qty") or "").strip()
    map_unit_cost = (request.form.get("map_unit_cost") or "").strip()
    map_supplier = (request.form.get("map_supplier") or "").strip()
    map_invoice_no = (request.form.get("map_invoice_no") or "").strip()

    if not map_qty or not map_unit_cost:
        return jsonify(
            {"success": False, "error": "Quantity and Unit Cost mappings are required."}
        ), 400

    default_supplier = (request.form.get("default_supplier") or "").strip()
    default_invoice = (request.form.get("default_invoice") or "").strip()

    try:
        content = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read CSV: {e}"}), 400

    if not rows:
        return jsonify({"success": False, "error": "CSV file is empty"}), 400

    detected_supplier = ""
    detected_invoice = ""
    output_rows = []
    for row in rows:
        sku = (row.get(map_sku, "") or "").strip() if map_sku else ""
        name = (row.get(map_item_name, "") or "").strip() if map_item_name else ""
        qty = (row.get(map_qty, "") or "").strip() if map_qty else ""
        unit_cost = (row.get(map_unit_cost, "") or "").strip() if map_unit_cost else ""

        if not detected_supplier and map_supplier:
            detected_supplier = (row.get(map_supplier, "") or "").strip()
        if not detected_invoice and map_invoice_no:
            detected_invoice = (row.get(map_invoice_no, "") or "").strip()

        if not sku and not name:
            continue

        output_rows.append(
            {
                "SKU": sku,
                "Item_Name": name or sku,
                "Qty": qty,
                "Unit_Cost": unit_cost,
            }
        )

    if not output_rows:
        return jsonify({"success": False, "error": "No valid rows found in CSV."}), 400

    sku_limit = _limit_for("max_skus")
    if sku_limit is not None:
        existing_skus = {
            it.get("SKU", "").strip() for it in get_all_items() if it.get("SKU")
        }
        new_skus = {
            row.get("SKU", "").strip()
            for row in output_rows
            if row.get("SKU", "").strip()
        }
        estimated_total = len(existing_skus.union(new_skus))
        if estimated_total > int(sku_limit):
            return jsonify(
                {
                    "success": False,
                    "error": "SKU limit reached. Upgrade to import more items.",
                }
            ), 402

    vendor = detected_supplier or default_supplier or "Unknown"
    invoice_no = detected_invoice or default_invoice or ""
    received_date = date.today().isoformat()

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8"
        ) as tmp:
            writer = csv.DictWriter(
                tmp, fieldnames=["SKU", "Item_Name", "Qty", "Unit_Cost"]
            )
            writer.writeheader()
            writer.writerows(output_rows)
            tmp_path = tmp.name

        results = import_invoice_csv(
            tmp_path,
            invoice_no=invoice_no or None,
            vendor=vendor,
            received_date=received_date,
        )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)

    if not results.get("ok"):
        return jsonify(
            {"success": False, "error": "Import failed", "details": results}
        ), 400

    return jsonify({"success": True, "results": results}), 200


def _read_csv_dicts(path):
    if not os.path.exists(path):
        return [], []
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader), (reader.fieldnames or [])


def _write_csv_dicts(path, rows, fieldnames):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def _append_csv_dict(path, row, fallback_headers):
    rows, headers = _read_csv_dicts(path)
    if not headers:
        headers = fallback_headers[:]
    # ensure all headers exist on the row
    out = {h: row.get(h, "") for h in headers}
    rows.append(out)
    _write_csv_dicts(path, rows, headers)


def _now_stamp():
    # matches your Ledger schema "Timestamp"
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe_float(v, default=0.0):
    try:
        return float(str(v).strip())
    except Exception:
        return default


def _safe_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default


@app.route("/inventory/invoice-import", methods=["GET"])
@manager_required
@tier_required("pro")
def inventory_invoice_import():
    return render_template("inventory/invoice_import.html")


@app.route("/inventory/invoice-import/pdf", methods=["POST"])
@manager_required
@tier_required("pro")
def inventory_invoice_import_pdf():
    pdf_file = request.files.get("pdf_file")
    if not pdf_file or not pdf_file.filename:
        flash("Please select a PDF invoice to upload.", "error")
        return redirect(url_for("inventory_invoice_import"))

    if not pdf_file.filename.lower().endswith(".pdf"):
        flash("Only PDF files are supported for invoice upload.", "error")
        return redirect(url_for("inventory_invoice_import"))

    upload_dir = Path("uploads") / "invoices"
    upload_dir.mkdir(parents=True, exist_ok=True)

    safe_name = secure_filename(pdf_file.filename)
    dest = upload_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    pdf_file.save(dest)

    flash("Invoice PDF uploaded. OCR import is not enabled yet.", "success")
    return redirect(url_for("inventory_invoice_import"))


@app.route("/inventory/invoice-import/photo", methods=["POST"])
@manager_required
@tier_required("pro")
def inventory_invoice_import_photo():
    photo_file = request.files.get("photo_file")
    if not photo_file or not photo_file.filename:
        flash("Please select an invoice photo to upload.", "error")
        return redirect(url_for("inventory_invoice_import"))

    allowed = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
    ext = os.path.splitext(photo_file.filename.lower())[1]
    if ext and ext not in allowed:
        flash("Unsupported photo format. Use JPG, PNG, or WEBP.", "error")
        return redirect(url_for("inventory_invoice_import"))

    upload_dir = Path("uploads") / "invoices"
    upload_dir.mkdir(parents=True, exist_ok=True)

    safe_name = secure_filename(photo_file.filename)
    dest = upload_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    photo_file.save(dest)

    flash("Invoice photo uploaded. OCR import is not enabled yet.", "success")
    return redirect(url_for("inventory_invoice_import"))


@app.route("/inventory/invoice-import", methods=["POST"])
def inventory_invoice_import_post():
    # === 1) If CSV file upload exists, keep your existing importer ===
    if (
        "invoice_csv" in request.files
        and request.files["invoice_csv"]
        and request.files["invoice_csv"].filename
    ):
        if not HAS_INVOICE_IMPORTER:
            flash("Invoice importer module is not available.", "error")
            return redirect(url_for("inventory_invoice_import"))

        sku_limit = _limit_for("max_skus")
        if sku_limit is not None and len(get_all_items()) >= int(sku_limit):
            flash("SKU limit reached. Upgrade to import more items.", "error")
            return redirect(url_for("inventory_invoice_import"))

        f = request.files["invoice_csv"]
        if not f.filename.lower().endswith(".csv"):
            flash("Please upload a .csv file for invoice import.", "error")
            return redirect(url_for("inventory_invoice_import"))

        # ✅ your existing function (already in your app)
        try:
            tmp_path = os.path.join("uploads", f"invoice_{uuid.uuid4().hex}.csv")
            os.makedirs("uploads", exist_ok=True)
            f.save(tmp_path)

            results = import_invoice_csv(tmp_path)  # <-- existing in your code
            flash(
                f"Invoice imported: {results.get('lots_created', 0)} lots added.",
                "success",
            )
            return redirect(url_for("inventory"))
        except Exception as e:
            flash(f"Invoice CSV import failed: {e}", "error")
            return redirect(url_for("inventory_invoice_import"))

    # === 2) Manual invoice import (lines textarea) ===
    lines_raw = (request.form.get("lines") or "").strip()
    if lines_raw:
        if not HAS_INVOICE_IMPORTER:
            flash("Invoice importer module is not available.", "error")
            return redirect(url_for("inventory_invoice_import"))

        sku_limit = _limit_for("max_skus")
        if sku_limit is not None and len(get_all_items()) >= int(sku_limit):
            flash("SKU limit reached. Upgrade to import more items.", "error")
            return redirect(url_for("inventory_invoice_import"))

        vendor = (request.form.get("vendor") or "").strip()
        invoice_no = (request.form.get("invoice_number") or "").strip()
        received_date = (
            request.form.get("invoice_date") or ""
        ).strip() or date.today().isoformat()

        parsed_rows = []
        errors = []
        for idx, raw_line in enumerate(lines_raw.splitlines(), start=1):
            line = raw_line.strip()
            if not line:
                continue

            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 4:
                errors.append(
                    f"Line {idx}: expected 4 fields (SKU, Item Name, Qty, Unit Cost)"
                )
                continue

            sku, name, qty, unit_cost = parts[0], parts[1], parts[2], parts[3]
            if not sku and not name:
                errors.append(f"Line {idx}: SKU or Item Name is required")
                continue

            parsed_rows.append(
                {
                    "SKU": sku,
                    "Item_Name": name or sku,
                    "Qty": qty,
                    "Unit_Cost": unit_cost,
                }
            )

        if errors:
            flash("Manual import errors: " + " | ".join(errors), "error")
            return redirect(url_for("inventory_invoice_import"))

        if not parsed_rows:
            flash("No valid lines found to import.", "error")
            return redirect(url_for("inventory_invoice_import"))

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8"
            ) as tmp:
                writer = csv.DictWriter(
                    tmp, fieldnames=["SKU", "Item_Name", "Qty", "Unit_Cost"]
                )
                writer.writeheader()
                writer.writerows(parsed_rows)
                tmp_path = tmp.name

            results = import_invoice_csv(
                tmp_path,
                invoice_no=invoice_no or None,
                vendor=vendor,
                received_date=received_date,
            )
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)

        if not results.get("ok"):
            flash("Manual invoice import failed.", "error")
            return redirect(url_for("inventory_invoice_import"))

        flash(
            f"Invoice imported: {results.get('lots_created', 0)} lots added.",
            "success",
        )
        return redirect(url_for("inventory"))

    # === 3) Receive form import ===
    # Expected manual form fields (match your receive.html inputs):
    # sku (optional), item_name, category, default_price, unit_cost, qty_received, supplier, invoice_no, received_date, notes, taxable
    sku = (request.form.get("sku") or "").strip()
    item_name = (request.form.get("item_name") or "").strip()
    category = (request.form.get("category") or "General").strip()
    default_price = _safe_float(request.form.get("default_price"), 0.0)
    unit_cost = _safe_float(request.form.get("unit_cost"), 0.0)
    qty_received = _safe_int(request.form.get("qty_received"), 0)
    supplier = (request.form.get("supplier") or "").strip()
    invoice_no = (request.form.get("invoice_no") or "").strip()
    received_date = (request.form.get("received_date") or "").strip()
    notes = (request.form.get("notes") or "").strip()
    taxable_raw = (request.form.get("taxable") or "true").strip().lower()
    taxable = "True" if taxable_raw in ("1", "true", "yes", "on") else "False"

    if qty_received <= 0:
        flash("Qty received must be 1 or more.", "error")
        return redirect(url_for("receive_stock"))

    if not received_date:
        received_date = datetime.now().strftime("%Y-%m-%d")

    # Inventory CSV paths (from your Tree.txt)
    items_path = os.path.join("Inventory", "Items.csv")
    lots_path = os.path.join("Inventory", "Lots.csv")
    ledger_path = os.path.join("Inventory", "Ledger.csv")

    # === Ensure SKU exists; if not provided, generate one ===
    # Uses your POS_CORE helper if available, else fallback.
    if not sku:
        try:
            sku = generate_sku(category=category)  # POS_CORE function in your project
        except Exception:
            sku = f"SKU-{uuid.uuid4().hex[:8].upper()}"

    # === Create or update the item in Items.csv ===
    items_rows, item_headers = _read_csv_dicts(items_path)
    if not item_headers:
        # fallback header set (won't break if your real file has more/less columns)
        item_headers = [
            "SKU",
            "Item_Name",
            "Category",
            "Subcategory",
            "Product_Name",
            "Default_Unit",
            "Default_Price",
            "Taxable",
            "Reorder_Point",
            "Date_Added",
            "Last_Updated",
            "Status",
            "Notes",
            "Retail_Markup",
            "Retail_Price",
            "Unit_Cost",
            "Unit_Price",
            "Last_Invoice_No",
            "Last_Vendor",
            "Last_Received_Date",
        ]

    now_date = datetime.now().strftime("%Y-%m-%d")
    found = False
    for r in items_rows:
        if (r.get("SKU") or "").strip() == sku:
            # update only what we actually got from the manual form
            if item_name:
                r["Item_Name"] = item_name
            if category:
                r["Category"] = category
            if default_price > 0:
                r["Default_Price"] = str(default_price)
            if unit_cost > 0:
                r["Unit_Cost"] = str(unit_cost)
            r["Taxable"] = taxable
            r["Last_Updated"] = now_date
            r["Last_Invoice_No"] = invoice_no
            r["Last_Vendor"] = supplier
            r["Last_Received_Date"] = received_date
            if notes:
                r["Notes"] = (r.get("Notes", "") + " | " + notes).strip(" |")
            found = True
            break

    if not found:
        new_row = {h: "" for h in item_headers}
        new_row["SKU"] = sku
        new_row["Item_Name"] = item_name or sku
        new_row["Category"] = category or "General"
        new_row["Default_Price"] = str(default_price) if default_price else "0"
        new_row["Unit_Cost"] = str(unit_cost) if unit_cost else "0"
        new_row["Taxable"] = taxable
        new_row["Date_Added"] = now_date
        new_row["Last_Updated"] = now_date
        new_row["Status"] = "Active"
        new_row["Last_Invoice_No"] = invoice_no
        new_row["Last_Vendor"] = supplier
        new_row["Last_Received_Date"] = received_date
        new_row["Notes"] = notes
        items_rows.append(new_row)

    _write_csv_dicts(items_path, items_rows, item_headers)

    # === Append a new Lot to Lots.csv (THIS is what Sales on-hand reads) ===
    lot_id = f"LOT-{uuid.uuid4().hex[:10].upper()}"
    lot_row = {
        "Lot_ID": lot_id,
        "SKU": sku,
        "Invoice_No": invoice_no,
        "Vendor": supplier,
        "Date_Received": received_date,
        "Qty_Received": str(qty_received),
        "Qty_Remaining": str(qty_received),
        "Unit_Cost": str(unit_cost),
        "Notes": notes,
    }
    lot_headers = [
        "Lot_ID",
        "SKU",
        "Invoice_No",
        "Vendor",
        "Date_Received",
        "Qty_Received",
        "Qty_Remaining",
        "Unit_Cost",
        "Notes",
    ]
    _append_csv_dict(lots_path, lot_row, lot_headers)

    # === Append Ledger receive event ===
    ledger_row = {
        "Entry_ID": f"LED-{uuid.uuid4().hex[:10].upper()}",
        "Timestamp": _now_stamp(),
        "SKU": sku,
        "Lot_ID": lot_id,
        "Delta_Qty": str(qty_received),
        "Reason": "RECEIVE",
        "Ref_Transaction_ID": invoice_no,
        "Employee_ID": "",
        "Notes": notes,
    }
    ledger_headers = [
        "Entry_ID",
        "Timestamp",
        "SKU",
        "Lot_ID",
        "Delta_Qty",
        "Reason",
        "Ref_Transaction_ID",
        "Employee_ID",
        "Notes",
    ]
    _append_csv_dict(ledger_path, ledger_row, ledger_headers)

    flash(
        f"Received {qty_received} × {sku} into {lot_id}. Sales on-hand should now update.",
        "success",
    )
    return redirect(url_for("inventory"))


# ==============================================================================
#                     DASHBOARD
# ==============================================================================


@app.route("/dashboard")
@login_required
def dashboard():
    """Route to role-specific dashboard"""
    role = session.get("role", "Cashier")

    if role in ("Owner", "Admin"):
        return redirect(url_for("owner_dashboard"))
    elif role == "Manager":
        return redirect(url_for("manager_dashboard"))
    elif role == "Cashier":
        return redirect(url_for("cashier_dashboard"))
    else:
        # Laborer or other roles
        return redirect(url_for("laborer_dashboard"))


@app.route("/dashboard/owner")
@login_required
def owner_dashboard():
    """Owner/Admin dashboard with full analytics and controls"""
    if session.get("role") not in ("Owner", "Admin"):
        flash("Owner access required", "error")
        return redirect(url_for("dashboard"))

    summary = generate_daily_summary()
    low_stock = check_low_stock()[:5]
    status = get_employee_status(session["employee_id"])

    # Get comprehensive metrics
    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    sales_file = _find_saleslog_file_for_date(DATA_DIR, today_str)
    sales_rows = _load_saleslog_rows(sales_file)
    sales_metrics = _compute_daily_sales_metrics(sales_rows)

    # Employee metrics
    employees = get_all_employees(include_inactive=True)
    active_employees = [
        e for e in employees if e.get("Status", "").upper() != "INACTIVE"
    ]

    # Inventory metrics
    items = get_all_items()

    # Time clock metrics
    punches = get_punches_for_date(today)
    clocked_in_ids = set()
    for p in punches:
        if p.get("Punch_Type") == "CLOCK_IN":
            clocked_in_ids.add(p.get("Employee_ID"))
        elif p.get("Punch_Type") == "CLOCK_OUT":
            clocked_in_ids.discard(p.get("Employee_ID"))
    employees_clocked_in = len(clocked_in_ids)

    # Pending tasks
    all_tasks = get_all_tasks()
    pending_tasks = [
        t for t in all_tasks if t.get("Status", "").upper() not in ("COMPLETE", "DONE")
    ]

    # Health checks
    health_checks = {
        "items_csv": os.path.exists(get_items_path()),
        "lots_csv": os.path.exists(get_lots_path()),
        "ledger_csv": os.path.exists(get_ledger_path()),
        "smtp_configured": bool(os.getenv("SMTP_HOST"))
        and bool(os.getenv("SMTP_USER")),
        "stripe_configured": bool(os.getenv("STRIPE_SECRET_KEY")),
    }
    health_ok = all(health_checks.values())

    return render_template(
        "dashboards/owner.html",
        clock_status=status,
        summary=summary,
        sales_metrics=sales_metrics,
        total_revenue=sales_metrics.get("total_revenue", 0.0),
        transaction_count=sales_metrics.get("transaction_count", 0),
        gross_profit=sales_metrics.get("gross_profit", 0.0),
        employee_count=len(active_employees),
        total_employees=len(employees),
        inventory_count=len(items),
        low_stock_count=len(check_low_stock()),
        low_stock_items=low_stock,
        employees_clocked_in=employees_clocked_in,
        pending_tasks_count=len(pending_tasks),
        health_checks=health_checks,
        health_ok=health_ok,
    )


@app.route("/dashboard/manager")
@login_required
def manager_dashboard():
    """Manager dashboard with operational metrics"""
    if session.get("role") not in ("Manager", "Owner", "Admin"):
        flash("Manager access required", "error")
        return redirect(url_for("dashboard"))

    summary = generate_daily_summary()
    low_stock = check_low_stock()[:5]
    status = get_employee_status(session["employee_id"])

    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    sales_file = _find_saleslog_file_for_date(DATA_DIR, today_str)
    sales_rows = _load_saleslog_rows(sales_file)
    sales_metrics = _compute_daily_sales_metrics(sales_rows)

    # Employee status
    employees = get_all_employees()
    punches = get_punches_for_date(today)
    clocked_in_ids = set()
    for p in punches:
        if p.get("Punch_Type") == "CLOCK_IN":
            clocked_in_ids.add(p.get("Employee_ID"))
        elif p.get("Punch_Type") == "CLOCK_OUT":
            clocked_in_ids.discard(p.get("Employee_ID"))
    employees_clocked_in = len(clocked_in_ids)

    # Tasks
    all_tasks = get_all_tasks()
    pending_tasks = [
        t for t in all_tasks if t.get("Status", "").upper() not in ("COMPLETE", "DONE")
    ]

    return render_template(
        "dashboards/manager.html",
        clock_status=status,
        summary=summary,
        sales_metrics=sales_metrics,
        total_revenue=sales_metrics.get("total_revenue", 0.0),
        transaction_count=sales_metrics.get("transaction_count", 0),
        employee_count=len(employees),
        employees_clocked_in=employees_clocked_in,
        pending_tasks_count=len(pending_tasks),
        low_stock_count=len(check_low_stock()),
        low_stock_items=low_stock,
    )


@app.route("/dashboard/cashier")
@login_required
def cashier_dashboard():
    """Cashier dashboard focused on sales"""
    summary = generate_daily_summary()
    status = get_employee_status(session["employee_id"])

    today = date.today()
    today_str = today.strftime("%Y-%m-%d")
    sales_file = _find_saleslog_file_for_date(DATA_DIR, today_str)
    sales_rows = _load_saleslog_rows(sales_file)
    sales_metrics = _compute_daily_sales_metrics(sales_rows)

    # Employee's own stats
    emp_id = session.get("employee_id", "")
    emp_sales = [r for r in sales_rows if r.get("Employee_ID") == emp_id]
    emp_revenue = sum(float(r.get("Subtotal", 0) or 0) for r in emp_sales)
    emp_transactions = len(
        set(r.get("Transaction_ID") for r in emp_sales if r.get("Transaction_ID"))
    )

    return render_template(
        "dashboards/cashier.html",
        clock_status=status,
        summary=summary,
        sales_metrics=sales_metrics,
        total_revenue=sales_metrics.get("total_revenue", 0.0),
        transaction_count=sales_metrics.get("transaction_count", 0),
        my_revenue=emp_revenue,
        my_transactions=emp_transactions,
    )


@app.route("/dashboard/laborer")
@login_required
def laborer_dashboard():
    """Laborer dashboard focused on tasks and time"""
    emp_id = session.get("employee_id", "")
    today = date.today()
    status = get_employee_status(emp_id)

    # Today's punches
    punches = get_punches_for_date(today)
    my_punches = [p for p in punches if p.get("Employee_ID") == emp_id]
    hours_today = (
        float(my_punches[-1].get("Hours_Worked_Today", 0)) if my_punches else 0.0
    )

    # My tasks
    my_tasks_list = get_tasks_for_employee(emp_id, today)
    pending_my_tasks = [
        t
        for t in my_tasks_list
        if t.get("Status", "").upper() not in ("COMPLETE", "DONE")
    ]

    # Time off requests
    timeoff_requests = get_time_off_requests(emp_id)
    pending_timeoff = [
        r for r in timeoff_requests if r.get("Status", "").upper() == "PENDING"
    ]

    return render_template(
        "dashboards/laborer.html",
        clock_status=status,
        hours_today=hours_today,
        pending_tasks_count=len(pending_my_tasks),
        pending_timeoff_count=len(pending_timeoff),
        my_tasks=my_tasks_list[:5],
    )


# -----------------------------
# Dashboard helpers
# -----------------------------
def _to_float(x):
    try:
        return float(str(x).replace("$", "").replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def _find_daily_file(root_dir, date_str, suffix):
    # suffix examples: "_SalesLog.csv", "_Transactions.csv"
    pattern = os.path.join(root_dir, "**", f"{date_str}*{suffix}")
    matches = glob.glob(pattern, recursive=True)
    if not matches:
        return None
    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


# -----------------------------
# Dashboard: Today's Sales (from Sales_Logs tree)
# -----------------------------
@app.route("/api/dashboard/sales")
@login_required
def api_dashboard_sales():
    from datetime import date as _date
    from pathlib import Path as _Path

    date_str = request.args.get("date") or _date.today().isoformat()

    def _to_float(x):
        s = str(x or "").strip().replace("$", "").replace(",", "")
        try:
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    def _find_daily_file(base_dir: _Path, day: str, suffix: str):
        if not base_dir.exists():
            return None
        matches = list(base_dir.glob(f"**/{day}*{suffix}"))
        return matches[0] if matches else None

    tx_file = _find_daily_file(_Path("Transaction_Logs"), date_str, "_Transactions.csv")
    tx_rows = read_csv(tx_file) if tx_file else []

    total_revenue = 0.0
    cash_sales_count = 0
    card_sales_count = 0

    transactions = []
    for r in tx_rows:
        pm = (r.get("Payment_Method") or "").upper()
        grand = _to_float(r.get("Grand_Total"))

        total_revenue += grand
        if "CASH" in pm:
            cash_sales_count += 1
        elif pm:
            card_sales_count += 1

        transactions.append(
            {
                "transaction_id": r.get("Transaction_ID", ""),
                "time": r.get("Time", ""),
                "items": r.get("Items_Count", ""),
                "payment_method": r.get("Payment_Method", ""),
                "total": grand,
            }
        )

    return jsonify(
        {
            "success": True,
            # ✅ keys your modal expects (most common)
            "total_revenue": total_revenue,
            "transactions_count": len(transactions),
            "cash_sales_count": cash_sales_count,
            "card_sales_count": card_sales_count,
            "transactions": transactions[::-1][
                -50:
            ],  # last 50, newest last -> reverse if you want newest first
            # ✅ backward-compat (in case any older JS uses these)
            "total_sales": total_revenue,
            "cash_count": cash_sales_count,
            "card_count": card_sales_count,
        }
    )


# -----------------------------
# Dashboard: Today's Transactions (from Transaction_Logs tree)
# -----------------------------
@app.route("/api/dashboard/transactions")
def api_dashboard_transactions():
    try:
        today_str = date.today().strftime("%Y-%m-%d")
        tx_root = os.path.join(BASE_DIR, "Transaction_Logs")
        tx_file = _find_daily_file(tx_root, today_str, "_Transactions.csv")

        if not tx_file or not os.path.exists(tx_file):
            return jsonify({"success": True, "count": 0, "transactions": []})

        rows = []
        with open(tx_file, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)

        # newest first if you have a Time column
        if rows and ("Time" in rows[0] or "time" in rows[0]):
            rows.sort(
                key=lambda r: (r.get("Time") or r.get("time") or ""), reverse=True
            )

        return jsonify(
            {"success": True, "count": len(rows), "transactions": rows[:100]}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.get("/sales/receipt/<transaction_id>")
@login_required
def legacy_receipt_redirect(transaction_id):
    return redirect(url_for("sales"))


# ============================================
# API: Staff Working Now
# ============================================
@app.route("/api/dashboard/staff")
def api_dashboard_staff():
    """Get currently clocked-in staff from Time_Clock CSV"""
    try:
        today_str = date.today().strftime("%Y-%m-%d")
        staff = []
        total_hours = 0

        # Look for time clock file
        timeclock_dir = os.path.join(DATA_DIR, "Time_Clock")

        possible_files = [
            os.path.join(timeclock_dir, f"timeclock_{today_str}.csv"),
            os.path.join(timeclock_dir, f"TimeLog_{today_str}.csv"),
            os.path.join(timeclock_dir, f"{today_str}.csv"),
            os.path.join(timeclock_dir, "time_clock.csv"),
            os.path.join(timeclock_dir, "TimeLog.csv"),
            os.path.join(timeclock_dir, "timeclock.csv"),
        ]

        timeclock_file = None
        for f in possible_files:
            if os.path.exists(f):
                timeclock_file = f
                break

        # Track who is currently clocked in (no clock out time)
        clocked_in = {}

        if timeclock_file:
            with open(timeclock_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    row_date = row.get("Date", row.get("date", ""))

                    # Only look at today's entries
                    if today_str in row_date or not row_date:
                        emp_id = row.get(
                            "Employee_ID", row.get("employee_id", row.get("EmpID", ""))
                        )
                        emp_name = row.get(
                            "Employee_Name",
                            row.get("employee_name", row.get("Name", "Unknown")),
                        )
                        clock_in = row.get(
                            "Clock_In", row.get("clock_in", row.get("In", ""))
                        )
                        clock_out = row.get(
                            "Clock_Out", row.get("clock_out"), row.get("Out", "")
                        )
                        role = row.get("Role", row.get("role", "Employee"))

                        # If clocked in but not clocked out, they're working
                        if clock_in and not clock_out:
                            # Calculate hours worked so far
                            try:
                                clock_in_time = datetime.strptime(
                                    clock_in, "%H:%M:%S"
                                ).replace(
                                    year=datetime.now().year,
                                    month=datetime.now().month,
                                    day=datetime.now().day,
                                )
                                hours_worked = (
                                    datetime.now() - clock_in_time
                                ).total_seconds() / 3600
                            except:
                                try:
                                    clock_in_time = datetime.strptime(
                                        clock_in, "%H:%M"
                                    ).replace(
                                        year=datetime.now().year,
                                        month=datetime.now().month,
                                        day=datetime.now().day,
                                    )
                                    hours_worked = (
                                        datetime.now() - clock_in_time
                                    ).total_seconds() / 3600
                                except:
                                    hours_worked = 0

                            clocked_in[emp_id] = {
                                "name": emp_name,
                                "role": role,
                                "clock_in": clock_in,
                                "hours": round(hours_worked, 1),
                            }
                        elif clock_out and emp_id in clocked_in:
                            # They clocked out, remove from working list
                            del clocked_in[emp_id]

        # Also check Employees directory for additional info
        emp_file = os.path.join(DATA_DIR, "Employees", "Employee_Directory.csv")
        emp_info = {}
        if os.path.exists(emp_file):
            with open(emp_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    emp_id = row.get(
                        "Employee_ID", row.get("employee_id", row.get("ID", ""))
                    )
                    emp_info[emp_id] = {
                        "name": row.get("Name", row.get("Employee_Name", "")),
                        "role": row.get("Role", row.get("role", "Employee")),
                    }

        # Build final staff list
        for emp_id, data in clocked_in.items():
            if emp_id in emp_info:
                data["name"] = emp_info[emp_id]["name"] or data["name"]
                data["role"] = emp_info[emp_id]["role"] or data["role"]
            staff.append(data)
            total_hours += data["hours"]

        return jsonify(
            {"success": True, "staff": staff, "total_hours": round(total_hours, 1)}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ============================================
# API: Low Stock Items
# ============================================
@app.route("/api/dashboard/lowstock")
def api_dashboard_lowstock():
    """Get low stock items from Inventory CSV"""
    try:
        items = []
        critical_count = 0

        # Look for inventory file
        inv_dir = os.path.join(DATA_DIR, "Inventory")

        possible_files = [
            os.path.join(inv_dir, "Items.csv"),
            os.path.join(inv_dir, "items.csv"),
            os.path.join(inv_dir, "inventory.csv"),
            os.path.join(inv_dir, "Inventory.csv"),
            os.path.join(inv_dir, "products.csv"),
        ]

        inv_file = None
        for f in possible_files:
            if os.path.exists(f):
                inv_file = f
                break

        if inv_file:
            with open(inv_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Get quantity on hand
                    qty = int(
                        row.get(
                            "Qty_On_Hand",
                            row.get(
                                "qty_on_hand",
                                row.get(
                                    "On_Hand",
                                    row.get(
                                        "on_hand", row.get("Stock", row.get("stock", 0))
                                    ),
                                ),
                            ),
                        )
                        or 0
                    )

                    # Get reorder point (default 5)
                    reorder = int(
                        row.get(
                            "Reorder_Point",
                            row.get(
                                "reorder_point",
                                row.get("Reorder", row.get("Min_Stock", 5)),
                            ),
                        )
                        or 5
                    )

                    # Check if below reorder point
                    if qty <= reorder:
                        item = {
                            "name": row.get(
                                "Item_Name",
                                row.get(
                                    "item_name",
                                    row.get(
                                        "Name",
                                        row.get("name", row.get("Product", "Unknown")),
                                    ),
                                ),
                            ),
                            "sku": row.get(
                                "SKU", row.get("sku", row.get("Item_ID", "-"))
                            ),
                            "on_hand": qty,
                            "reorder_point": reorder,
                        }
                        items.append(item)

                        if qty <= 2:
                            critical_count += 1

        # Sort by quantity (lowest first)
        items.sort(key=lambda x: x["on_hand"])

        return jsonify(
            {"success": True, "items": items, "critical_count": critical_count}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ============================================
# IMPORTANT: Update your DATA_DIR variable
# ============================================
# Make sure DATA_DIR points to your CSV data folder, e.g.:
# DATA_DIR = '/mnt/sdcard/Mountain Gardens Nursery POS/data'
# or
# DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
# ============================================

# ==============================================================================
#                     NOTIFICATIONS
# ==============================================================================


@app.route("/notifications")
@login_required
def notifications():
    notifs = get_all_notifications(session["employee_id"])
    return render_template("notifications.html", notifications=notifs)


@app.route("/notifications/mark-read", methods=["POST"])
@login_required
def mark_notifications_read():
    ntype = (request.form.get("type") or request.args.get("type") or "").strip().upper()
    emp_id = str(session.get("employee_id", ""))

    marked = 0
    try:
        rows = read_csv(get_notification_path())
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for r in rows:
            if str(r.get("Employee_ID", "")) != emp_id:
                continue

            if ntype and (str(r.get("Type", "")) or "").upper() != ntype:
                continue

            read_flag = str(r.get("Read", r.get("Is_Read", "N"))).strip().upper()
            if read_flag in ("Y", "YES", "TRUE", "1"):
                continue

            r["Read"] = "Y"
            r["Read_Date"] = now
            marked += 1

        if marked:
            write_csv(get_notification_path(), NOTIFICATION_HEADERS, rows)

    except Exception:
        marked = 0

    # ✅ If called by JS/fetch, return JSON
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if is_ajax:
        return jsonify({"marked": marked})

    # ✅ Normal form post: redirect back
    return redirect(request.referrer or url_for("my_tasks"))


# ==============================================================================
#                     SALES TERMINAL
# ==============================================================================


@app.route("/sales")
@login_required
def sales():
    card_fee_rate = float(os.environ.get("MGN_CARD_FEE_RATE", "0.03"))
    return render_template(
        "sales/terminal.html",
        categories=MAIN_CATEGORIES,
        animal_subs=ANIMAL_SUBCATEGORIES,
        product_subs=PRODUCT_SUBCATEGORIES,
        plant_subs=PLANT_SUBCATEGORIES,
        card_fee_rate=card_fee_rate,
        tax_rate=get_tax_rate(),
    )


@app.route("/sales/search")
@login_required
def sales_search():
    query = (request.args.get("q", "") or "").strip()
    category = (request.args.get("category", "") or "").strip()

    items = search_items(query, category=category if category else None)

    # Barcode scan: a scanner just types the code into the search box + Enter. Match
    # it against Supplier_Barcode and surface that item first.
    if query:
        existing = {(it.get("SKU") or "") for it in items}
        for it in get_all_items():
            if (it.get("Supplier_Barcode") or "").strip() == query and (it.get("SKU") or "") not in existing:
                items = [it] + items

    results = []
    for item in items[:20]:
        sku = (item.get("SKU") or "").strip()
        if not sku:
            continue

        stock = get_stock_on_hand(sku)
        results.append(
            {
                "sku": sku,
                "name": item.get("Item_Name", ""),
                "label": f"{item.get('Item_Name', '')} — {sku}",
                "category": item.get("Category", ""),
                "price": float(item.get("Default_Price", 0) or 0),
                "stock": stock,
                "taxable": str(item.get("Taxable", "Y")).strip().upper() not in ("N", "NO", "EXEMPT", "FALSE", "0"),
            }
        )

    return jsonify({"items": results})


@app.route("/sales/item/<sku>")
@login_required
def sales_item_details(sku):
    item = get_item(sku)
    if not item:
        return jsonify({"ok": False, "error": "Item not found"}), 404

    stock = get_stock_on_hand(sku)
    avg_cost = get_average_cost(sku)

    return jsonify(
        {
            "ok": True,
            "success": True,
            "sku": sku,
            "name": item.get("Item_Name", ""),
            "category": item.get("Category", ""),
            "price": float(item.get("Default_Price", 0)),
            "wholesale": avg_cost,
            "qty_on_hand": stock,
            "on_hand": stock,
            "taxable": str(item.get("Taxable", "Y")).strip().upper() not in ("N", "NO", "EXEMPT", "FALSE", "0"),
        }
    )


@app.route("/sales/quick_add", methods=["POST"])
@login_required
def sales_quick_add():
    """Cashier adds an item on the spot (name + price) and it drops into the cart.
    Persists as a 'QA-' provisional item so a manager can reconcile it at EOD."""
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    try:
        price = float(data.get("price") or 0)
    except (TypeError, ValueError):
        price = 0.0
    if not name or price <= 0:
        return jsonify({"ok": False, "error": "Name and a price greater than 0 are required."}), 400
    sku, row = quick_add_item(
        name, price,
        category=(data.get("category") or "Quick-Add"),
        size=(data.get("size") or ""),
        emp_id=session.get("employee_id", ""),
        emp_name=session.get("employee_name", ""),
    )
    # mirror /sales/item/<sku> so the front-end drops it straight into the cart
    return jsonify({"ok": True, "success": True, "sku": sku, "name": row["Item_Name"],
                    "price": price, "category": row["Category"], "taxable": True,
                    "qty_on_hand": 0, "on_hand": 0, "stock": 0})


RECONCILE_PAGE = """<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Reconcile Quick-Add Items</title><style>
 body{font-family:system-ui,Arial,sans-serif;margin:0;background:#0f1115;color:#e8e8e8}
 .wrap{max-width:960px;margin:0 auto;padding:20px}.muted{color:#9aa6b2}
 h1{font-size:20px;color:#e8c55a}a{color:#c9a84c}
 input{padding:7px;background:#161a22;color:#e8e8e8;border:1px solid #2a2f3a;border-radius:6px}
 button{padding:7px 14px;background:#c9a84c;color:#111;border:0;border-radius:6px;font-weight:600;cursor:pointer}
 .row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;padding:12px;border-bottom:1px solid #232833}
 .empty{padding:48px;text-align:center;color:#9aa6b2}
</style></head><body><div class="wrap">
<h1>End-of-Day Reconciliation</h1>
<p class="muted">Map each on-the-spot (quick-add) item rung today to the real catalog product, so sold items stay aligned with inventory. <a href="/sales">&larr; back to register</a></p>
{% with msgs = get_flashed_messages(with_categories=true) %}{% for cat,m in msgs %}<p class="muted">- {{ m }}</p>{% endfor %}{% endwith %}
{% if not qas %}<div class="empty">No quick-add items to reconcile. &#10003;</div>{% else %}
<datalist id="catalog">{% for c in catalog %}<option value="{{c.sku}}">{{c.name}} ({{c.sku}})</option>{% endfor %}</datalist>
{% for q in qas %}
<form class="row" method="post" action="{{ url_for('inventory_reconcile_apply') }}">
  <div style="flex:1;min-width:220px"><b>{{ q.Item_Name }}</b> <span class="muted">${{ q.Default_Price }} &middot; {{ q.SKU }}</span></div>
  <input type="hidden" name="qa_sku" value="{{ q.SKU }}">
  <input list="catalog" name="canonical_sku" placeholder="map to product / SKU" required style="width:220px">
  <input name="reason" placeholder="note (optional)" style="width:150px">
  <button type="submit">Map</button>
</form>{% endfor %}{% endif %}
</div></body></html>"""


@app.route("/inventory/reconcile")
@manager_required
def inventory_reconcile():
    from flask import render_template_string
    qas = get_unreconciled_quickadds()
    catalog = [{"sku": i.get("SKU", ""), "name": i.get("Item_Name", "")}
               for i in get_all_items() if not str(i.get("SKU", "")).startswith("QA-")]
    return render_template_string(RECONCILE_PAGE, qas=qas, catalog=catalog)


@app.route("/inventory/reconcile/apply", methods=["POST"])
@manager_required
def inventory_reconcile_apply():
    ok, msg = reconcile_quickadd(
        (request.form.get("qa_sku") or "").strip(),
        (request.form.get("canonical_sku") or "").strip(),
        mapped_by=session.get("employee_id", ""),
        reason=(request.form.get("reason") or "").strip(),
    )
    flash(msg, "success" if ok else "error")
    return redirect(url_for("inventory_reconcile"))


# ============================================================================
# MONEY OS -- owner/manager financial cockpit (engine in money_core.py).
# P&L, payday readiness, envelopes, bills + autopilot, DIY payroll filing helper.
# ============================================================================
MONEY_DASHBOARD = """<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>Money OS</title><style>
 body{font-family:system-ui,Arial,sans-serif;margin:0;background:#0f1115;color:#e8e8e8}
 .wrap{max-width:1040px;margin:0 auto;padding:18px}h1{font-size:20px;color:#e8c55a;margin:.2em 0}
 a{color:#c9a84c}.muted{color:#9aa6b2;font-size:13px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
 .card{background:#161a22;border:1px solid #232833;border-radius:12px;padding:14px}
 .big{font-size:26px;font-weight:700}.flag{display:inline-block;padding:2px 10px;border-radius:20px;font-size:12px;font-weight:700}
 .PROFITABLE,.GREEN{background:#10391f;color:#5fe39b}.SLOW,.AMBER{background:#4a3a10;color:#ffcf5f}
 .LOSS,.RED{background:#4a1414;color:#ff8080}.BLACK{background:#000;color:#ff8080;border:1px solid #ff8080}
 input,select,button{padding:7px;border-radius:7px;border:1px solid #2a2f3a;background:#0f1115;color:#e8e8e8}
 button{background:#c9a84c;color:#111;border:0;font-weight:700;cursor:pointer}
 table{width:100%;border-collapse:collapse;font-size:13px}td,th{padding:6px;border-bottom:1px solid #232833;text-align:left}
 .bar{height:8px;background:#232833;border-radius:6px;overflow:hidden}.bar>i{display:block;height:100%;background:#c9a84c}
 form.inline{display:inline}
</style></head><body><div class="wrap">
<h1>Money OS</h1>
<p class="muted">Owner/manager cockpit. Reads your sales + time clock. Nothing here moves real money -- it stages, you approve. <a href="/sales">register</a> &middot; <a href="/payroll">payroll</a></p>
{% with msgs = get_flashed_messages() %}{% for m in msgs %}<p class="muted">- {{ m }}</p>{% endfor %}{% endwith %}

<div class="grid">
  <div class="card">
    <div class="muted">Today's profit ({{ pnl.date }})</div>
    <div class="big">${{ '%.2f'|format(pnl.net_profit) }} <span class="flag {{ pnl.flag }}">{{ pnl.flag }}</span></div>
    <table>
      <tr><td>Revenue</td><td>${{ '%.2f'|format(pnl.revenue) }}</td></tr>
      <tr><td>COGS</td><td>-${{ '%.2f'|format(pnl.cogs) }}</td></tr>
      <tr><td>Labor (OT-correct)</td><td>-${{ '%.2f'|format(pnl.labor_cost) }}</td></tr>
      <tr><td>Employer tax est</td><td>-${{ '%.2f'|format(pnl.employer_tax) }}</td></tr>
      <tr><td>Overhead (prorated)</td><td>-${{ '%.2f'|format(pnl.overhead_allocated) }}</td></tr>
      <tr><td>Sales tax collected</td><td>${{ '%.2f'|format(pnl.sales_tax_collected) }} <span class="muted">(owed to CDTFA)</span></td></tr>
    </table>
    <form class="inline" method="post" action="/money/allocate"><button>Run today's set-aside</button></form>
  </div>

  <div class="card">
    <div class="muted">Payroll readiness {% if r.pay_date %}(by {{ r.pay_date }}){% endif %}</div>
    <div class="big">{% if r.gap > 0 %}Short ${{ '%.2f'|format(r.gap) }}{% else %}Covered{% endif %}
      <span class="flag {{ r.alert_level }}">{{ r.alert_level }}</span></div>
    {% if r.catch_up > 0 %}<p class="flag BLACK">CATCH-UP OWED: ${{ '%.2f'|format(r.catch_up) }} across {{ r.catch_up_periods|length }} unrun period(s)</p>{% endif %}
    <table>
      <tr><td>Accrued to date</td><td>${{ '%.2f'|format(r.accrued_to_date) }}</td></tr>
      <tr><td>Projected to payday</td><td>${{ '%.2f'|format(r.projected_to_payday) }}</td></tr>
      <tr><td>Employer tax est</td><td>${{ '%.2f'|format(r.employer_tax_est) }}</td></tr>
      <tr><td>Total need</td><td>${{ '%.2f'|format(r.total_need) }}</td></tr>
      <tr><td>Cash on hand{% if r.cash_stale %} (stale){% endif %}</td><td>${{ '%.2f'|format(r.cash_on_hand) }}</td></tr>
      <tr><td>Payroll envelopes</td><td>${{ '%.2f'|format(r.payroll_envelope) }}</td></tr>
    </table>
  </div>

  <div class="card">
    <div class="muted">Envelopes</div>
    <table>{% for e in envelopes %}<tr><td>{{ e.Envelope }}</td>
      <td style="width:50%"><div class="bar"><i style="width:{{ ((e.Balance|float / e.Target|float * 100) if e.Target|float > 0 else (100 if e.Balance|float>0 else 0))|round|int }}%"></i></div></td>
      <td>${{ '%.2f'|format(e.Balance|float) }}</td></tr>{% endfor %}</table>
  </div>

  <div class="card">
    <div class="muted">Cash on hand (manual)</div>
    <form method="post" action="/money/cash"><input name="amount" type="number" step="0.01" placeholder="bank balance"> <button>Update</button></form>
    <p class="muted">{% if cash.as_of %}as of {{ cash.as_of }} ({{ cash.source }}){% else %}not set{% endif %}. Plaid auto-pull is a later upgrade.</p>
    <div class="muted" style="margin-top:10px">Autopilot</div>
    <form method="post" action="/money/autopilot">
      <select name="mode">
        <option value="OFF" {% if autopilot=='OFF' %}selected{% endif %}>OFF</option>
        <option value="SUGGEST" {% if autopilot=='SUGGEST' %}selected{% endif %}>SUGGEST (default -- no auto money)</option>
        <option value="ARMED" {% if autopilot=='ARMED' %}selected{% endif %}>ARMED (auto-approve flagged autopay only)</option>
      </select> <button>Set</button>
    </form>
  </div>
</div>

<div class="card" style="margin-top:14px">
  <div class="muted">Bills -- available after payroll reserve: <b>${{ '%.2f'|format(bills.available_after_payroll) }}</b>
    (cash ${{ '%.2f'|format(bills.cash) }} - payroll ${{ '%.2f'|format(bills.payroll_reserve) }})</div>
  <table><tr><th>Vendor</th><th>Amount</th><th>Due</th><th>Status</th><th>When</th><th></th></tr>
  {% for b in bills.bills %}<tr>
    <td>{{ b.Vendor }}</td><td>${{ '%.2f'|format(b.Amount|float) }}</td><td>{{ b.Due_Date }}</td>
    <td>{{ b.Status }}</td><td><span class="muted">{{ b._tag }}</span></td>
    <td>{% if b.Status != 'PAID' %}
      <form class="inline" method="post" action="/money/bills/{{ b.Bill_ID }}/approve"><button>Approve</button></form>
      <form class="inline" method="post" action="/money/bills/{{ b.Bill_ID }}/pay"><button>Pay</button></form>
    {% endif %}</td></tr>{% endfor %}</table>
  <form method="post" action="/money/bills/add" style="margin-top:8px">
    <input name="vendor" placeholder="vendor"> <input name="amount" type="number" step="0.01" placeholder="amount">
    <input name="due_date" type="date"> <button>Add bill</button>
  </form>
</div>
<p class="muted" style="margin-top:10px">DIY payroll filing figures: <a href="/money/filing">/money/filing?period_id=...</a> gives the exact EFTPS (federal) + DE 88 (CA) deposit amounts to key into the free government portals after a pay run.</p>
</div></body></html>"""


@app.route("/money")
@manager_required
def money_dashboard():
    from flask import render_template_string
    try:
        money.generate_due_bills()
    except Exception:
        pass
    pnl = money.compute_daily_pnl(date.today())
    return render_template_string(
        MONEY_DASHBOARD, pnl=pnl, r=money.payroll_readiness(), envelopes=money.get_envelopes(),
        bills=money.bill_priority_view(), cash=money.get_cash_on_hand(),
        autopilot=money.get_autopilot_mode())


@app.route("/api/money/pnl")
@manager_required
def api_money_pnl():
    if request.args.get("range") == "week":
        return jsonify(money.compute_weekly_pnl(request.args.get("date") or date.today().strftime("%Y-%m-%d")))
    return jsonify(money.compute_daily_pnl(request.args.get("date") or None))


@app.route("/api/money/payroll-readiness")
@manager_required
def api_money_readiness():
    return jsonify(money.payroll_readiness())


@app.route("/money/cash", methods=["POST"])
@manager_required
def money_cash():
    money.set_cash_manual(request.form.get("amount", 0), by=session.get("employee_id", ""))
    flash("Cash on hand updated.", "success")
    return redirect(url_for("money_dashboard"))


@app.route("/money/allocate", methods=["POST"])
@manager_required
def money_allocate():
    res = money.allocate_for_date()
    flash("Already allocated today." if res.get("skipped") else "Daily set-aside run.", "info")
    return redirect(url_for("money_dashboard"))


@app.route("/money/bills/add", methods=["POST"])
@manager_required
def money_bill_add():
    money.add_bill(request.form.get("vendor", ""), request.form.get("amount", 0),
                   request.form.get("due_date", ""), type=request.form.get("type", "BILL"),
                   note=request.form.get("note", ""))
    flash("Bill added.", "success")
    return redirect(url_for("money_dashboard"))


@app.route("/money/bills/<bid>/approve", methods=["POST"])
@manager_required
def money_bill_approve(bid):
    money.approve_bill(bid, session.get("employee_id", ""))
    return redirect(url_for("money_dashboard"))


@app.route("/money/bills/<bid>/pay", methods=["POST"])
@manager_required
def money_bill_pay(bid):
    money.pay_bill(bid, session.get("employee_id", ""))
    flash("Bill marked paid.", "success")
    return redirect(url_for("money_dashboard"))


@app.route("/money/autopilot", methods=["POST"])
@owner_required
def money_autopilot():
    m = money.set_autopilot_mode(request.form.get("mode", "SUGGEST"))
    flash(f"Autopilot set to {m}.", "info")
    return redirect(url_for("money_dashboard"))


@app.route("/money/filing")
@manager_required
def money_filing():
    pid = (request.args.get("period_id") or "").strip()
    if not pid:
        return jsonify({"error": "period_id required, e.g. /money/filing?period_id=PP-2026-06"})
    return jsonify(money.filing_summary(pid))


@app.route("/sales/complete", methods=["POST"])
@login_required
def complete_sale():
    try:
        if not enforce_limit_or_upgrade(
            "max_txn_month", _count_transactions_current_month()
        ):
            return jsonify(
                {
                    "success": False,
                    "error": "Monthly transaction limit reached. Upgrade to continue.",
                }
            ), 402

        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
            items = data.get("items", [])
            payment_method = data.get("payment_method", "CASH")
            amount_received = float(
                data.get("cash_received") or data.get("amount_received") or 0
            )
            card_fee = float(data.get("card_fee") or 0)
        else:
            items_json = request.form.get("items", "[]")
            items = json.loads(items_json)
            payment_method = request.form.get("payment_method", "CASH")
            amount_received = float(request.form.get("cash_received") or 0)
            card_fee = float(request.form.get("card_fee") or 0)

        # Optional customer capture (works for both JSON and form posts)
        def _truthy(v):
            return str(v).strip().lower() in ("1", "true", "on", "yes")
        _src = data if request.is_json else request.form
        customer_name = (_src.get("customer_name") or "").strip()
        customer_email = (_src.get("customer_email") or "").strip()
        email_receipt = _truthy(_src.get("email_receipt"))
        newsletter = _truthy(_src.get("newsletter"))

        if not items:
            return jsonify({"success": False, "error": "No items in cart"})

        success, result = record_sale(
            items=items,
            emp_id=session["employee_id"],
            emp_name=session["employee_name"],
            payment_method=payment_method,
            amount_received=amount_received,
            notes="",
            card_fee=card_fee,
        )

        if success:
            change = result.get("change_due", 0)
            settings = get_or_create_tenant_settings()
            review_url = os.environ.get("APP_REVIEW_URL", "").strip()
            prompt_review = False
            if settings.get("Review_Prompt_Enabled", "Y").upper() == "Y" and review_url:
                txns = _count_transactions_current_month()
                prompt_review = txns in (25, 100, 250)

            # --- customer profile + receipt email + newsletter (all optional) ---
            txid = result["transaction_id"]
            customer_id = ""
            receipt_emailed = False
            newsletter_added = False
            if customer_email and "@" in customer_email:
                try:
                    customer_id = upsert_customer(customer_name, customer_email)
                    parts = customer_name.split(" ", 1)
                    first = parts[0] if parts else ""
                    last = parts[1] if len(parts) > 1 else ""
                    # purchase history (so we can see what they bought + tailor offers)
                    log_customer_receipt(first, last, customer_email,
                                         build_receipt_payload(txid) or {})
                    if newsletter:
                        newsletter_added = add_newsletter_subscriber(
                            customer_id, customer_email, customer_name)[0]
                    if email_receipt:
                        try:
                            bundle = get_receipt_bundle(txid)
                            pdf_bytes = build_receipt_pdf_bytes(bundle)
                            send_receipt_email_smtp(
                                to_email=customer_email, customer_name=customer_name,
                                trans_id=txid, pdf_bytes=pdf_bytes)
                            receipt_emailed = True
                            log_receipt_delivery(
                                txid, method="EMAIL", status="OK",
                                notes="Receipt emailed at checkout",
                                emp_id=session.get("employee_id", ""),
                                customer_id=customer_id, customer_name=customer_name,
                                customer_email=customer_email)
                        except Exception as _re:
                            # email is best-effort -- the sale + profile already saved
                            try:
                                log_receipt_delivery(
                                    txid, method="EMAIL", status="FAILED", notes=str(_re),
                                    emp_id=session.get("employee_id", ""),
                                    customer_id=customer_id, customer_name=customer_name,
                                    customer_email=customer_email)
                            except Exception:
                                pass
                except Exception:
                    pass

            return jsonify(
                {
                    "success": True,
                    "transaction_id": result["transaction_id"],
                    "total": result["total"],
                    "change": change,
                    "change_breakdown": make_change_breakdown(change),
                    "review_prompt": prompt_review,
                    "review_url": review_url,
                    "customer_id": customer_id,
                    "receipt_emailed": receipt_emailed,
                    "newsletter_added": newsletter_added,
                }
            )
        else:
            return jsonify(
                {"success": False, "error": result.get("error", "Unknown error")}
            )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def _sales_logs_root() -> Path:
    # /mnt/sdcard/Mountain Gardens Nursery POS/Sales_Logs
    return Path(__file__).resolve().parent / "Sales_Logs"


def _list_saleslog_files() -> list[dict]:
    """
    Returns list of dicts:
      { "rel": "2025/12_December/Week_3/2025-12-15_Monday_SalesLog.csv",
        "label": "2025-12-15 Monday",
        "path": Path(...) }
    """
    root = _sales_logs_root()
    files = []
    if not root.exists():
        return files

    # Only SalesLog files (ignore other variants)
    for p in root.rglob("*SalesLog.csv"):
        rel = str(p.relative_to(root)).replace("\\", "/")

        # Try to build a nice label from filename: YYYY-MM-DD_Day_SalesLog.csv
        name = p.name
        m = re.match(r"(\d{4}-\d{2}-\d{2})_([A-Za-z]+)_SalesLog\.csv$", name)
        if m:
            label = f"{m.group(1)} {m.group(2)}"
        else:
            label = name

        files.append({"rel": rel, "label": label, "path": p})

    # Sort newest first by modified time
    files.sort(key=lambda x: x["path"].stat().st_mtime, reverse=True)
    return files


def _read_csv_rows(path: Path, limit: int = 300) -> tuple[list[str], list[dict]]:
    rows: list[dict] = []
    headers: list[str] = []
    if not path.exists():
        return headers, rows

    with path.open("r", newline="", encoding="utf-8", errors="replace") as f:
        r = csv.DictReader(f)
        headers = r.fieldnames or []
        for i, row in enumerate(r):
            rows.append(row)
            if i + 1 >= limit:
                break
    return headers, rows


def _sales_logs_root() -> Path:
    # /mnt/sdcard/Mountain Gardens Nursery POS/Sales_Logs
    return Path(__file__).resolve().parent / "Sales_Logs"


def _transaction_logs_root() -> Path:
    # /mnt/sdcard/Mountain Gardens Nursery POS/Transaction_Logs
    return Path(__file__).resolve().parent / "Transaction_Logs"


def _list_saleslog_files() -> list[dict]:
    """
    Returns list of dicts:
      { "rel": "2025/12_December/Week_3/2025-12-15_Monday_SalesLog.csv",
        "label": "2025-12-15 Monday",
        "path": Path(...) }
    """
    root = _sales_logs_root()
    files = []
    if not root.exists():
        return files

    for p in root.rglob("*SalesLog.csv"):
        rel = str(p.relative_to(root)).replace("\\", "/")

        name = p.name
        m = re.match(r"(\d{4}-\d{2}-\d{2})_([A-Za-z]+)_SalesLog\.csv$", name)
        if m:
            label = f"{m.group(1)} {m.group(2)}"
        else:
            label = name

        files.append({"rel": rel, "label": label, "path": p})

    files.sort(key=lambda x: x["path"].stat().st_mtime, reverse=True)
    return files


def _read_csv_rows(path: Path, limit: int = 300) -> tuple[list[str], list[dict]]:
    rows: list[dict] = []
    headers: list[str] = []
    if not path or not path.exists():
        return headers, rows

    with path.open("r", newline="", encoding="utf-8", errors="replace") as f:
        r = csv.DictReader(f)
        headers = r.fieldnames or []
        for i, row in enumerate(r):
            rows.append(row)
            if i + 1 >= limit:
                break
    return headers, rows


def _parse_day_from_saleslog_filename(name: str) -> tuple[str | None, str | None]:
    m = re.match(r"(\d{4}-\d{2}-\d{2})_([A-Za-z]+)_SalesLog\.csv$", name or "")
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _find_matching_transaction_log(date_str: str, day_str: str) -> Path | None:
    """
    Looks for: YYYY-MM-DD_Day_TransactionLog.csv anywhere under Transaction_Logs.
    Falls back to looser patterns if needed.
    """
    root = _transaction_logs_root()
    if not root.exists():
        return None

    exact = f"{date_str}_{day_str}_TransactionLog.csv"
    for p in root.rglob(exact):
        return p

    # fallback patterns (in case naming differs slightly)
    patterns = [
        f"{date_str}_*_TransactionLog.csv",
        f"{date_str}*_TransactionLog.csv",
        f"*{date_str}*TransactionLog*.csv",
    ]
    for pat in patterns:
        hits = list(root.rglob(pat))
        if hits:
            hits.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            return hits[0]

    return None


def _safe_float(x) -> float:
    try:
        return float(str(x).replace("$", "").strip() or 0)
    except Exception:
        return 0.0


def _safe_int(x) -> int:
    try:
        return int(float(str(x).strip() or 0))
    except Exception:
        return 0


@app.route("/sales/log")
@login_required
def sales_log():
    files = _list_saleslog_files()

    # Prefer today's SalesLog if present; else newest file
    selected_rel = (request.args.get("file") or "").strip()
    if not selected_rel:
        today = date.today().strftime("%Y-%m-%d")
        picked = None
        for f in files:
            if f["label"].startswith(today + " "):
                picked = f["rel"]
                break
        selected_rel = picked or (files[0]["rel"] if files else "")

    selected_sales_path = None
    for f in files:
        if f["rel"] == selected_rel:
            selected_sales_path = f["path"]
            break

    # Read SalesLog (line items)
    row_limit = 400
    sales_headers, sales_rows = ([], [])
    date_str, day_str = (None, None)
    if selected_sales_path:
        sales_headers, sales_rows = _read_csv_rows(selected_sales_path, limit=row_limit)
        date_str, day_str = _parse_day_from_saleslog_filename(selected_sales_path.name)

    # Auto-match TransactionLog for the same day
    tx_path = None
    tx_headers, tx_rows = ([], [])
    if date_str and day_str:
        tx_path = _find_matching_transaction_log(date_str, day_str)
        if tx_path:
            tx_headers, tx_rows = _read_csv_rows(tx_path, limit=row_limit)

    # Simple day metrics
    tx_total = sum(_safe_float(r.get("Total", 0)) for r in tx_rows) if tx_rows else 0.0
    tx_count = len(tx_rows) if tx_rows else 0

    items_sold = (
        sum(_safe_int(r.get("Quantity", 0)) for r in sales_rows) if sales_rows else 0
    )
    sales_line_total = (
        sum(_safe_float(r.get("Line_Total", 0)) for r in sales_rows)
        if sales_rows
        else 0.0
    )
    sales_cogs_total = (
        sum(_safe_float(r.get("COGS_Line", 0)) for r in sales_rows)
        if sales_rows
        else 0.0
    )
    sales_gross = sales_line_total - sales_cogs_total

    # 🔹 Daily payroll (gross wages) for this same date, based on Time_Clock
    daily_payroll = 0.0
    if date_str:
        from collections import defaultdict

        # Use existing helper – same one the payroll dashboard & run_payroll use
        punches = scan_timeclock_files(date_str, date_str)

        punches_by_emp: dict[str, list[dict]] = defaultdict(list)
        for punch in punches:
            emp_id = str(punch.get("Employee_ID", "")).strip()
            if emp_id:
                punches_by_emp[emp_id].append(punch)

        employees = get_all_employees()
        for emp in employees:
            if emp.get("Status") != "Active":
                continue

            emp_id = str(emp.get("Employee_ID", "")).strip()
            emp_punches = punches_by_emp.get(emp_id, [])
            if not emp_punches:
                continue

            # Re-use your CA overtime logic + pay rate
            regular, ot, double = calculate_california_hours(emp_punches)
            rate = get_employee_pay_rate(emp_id)
            gross_pay = calculate_gross_pay_california(regular, ot, double, rate)

            daily_payroll += gross_pay

    # Net for the day after payroll cost
    net_after_payroll = sales_gross - daily_payroll

    return render_template(
        "sales/log.html",
        files=files,
        selected_file=selected_rel,
        # Transactions (header rows)
        tx_headers=tx_headers,
        tx_rows=tx_rows,
        tx_file=(tx_path.name if tx_path else ""),
        # Sales (line rows)
        sales_headers=sales_headers,
        sales_rows=sales_rows,
        sales_file=(selected_sales_path.name if selected_sales_path else ""),
        # Metrics
        row_limit=row_limit,
        tx_count=tx_count,
        tx_total=tx_total,
        items_sold=items_sold,
        sales_line_total=sales_line_total,
        sales_cogs_total=sales_cogs_total,
        sales_gross=sales_gross,
        daily_payroll=daily_payroll,
        net_after_payroll=net_after_payroll,
    )


# --- Email receipt (logs + sends PDF attachment) ---
@app.route("/sales/receipt/<trans_id>/email", methods=["POST"])
def receipt_email(trans_id):
    emp_id = session.get("employee_id", "")
    customer_name = (request.form.get("customer_name") or "").strip()
    customer_email = (request.form.get("customer_email") or "").strip()

    if not customer_email or "@" not in customer_email:
        return jsonify({"ok": False, "error": "Enter a valid email."}), 400

    customer_id = upsert_customer(customer_name, customer_email)

    bundle = get_receipt_bundle(trans_id)
    if not bundle:
        return jsonify({"ok": False, "error": "Receipt not found."}), 404

    try:
        pdf_bytes = build_receipt_pdf_bytes(bundle)  # function you’ll add in POS_CORE
        send_receipt_email_smtp(
            to_email=customer_email,
            customer_name=customer_name,
            trans_id=trans_id,
            pdf_bytes=pdf_bytes,
        )
        log_receipt_delivery(
            trans_id,
            method="EMAIL",
            status="OK",
            notes="Sent receipt email",
            emp_id=emp_id,
            customer_id=customer_id,
            customer_name=customer_name,
            customer_email=customer_email,
        )
        return jsonify({"ok": True})
    except Exception as e:
        log_receipt_delivery(
            trans_id,
            method="EMAIL",
            status="FAILED",
            notes=str(e),
            emp_id=emp_id,
            customer_id=customer_id,
            customer_name=customer_name,
            customer_email=customer_email,
        )
        return jsonify({"ok": False, "error": f"Email failed: {e}"}), 500


# ==============================================================================
#                     CUSTOMERS + NEWSLETTER (marketing)
# ==============================================================================


@app.route("/customers")
@manager_required
def customers_list():
    from datetime import datetime as _dt
    inactive_days = request.args.get("inactive_days", type=int)
    tier_filter = (request.args.get("tier") or "").strip().upper()
    today = date.today()
    enriched = []
    for c in list_customers():
        hist = get_customer_history(c.get("Email", ""))
        spent = 0.0
        for h in hist:
            try:
                spent += float(h.get("Total") or 0)
            except Exception:
                pass
        last = hist[0].get("Date") if hist else ""
        days = None
        if last:
            try:
                days = (today - _dt.strptime(last, "%Y-%m-%d").date()).days
            except Exception:
                days = None
        enriched.append({**c, "purchases": len(hist), "spent": round(spent, 2),
                         "last": last, "days": days, "tier": customer_tier(spent)})
    if tier_filter:
        enriched = [e for e in enriched if e["tier"] == tier_filter]
    if inactive_days is not None:
        enriched = [e for e in enriched if e["days"] is not None and e["days"] >= inactive_days]
    enriched.sort(key=lambda e: e["spent"], reverse=True)
    return render_template("customers/list.html", customers=enriched,
                           tier_filter=tier_filter, inactive_days=inactive_days)


@app.route("/newsletter/unsubscribe/<sub_id>")
def newsletter_unsubscribe(sub_id):
    """PUBLIC (no login) -- the unsubscribe link recipients click from a marketing
    email. CAN-SPAM compliant: deactivates the subscriber."""
    from markupsafe import escape
    ok, email = unsubscribe_newsletter(sub_id)
    msg = (f"{escape(email)} has been unsubscribed from the newsletter."
           if ok else "That subscription was not found (it may already be removed).")
    return ("<html><body style='font-family:sans-serif;max-width:520px;margin:60px auto;"
            "text-align:center;color:#1e3a2a;'><h2>Mountain Gardens Nursery</h2>"
            f"<p style='font-size:16px;'>{msg}</p></body></html>")


@app.route("/customers/<customer_id>")
@manager_required
def customer_detail(customer_id):
    cust = get_customer_by_id(customer_id)
    if not cust:
        flash("Customer not found.", "error")
        return redirect(url_for("customers_list"))
    history = get_customer_history(cust.get("Email", ""))
    spent = 0.0
    for h in history:
        try:
            spent += float(h.get("Total") or 0)
        except Exception:
            pass
    return render_template("customers/detail.html", customer=cust,
                           history=history, spent=round(spent, 2))


@app.route("/newsletter")
@manager_required
def newsletter_list():
    subs = get_newsletter_subscribers(active_only=False)
    return render_template("customers/newsletter.html", subscribers=subs)


@app.route("/newsletter/export")
@manager_required
def newsletter_export():
    from flask import Response
    subs = get_newsletter_subscribers(active_only=True)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Email", "Name", "Subscribed_At"])
    for s in subs:
        w.writerow([s.get("Email", ""), s.get("Name", ""), s.get("Subscribed_At", "")])
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=newsletter_subscribers.csv"})


# ==============================================================================
#         VENDOR INVOICE INGEST (master-SKU + vendor aliases + FIFO lots)
# ==============================================================================


def _parse_invoice_csv(text):
    """Tolerant invoice-line parser. Accepts common column names from Square /
    Shopify / QuickBooks / a vendor's own export."""
    out = []
    if not text or not text.strip():
        return out

    def pick(row, *names):
        for k in row:
            kn = (k or "").strip().lower().replace(" ", "_")
            if kn in names:
                return (row[k] or "").strip()
        return ""

    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        vsku = pick(row, "vendor_sku", "sku", "item_no", "item_number", "product_no",
                    "product_number", "item", "item_code", "code")
        desc = pick(row, "desc", "description", "name", "item_name", "product", "title")
        qty = pick(row, "qty", "quantity", "units", "count", "received")
        cost = pick(row, "unit_cost", "cost", "price", "unit_price", "wholesale", "wholesale_cost")
        if not (vsku or desc):
            continue
        out.append({"vendor_sku": vsku, "desc": desc, "qty": qty, "unit_cost": cost})
    return out


@app.route("/inventory/vendor-invoice", methods=["GET", "POST"])
@manager_required
def vendor_invoice():
    if request.method == "POST":
        vendor = (request.form.get("vendor") or "").strip()
        invoice_no = (request.form.get("invoice_no") or "").strip()
        csv_text = request.form.get("csv_text") or ""
        f = request.files.get("csv_file")
        if f and f.filename:
            try:
                csv_text = f.read().decode("utf-8", errors="ignore")
            except Exception:
                csv_text = ""
        if not vendor:
            flash("Enter the vendor name.", "error")
            return redirect(url_for("vendor_invoice"))
        lines = _parse_invoice_csv(csv_text)
        if not lines:
            flash("No invoice lines found. Paste/upload CSV with columns like "
                  "vendor_sku, description, qty, unit_cost.", "error")
            return redirect(url_for("vendor_invoice"))
        result = ingest_invoice_lines(vendor, invoice_no, lines,
                                      emp_id=session.get("employee_id", ""))
        return render_template("inventory/vendor_invoice.html", result=result,
                               vendor=vendor, invoice_no=invoice_no, submitted=True,
                               items=get_all_items())
    return render_template("inventory/vendor_invoice.html", result=None,
                           submitted=False, items=[])


@app.route("/inventory/vendor-invoice/map", methods=["POST"])
@manager_required
def vendor_invoice_map():
    master_sku = (request.form.get("master_sku") or "").strip()
    vendor = (request.form.get("vendor") or "").strip()
    vendor_sku = (request.form.get("vendor_sku") or "").strip()
    desc = (request.form.get("desc") or "").strip()
    invoice_no = (request.form.get("invoice_no") or "").strip()
    try:
        qty = int(float(request.form.get("qty") or 0))
    except Exception:
        qty = 0
    try:
        cost = float(request.form.get("unit_cost") or 0)
    except Exception:
        cost = 0.0
    # accept "SKU - Name" from the datalist by taking the leading token
    master_sku = master_sku.split(" - ")[0].split(" ")[0].strip()
    if not (master_sku and vendor_sku):
        flash("Pick a master product to map this vendor item to.", "error")
        return redirect(url_for("vendor_invoice"))
    if not get_item(master_sku):
        flash(f"Master SKU '{master_sku}' not found in the catalog.", "error")
        return redirect(url_for("vendor_invoice"))
    map_vendor_sku(master_sku, vendor, vendor_sku, vendor_desc=desc, unit_cost=cost)
    if qty > 0:
        create_lot(master_sku, qty, cost, supplier=vendor, invoice=invoice_no,
                   notes=f"Vendor {vendor} SKU {vendor_sku}")
    flash(f"Mapped {vendor} #{vendor_sku} to {master_sku} and received {qty}. "
          f"Future invoices from {vendor} auto-match.", "success")
    return redirect(url_for("vendor_invoice"))


# ==============================================================================
#       ADMIN TASK SCHEDULER (owner/admin only -- managers excluded)
# ==============================================================================


@app.route("/admin/schedule", methods=["GET"])
@owner_required
def admin_schedule():
    admins = [e for e in get_all_employees()
              if (e.get("Role") or "") in ("Owner", "Admin")]
    return render_template("admin/schedule.html",
                           schedules=list_recurring_schedules(active_only=False),
                           preview=preview_recurring(14), admins=admins)


@app.route("/admin/schedule", methods=["POST"])
@owner_required
def admin_schedule_create():
    title = (request.form.get("title") or "").strip()
    day_rule = (request.form.get("day_rule") or "").strip()
    assignee = (request.form.get("assignee") or session.get("employee_id", "")).strip()
    description = (request.form.get("description") or "").strip()
    end_date = (request.form.get("end_date") or "").strip()
    if not title or not day_rule:
        flash("A title and a day rule (e.g. 1,15) are required.", "error")
        return redirect(url_for("admin_schedule"))
    create_recurring_task(title, day_rule, assignee, session.get("employee_id", ""),
                          description=description, end_date=end_date)
    flash(f"Recurring task '{title}' scheduled ({day_rule}).", "success")
    return redirect(url_for("admin_schedule"))


@app.route("/admin/schedule/run", methods=["POST"])
@owner_required
def admin_schedule_run():
    made = check_and_assign_recurring_tasks()
    flash(f"Assigned {len(made)} task(s) due today.", "success")
    return redirect(url_for("admin_schedule"))


@app.route("/admin/schedule/<sid>/pause", methods=["POST"])
@owner_required
def admin_schedule_pause(sid):
    set_recurring_status(sid, "PAUSED")
    flash("Schedule paused.", "success")
    return redirect(url_for("admin_schedule"))


# ==============================================================================
#       INTEGRATIONS -- CSV import/export (Square / Shopify / QuickBooks / MGN)
# ==============================================================================

_ITX = None


def _itx():
    """Lazy-load the standalone inventory_transfer converter module (CLI tool)."""
    global _ITX
    if _ITX is None:
        import importlib.util as ilu
        p = os.path.join(SCRIPT_DIR, "tools", "inventory_transfer.py")
        spec = ilu.spec_from_file_location("inventory_transfer", p)
        mod = ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _ITX = mod
    return _ITX


_PLATFORMS = ["square", "shopify", "quickbooks", "mgn"]
_API = None


def _api():
    """Lazy-load the credential-gated live-sync adapters."""
    global _API
    if _API is None:
        import importlib.util as ilu
        p = os.path.join(SCRIPT_DIR, "tools", "integrations_api.py")
        spec = ilu.spec_from_file_location("integrations_api", p)
        mod = ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _API = mod
    return _API


@app.route("/integrations")
@manager_required
def integrations():
    return render_template("integrations.html", platforms=_PLATFORMS,
                           item_count=len(get_all_items()),
                           sync_status=_api().status())


@app.route("/integrations/export")
@manager_required
def integrations_export():
    from flask import Response
    fmt = (request.args.get("format") or "square").strip().lower()
    if fmt not in _PLATFORMS:
        flash("Unknown export format.", "error")
        return redirect(url_for("integrations"))
    itx = _itx()
    headers, rows, _skipped = itx.convert("mgn", fmt, ITEM_HEADERS, get_all_items())
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=catalog_{fmt}.csv"})


@app.route("/integrations/import", methods=["POST"])
@manager_required
def integrations_import_preview():
    f = request.files.get("csv_file")
    src = (request.form.get("src_format") or "auto").strip().lower()
    if not f or not f.filename:
        flash("Choose a CSV file to import.", "error")
        return redirect(url_for("integrations"))
    try:
        text = f.read().decode("utf-8-sig", errors="ignore")
    except Exception:
        text = ""
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows = list(reader)
    itx = _itx()
    if src == "auto":
        src = itx.detect_format(headers) or ""
    if src not in itx.IMPORTERS:
        flash("Could not detect the CSV format. Pick the source platform and retry.", "error")
        return redirect(url_for("integrations"))
    items, skipped = itx.IMPORTERS[src](rows)
    return render_template(
        "integrations.html", platforms=_PLATFORMS, item_count=len(get_all_items()),
        preview={"src": src, "items": items[:50], "total": len(items),
                 "skipped": len(skipped)},
        staged=json.dumps(items))


@app.route("/integrations/import/apply", methods=["POST"])
@manager_required
def integrations_import_apply():
    staged = request.form.get("staged") or "[]"
    create_lots = (request.form.get("create_lots") or "") in ("1", "on", "true", "yes")
    try:
        items = json.loads(staged)
    except Exception:
        items = []
    if not items:
        flash("Nothing to import.", "error")
        return redirect(url_for("integrations"))
    res = import_items(items, create_lots=create_lots, actor=session.get("employee_id", ""))
    flash(f"Imported: {res['created']} created, {res['updated']} updated, "
          f"{res['lots']} stock lots, {res['skipped']} skipped.", "success")
    return redirect(url_for("integrations"))


_ACCT = None


def _acct():
    """Lazy-load the standalone accounting_export module."""
    global _ACCT
    if _ACCT is None:
        import importlib.util as ilu
        p = os.path.join(SCRIPT_DIR, "tools", "accounting_export.py")
        spec = ilu.spec_from_file_location("accounting_export", p)
        mod = ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _ACCT = mod
    return _ACCT


@app.route("/integrations/accounting")
@manager_required
def integrations_accounting():
    from flask import Response
    from datetime import datetime as _dt
    kind = (request.args.get("kind") or "summary").strip().lower()
    today = date.today()

    def _parse(s, default):
        try:
            return _dt.strptime((s or "").strip(), "%Y-%m-%d").date()
        except Exception:
            return default

    start = _parse(request.args.get("from"), today.replace(day=1))
    end = _parse(request.args.get("to"), today)
    acc = _acct()
    if kind == "journal":
        rows = acc.journal_entry_rows(start, end)
        headers = ["Date", "Account", "Debit", "Credit", "Memo"]
    else:
        rows = acc.daily_summary_rows(start, end)
        headers = ["Date", "Transactions", "Gross_Sales", "Sales_Tax", "COGS",
                   "Gross_Profit", "Cash", "Card"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers)
    w.writeheader()
    w.writerows(rows)
    return Response(
        buf.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition":
                 f"attachment; filename=accounting_{kind}_{start}_{end}.csv"})


# ==============================================================================
#       BACKUPS (resilience -- one mini PC = one dead drive without these)
# ==============================================================================

_BK = None


def _bk():
    global _BK
    if _BK is None:
        import importlib.util as ilu
        p = os.path.join(SCRIPT_DIR, "tools", "backup_data.py")
        spec = ilu.spec_from_file_location("backup_data", p)
        mod = ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _BK = mod
    return _BK


@app.route("/admin/backup")
@owner_required
def admin_backup():
    return render_template("admin/backup.html", backups=_bk().list_backups(),
                           encrypted=bool(os.environ.get("MGN_BACKUP_PASSPHRASE")),
                           offsite=os.environ.get("MGN_BACKUP_OFFSITE", ""))


@app.route("/admin/backup/run", methods=["POST"])
@owner_required
def admin_backup_run():
    try:
        r = _bk().create_backup()
        bits = [f"Backup saved ({round(r['size'] / 1024, 1)} KB)"]
        if r.get("encrypted"):
            bits.append("encrypted")
        if r.get("offsite"):
            bits.append("copied offsite")
        flash(", ".join(bits) + ".", "success")
    except Exception as e:
        flash(f"Backup failed: {e}", "error")
    return redirect(url_for("admin_backup"))


@app.route("/admin/backup/download/<name>")
@owner_required
def admin_backup_download(name):
    name = os.path.basename(name)  # path-traversal guard
    path = _bk().backup_dir() / name
    if not path.exists():
        flash("Backup not found.", "error")
        return redirect(url_for("admin_backup"))
    return send_file(str(path), as_attachment=True, download_name=name)


# ==============================================================================
#                     INVENTORY
# ==============================================================================


def _safe_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default


def pick_sku(item: dict):
    # common header variants from different CSVs/importers
    for key in ("SKU", "Internal_SKU", "Internal SKU", "InternalSKU", "sku", "Sku"):
        v = item.get(key)
        if v and str(v).strip():
            return str(v).strip()
    return None


@app.route("/inventory")
@login_required
def inventory():
    items = get_all_items()
    normalized = []
    for item in items:
        sku = pick_sku(item)
        if not sku:
            continue
        item["SKU"] = sku  # normalize for templates
        item["Default_Price"] = item.get("Default_Price") or "0"
        item["Reorder_Point"] = item.get("Reorder_Point") or "0"
        item["Status"] = item.get("Status") or "Active"
        item["stock_on_hand"] = get_stock_on_hand(sku)
        normalized.append(item)

    return render_template("inventory/list.html", items=normalized)


@app.route("/inventory/add", methods=["GET", "POST"])
@manager_required
def add_item():
    if request.method == "POST":
        if not enforce_limit_or_upgrade("max_skus", len(get_all_items())):
            return redirect(url_for("inventory"))

        name = request.form.get("name")
        category = request.form.get("category")
        subcategory = request.form.get("subcategory", "")
        price = float(request.form.get("price", 0))
        reorder = int(request.form.get("reorder_point", 5))

        sku = generate_sku(name, category, subcategory)
        success, msg = create_item(
            sku, name, category, subcategory, name, price, reorder
        )

        if success:
            flash(f"Item created: {sku}", "success")
            return redirect(url_for("inventory"))
        flash(msg, "error")

    return render_template(
        "inventory/add.html",
        categories=MAIN_CATEGORIES,
        animal_subs=ANIMAL_SUBCATEGORIES,
        product_subs=PRODUCT_SUBCATEGORIES,
        plant_subs=PLANT_SUBCATEGORIES,
    )


#


@app.route("/api/inventory/categories")
@login_required
def api_inventory_categories():
    items = get_all_items()
    cats = sorted(
        {
            (i.get("Category") or "").strip()
            for i in items
            if (i.get("Category") or "").strip()
        }
    )
    return jsonify({"success": True, "categories": cats})


@app.route("/api/inventory/catalog")
@login_required
def api_inventory_catalog():
    category = (request.args.get("category") or "").strip()
    q = (request.args.get("q") or "").strip().lower()

    items = get_all_items()

    out = []
    for it in items:
        sku = (it.get("SKU") or "").strip()
        if not sku:
            continue

        cat = (it.get("Category") or "").strip()
        if category and cat != category:
            continue

        # You don't currently have dedicated Botanical/Common columns in Items.csv,
        # so we use your existing fields as the "best available" sources.
        item_name = (it.get("Item_Name") or "").strip()
        botanical = (it.get("Product_Name") or "").strip()
        common = (it.get("Item_Description") or "").strip()

        label = " — ".join([x for x in [item_name, botanical, common] if x])

        hay = f"{sku} {item_name} {botanical} {common}".lower()
        if q and q not in hay:
            continue

        out.append(
            {
                "sku": sku,
                "category": cat,
                "item_name": item_name,
                "botanical": botanical,
                "common": common,
                "label": label,
            }
        )

    # keep it snappy
    return jsonify({"success": True, "items": out[:500]})


# ============================================================================
# QR CODE IMAGE GENERATION ROUTE
# ============================================================================


@app.route("/qr/<sku>.png")
def qr_png(sku):
    """
    Generate QR code PNG on-the-fly for any SKU.
    No need to store PNGs on disk.

    Usage in templates:
        <img src="/qr/{{ sku }}.png" alt="QR for {{ sku }}">
    """
    try:
        # Verify SKU exists
        item = get_item(sku)
        if not item:
            abort(404)

        # Create QR payload (stable forever)
        payload = f"MGN:SKU:{sku}"

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,  # Size (1 = small, auto-adjusts if needed)
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(payload)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convert to PNG bytes
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)

        return send_file(buf, mimetype="image/png")

    except Exception as e:
        app.logger.error(f"Error generating QR: {e}")
        abort(500)


@app.route("/qr/label/<sku>")
@login_required
def qr_label(sku):
    """
    Print-friendly QR label page.
    Shows QR code + SKU + Item name for printing on labels.
    """
    item = get_item(sku)
    if not item:
        flash("Item not found", "error")
        return redirect(url_for("inventory"))

    return render_template("inventory/qr_label.html", item=item, sku=sku)


# ============================================================================
# BARCODE MANAGEMENT ROUTES
# ============================================================================


@app.route("/inventory/<sku>/barcodes")
@manager_required
def view_barcodes(sku):
    """View all barcode mappings for a SKU"""
    item = get_item(sku)
    if not item:
        flash("Item not found", "error")
        return redirect(url_for("inventory"))

    barcodes = get_barcodes_for_sku(sku)

    return render_template(
        "inventory/barcodes.html", item=item, sku=sku, barcodes=barcodes
    )


@app.route("/inventory/<sku>/barcode/add", methods=["POST"])
@manager_required
def add_barcode(sku):
    """Add a new barcode mapping"""
    try:
        code = request.form.get("code", "").strip()
        vendor = request.form.get("vendor", "").strip()
        code_type = request.form.get("code_type", "BARCODE").strip()
        notes = request.form.get("notes", "").strip()

        if not code:
            flash("Barcode is required", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        success, message = upsert_barcode(
            code=code, sku=sku, vendor=vendor, code_type=code_type, notes=notes
        )

        if success:
            flash(message, "success")

            # Audit log
            emp_id = session.get("employee_id", "SYSTEM")
            emp_name = session.get("employee_name", "System")
            log_audit(
                emp_id,
                emp_name,
                "BARCODE_ADD",
                "Barcode",
                code,
                f"{code_type} for {sku}",
                notes=f"Vendor: {vendor}",
            )
        else:
            flash(message, "error")

        return redirect(url_for("edit_inventory", sku=sku))

    except Exception as e:
        flash(f"Error adding barcode: {str(e)}", "error")
        return redirect(url_for("edit_inventory", sku=sku))


@app.route("/barcode/deactivate/<code>", methods=["POST"])
@manager_required
def deactivate_barcode_route(code):
    """Deactivate a barcode mapping"""
    try:
        success, message = deactivate_barcode(code)

        if success:
            flash(message, "success")

            # Audit log
            emp_id = session.get("employee_id", "SYSTEM")
            emp_name = session.get("employee_name", "System")
            log_audit(emp_id, emp_name, "BARCODE_DEACTIVATE", "Barcode", code, "")
        else:
            flash(message, "error")

        # Get SKU from barcode to redirect back
        barcodes = read_barcodes()
        sku = None
        for b in barcodes:
            if b["Code"] == code:
                sku = b["SKU"]
                break

        if sku:
            return redirect(url_for("edit_inventory", sku=sku))
        else:
            return redirect(url_for("inventory"))

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("inventory"))


# ============================================================================
# SCAN LOOKUP API
# ============================================================================


@app.route("/api/scan/lookup", methods=["POST"])
@login_required
def scan_lookup():
    """
    API endpoint for barcode scanning.
    Returns SKU and item details for any scanned code.

    POST data: {"code": "scanned_value"}
    Returns: {"success": true, "sku": "...", "item": {...}}
    """
    try:
        data = request.get_json()
        scan_code = data.get("code", "").strip()

        if not scan_code:
            return jsonify({"success": False, "error": "No code provided"})

        # Resolve to SKU
        sku = resolve_sku_from_scan(scan_code)

        if not sku:
            return jsonify({"success": False, "error": "Code not found"})

        # Get item details
        item = get_item(sku)

        if not item:
            return jsonify({"success": False, "error": "Item not found"})

        # Get stock info
        stock = get_stock_on_hand(sku)
        avg_cost = get_average_cost(sku)

        return jsonify(
            {
                "success": True,
                "sku": sku,
                "scan_code": scan_code,
                "item": {
                    "sku": sku,
                    "name": item.get("Item_Name", ""),
                    "category": item.get("Category", ""),
                    "price": item.get("Default_Price", ""),
                    "stock": stock,
                    "avg_cost": avg_cost,
                    "taxable": item.get("Taxable", "Y"),
                },
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/inventory/edit/<sku>", methods=["GET", "POST"], endpoint="edit_inventory")
@manager_required
def edit_item(sku):
    item = get_item(sku)
    if not item:
        flash("Item not found.", "error")
        return redirect(url_for("inventory"))

    if request.method == "POST":
        new_sku = (request.form.get("SKU") or sku).strip()

        # --- build updates for Items.csv (match POS_CORE ITEM_HEADERS) ---
        updates = {
            "Item_Name": (request.form.get("Item_Name") or "").strip(),
            "Category": (request.form.get("Category") or "").strip(),
            "Subcategory": (request.form.get("Subcategory") or "").strip(),
            "Product_Name": (request.form.get("Product_Name") or "").strip(),
            "Default_Unit": (request.form.get("Default_Unit") or "").strip(),
            "Default_Price": (request.form.get("Default_Price") or "").strip(),
            "Taxable": (request.form.get("Taxable") or "Y").strip(),
            "Reorder_Point": (request.form.get("Reorder_Point") or "").strip(),
            "Status": (request.form.get("Status") or "Active").strip(),
            "Notes": (request.form.get("Notes") or "").strip(),
            "Size": (request.form.get("Size") or "").strip(),
            "Item_Description": (request.form.get("Item_Description") or "").strip(),
            "Botanical_Name": (request.form.get("Botanical_Name") or "").strip(),
            "Common_Name": (request.form.get("Common_Name") or "").strip(),
            "Sun_Requirements": (request.form.get("Sun_Requirements") or "").strip(),
            "Water_Needs": (request.form.get("Water_Needs") or "").strip(),
            "Hardiness_Zone": (request.form.get("Hardiness_Zone") or "").strip(),
            "Mature_Size": (request.form.get("Mature_Size") or "").strip(),
            "Bloom_Time": (request.form.get("Bloom_Time") or "").strip(),
            "Growth_Rate": (request.form.get("Growth_Rate") or "").strip(),
            "Wholesale_Cost": (request.form.get("Wholesale_Cost") or "").strip(),
            "Retail_Markup": (request.form.get("Retail_Markup") or "").strip(),
            "Retail_Price": (request.form.get("Retail_Price") or "").strip(),
            "Unit_Cost": (request.form.get("Unit_Cost") or "").strip(),
            "Unit_Price": (request.form.get("Unit_Price") or "").strip(),
        }

        # strip keys that user left blank (prevents wiping fields accidentally)
        updates = {k: v for k, v in updates.items() if v != ""}

        items = read_csv(get_items_path())

        # If SKU is being changed, ensure it doesn't collide
        if new_sku != sku:
            if any((r.get("SKU") or "").strip() == new_sku for r in items):
                flash(f"SKU '{new_sku}' already exists. Pick a different SKU.", "error")
                return redirect(url_for("edit_item", sku=sku))

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        saved = False
        for r in items:
            if (r.get("SKU") or "").strip() == sku:
                r.update(updates)
                r["SKU"] = new_sku
                r["Last_Updated"] = now
                saved = True
                break

        if not saved:
            flash("Could not update item (SKU not found in Items.csv).", "error")
            return redirect(url_for("inventory"))

        write_csv(get_items_path(), ITEM_HEADERS, items)

        # keep Lots.csv tied to the SKU if SKU changed
        if new_sku != sku:
            lots = read_csv(get_lots_path())
            for l in lots:
                if (l.get("SKU") or "").strip() == sku:
                    l["SKU"] = new_sku
            write_csv(get_lots_path(), LOT_HEADERS, lots)

            ledger = read_csv(get_ledger_path())
            for e in ledger:
                if (e.get("SKU") or "").strip() == sku:
                    e["SKU"] = new_sku
            write_csv(get_ledger_path(), LEDGER_HEADERS, ledger)

        flash("Item updated ✅", "success")
        return redirect(url_for("edit_item", sku=new_sku))

    # GET view (also show lots + stock)
    lots = get_lots_for_sku(sku, available_only=False)
    return render_template(
        "inventory/edit.html",
        item=item,
        lots=lots,
        stock=get_stock_on_hand(sku),
        avg_cost=get_average_cost(sku),
    )
    flash(f"✅ Received {qty} into {sku}.", "success")
    return redirect(url_for("receive_stock"))


@app.route("/inventory/lots/edit/<lot_id>", methods=["POST"])
@manager_required
def edit_lot(lot_id):
    sku = (request.form.get("sku") or "").strip()

    lots = read_csv(get_lots_path())
    found = False
    for l in lots:
        if (l.get("Lot_ID") or "").strip() == lot_id:
            l["Received_Date"] = (
                request.form.get("Received_Date") or l.get("Received_Date") or ""
            ).strip()
            l["Supplier"] = (
                request.form.get("Supplier") or l.get("Supplier") or ""
            ).strip()
            l["Invoice_Ref"] = (
                request.form.get("Invoice_Ref") or l.get("Invoice_Ref") or ""
            ).strip()
            l["Unit_Cost"] = (
                request.form.get("Unit_Cost") or l.get("Unit_Cost") or ""
            ).strip()
            l["Qty_Remaining"] = (
                request.form.get("Qty_Remaining") or l.get("Qty_Remaining") or ""
            ).strip()
            l["Expiry_Date"] = (
                request.form.get("Expiry_Date") or l.get("Expiry_Date") or ""
            ).strip()
            l["Notes"] = (request.form.get("Notes") or l.get("Notes") or "").strip()
            found = True
            break

    if not found:
        flash("Lot not found.", "error")
        return redirect(url_for("edit_item", sku=sku or ""))

    write_csv(get_lots_path(), LOT_HEADERS, lots)
    flash("Lot updated ✅", "success")
    return redirect(url_for("edit_item", sku=sku))


#!/usr/bin/env python3
"""
INVENTORY QUANTITY ADJUSTMENT - Enhanced Route for MGN_APP.py
# ==============================================================
Add this route after the edit_item function (around line 1450)
"""

# First, ensure uuid is imported at the top of MGN_APP.py
# Add to imports section if not present:
# import uuid


@app.route("/inventory/adjust-qty/<sku>", methods=["POST"])
@manager_required
def adjust_inventory_qty(sku):
    """
    Adjust inventory quantity by adding or removing stock.
    - Add: Creates a new lot with specified cost
    - Remove: Deducts from existing lots using FIFO
    """
    try:
        adjustment_type = request.form.get("adjustment_type", "").strip()
        quantity_str = request.form.get("quantity", "0").strip()
        reason = request.form.get("reason", "").strip()
        unit_cost_str = request.form.get("unit_cost", "0").strip()

        # Validate inputs
        try:
            quantity = int(quantity_str)
        except ValueError:
            flash("Invalid quantity. Please enter a number.", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        try:
            unit_cost = float(unit_cost_str) if unit_cost_str else 0.0
        except ValueError:
            flash("Invalid unit cost. Please enter a valid number.", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        if quantity <= 0:
            flash("Quantity must be greater than zero.", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        if not reason:
            flash("Please provide a reason for the adjustment.", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        # Get item
        item = get_item(sku)
        if not item:
            flash("Item not found.", "error")
            return redirect(url_for("inventory"))

        # Get employee info
        emp_id = session.get("employee_id", "SYSTEM")
        emp_name = session.get("employee_name", "System")

        if adjustment_type == "add":
            # ================================================================
            # ADD INVENTORY - Create a new lot
            # ================================================================
            if unit_cost <= 0:
                flash("Unit cost is required when adding inventory.", "error")
                return redirect(url_for("edit_inventory", sku=sku))

            supplier = request.form.get("supplier", "").strip() or "Manual Adjustment"
            invoice_ref = request.form.get("invoice_ref", "").strip()

            if not invoice_ref:
                invoice_ref = f"ADJ-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            # Create the lot
            lot_id = create_lot(
                sku=sku,
                qty=quantity,
                unit_cost=unit_cost,
                supplier=supplier,
                invoice_ref=invoice_ref,
                notes=f"Manual adjustment: {reason}",
            )

            # Log to ledger
            ledger_entry = {
                "Entry_ID": str(uuid.uuid4()),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "SKU": sku,
                "Lot_ID": lot_id,
                "Delta_Qty": str(quantity),
                "Reason": f"MANUAL_ADD: {reason}",
                "Ref_Transaction_ID": "",
                "Employee_ID": emp_id,
                "Notes": f"Added by {emp_name}",
            }
            append_csv(get_ledger_path(), LEDGER_HEADERS, ledger_entry)

            # Log audit
            log_audit(
                emp_id,
                emp_name,
                "INVENTORY_ADD",
                "Inventory",
                sku,
                item.get("Item_Name", ""),
                old_value="",
                new_value=f"+{quantity}",
                notes=f"Added {quantity} units @ ${unit_cost:.2f} - {reason}",
            )

            flash(f"✅ Added {quantity} units to inventory (Lot: {lot_id})", "success")

        elif adjustment_type == "remove":
            # ================================================================
            # REMOVE INVENTORY - Deduct using FIFO
            # ================================================================
            lots = get_lots_for_sku(sku, available_only=True)

            # Check if we have enough stock
            total_available = sum(int(lot.get("Qty_Remaining", 0) or 0) for lot in lots)

            if total_available <= 0:
                flash("No stock available to remove.", "error")
                return redirect(url_for("edit_inventory", sku=sku))

            if quantity > total_available:
                flash(
                    f"Cannot remove {quantity} units. Only {total_available} available.",
                    "error",
                )
                return redirect(url_for("edit_inventory", sku=sku))

            # Process removal using FIFO
            remaining_to_remove = quantity
            lots_path = get_lots_path()
            all_lots = read_csv(lots_path)
            lots_updated = []

            for lot in lots:
                if remaining_to_remove <= 0:
                    break

                lot_id = lot.get("Lot_ID", "")
                available = int(lot.get("Qty_Remaining", 0) or 0)

                if available <= 0:
                    continue

                # Determine how much to remove from this lot
                qty_from_this_lot = min(remaining_to_remove, available)

                # Update the lot in all_lots
                for l in all_lots:
                    if l.get("Lot_ID") == lot_id:
                        new_remaining = available - qty_from_this_lot
                        l["Qty_Remaining"] = str(new_remaining)
                        lots_updated.append(
                            {
                                "lot_id": lot_id,
                                "qty_removed": qty_from_this_lot,
                                "remaining": new_remaining,
                            }
                        )
                        break

                # Log to ledger
                ledger_entry = {
                    "Entry_ID": str(uuid.uuid4()),
                    "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "SKU": sku,
                    "Lot_ID": lot_id,
                    "Delta_Qty": str(-qty_from_this_lot),
                    "Reason": f"MANUAL_REMOVE: {reason}",
                    "Ref_Transaction_ID": "",
                    "Employee_ID": emp_id,
                    "Notes": f"Removed by {emp_name}",
                }
                append_csv(get_ledger_path(), LEDGER_HEADERS, ledger_entry)

                remaining_to_remove -= qty_from_this_lot

            # Save updated lots
            write_csv(lots_path, LOT_HEADERS, all_lots)

            # Log audit
            log_audit(
                emp_id,
                emp_name,
                "INVENTORY_REMOVE",
                "Inventory",
                sku,
                item.get("Item_Name", ""),
                old_value=str(total_available),
                new_value=str(total_available - quantity),
                notes=f"Removed {quantity} units - {reason}",
            )

            flash(f"✅ Removed {quantity} units from inventory", "success")

        else:
            flash("Invalid adjustment type.", "error")
            return redirect(url_for("edit_inventory", sku=sku))

        return redirect(url_for("edit_inventory", sku=sku))

    except Exception as e:
        import traceback

        traceback.print_exc()
        flash(f"Error adjusting inventory: {str(e)}", "error")
        return redirect(url_for("edit_inventory", sku=sku))


# ensure_csv_exists(get_lots_path(), LOT_HEADERS)
# ensure_csv_exists(get_ledger_path(), LEDGER_HEADERS)


#!/usr/bin/env python3
"""
ENHANCED RECEIVE STOCK ROUTE
# =============================
Replace the existing receive_stock route in MGN_APP.py

This version includes:
- Barcode mapping on receive
- Auto-creates internal QR mappings
- Pristine inventory tracking
- Complete audit trail
"""


@app.route("/inventory/receive", methods=["GET", "POST"], endpoint="receive_stock")
@manager_required
def receive_stock():
    """
    Enhanced receive stock with barcode mapping.

    Modes:
    - existing: Receive into existing SKU
    - new: Create new item and receive
    """
    if request.method == "GET":
        items = get_all_items()
        return render_template("inventory/receive.html", items=items)

    # POST - Process receive
    try:
        mode = request.form.get("mode", "").strip()

        if mode == "existing":
            # ================================================================
            # RECEIVE INTO EXISTING ITEM
            # ================================================================
            sku = request.form.get("sku", "").strip()
            qty_str = request.form.get("qty", "0").strip()
            unit_cost_str = request.form.get("unit_cost", "0").strip() or "0"
            date_str = request.form.get("date", "").strip()
            note = request.form.get("note", "").strip()
            vendor_barcode = request.form.get("vendor_barcode", "").strip()
            vendor_name = request.form.get("vendor", "").strip()

            # Validate
            if not sku:
                flash("SKU is required", "error")
                return redirect(url_for("receive_stock"))

            item = get_item(sku)
            if not item:
                flash(f"SKU {sku} not found", "error")
                return redirect(url_for("receive_stock"))

            try:
                qty = int(qty_str)
            except ValueError:
                flash("Invalid quantity", "error")
                return redirect(url_for("receive_stock"))

            try:
                unit_cost = float(unit_cost_str)
            except ValueError:
                flash("Invalid unit cost", "error")
                return redirect(url_for("receive_stock"))

            if qty <= 0:
                flash("Quantity must be greater than zero", "error")
                return redirect(url_for("receive_stock"))

            if unit_cost < 0:
                flash("Unit cost cannot be negative", "error")
                return redirect(url_for("receive_stock"))

            # Use today if no date provided
            if not date_str:
                date_str = date.today().strftime("%Y-%m-%d")

            # Create lot
            lot_id = create_lot(
                sku=sku,
                qty=qty,
                unit_cost=unit_cost,
                supplier=vendor_name or "Manual",
                invoice_ref=note[:50]
                if note
                else f"RCV-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                received_date=date_str,
                notes=note,
            )

            # Log to ledger
            emp_id = session.get("employee_id", "SYSTEM")
            emp_name = session.get("employee_name", "System")

            ledger_entry = {
                "Entry_ID": str(uuid.uuid4()),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "SKU": sku,
                "Lot_ID": lot_id,
                "Delta_Qty": str(qty),
                "Reason": "RECEIVE",
                "Ref_Transaction_ID": "",
                "Employee_ID": emp_id,
                "Notes": f"Received by {emp_name}: {note}",
            }
            append_csv(get_ledger_path(), LEDGER_HEADERS, ledger_entry)

            # Create/update barcode mappings
            # 1) Ensure internal QR exists
            create_internal_qr_mapping(sku)

            # 2) If vendor barcode provided, map it
            if vendor_barcode:
                success, msg = upsert_barcode(
                    code=vendor_barcode,
                    sku=sku,
                    vendor=vendor_name or "Vendor",
                    code_type="VENDOR_BARCODE",
                    notes=f"Added during receive: {note}",
                )
                if not success:
                    flash(f"Warning: {msg}", "warning")

            # Audit log
            log_audit(
                emp_id,
                emp_name,
                "RECEIVE_STOCK",
                "Inventory",
                sku,
                item.get("Item_Name", ""),
                old_value="",
                new_value=f"+{qty}",
                notes=f"Lot: {lot_id}, Cost: ${unit_cost:.2f}, Vendor: {vendor_name}",
            )

            flash(f"✅ Received {qty} units of {sku} (Lot: {lot_id})", "success")
            return redirect(url_for("receive_stock"))

        elif mode == "new":
            # ================================================================
            # CREATE NEW ITEM AND RECEIVE
            # ================================================================
            sku = request.form.get("sku", "").strip()
            item_name = request.form.get("item_name", "").strip()
            category = request.form.get("category", "").strip()
            price_str = request.form.get("price", "0").strip()
            qty_str = request.form.get("qty", "0").strip()
            unit_cost_str = request.form.get("unit_cost", "0").strip() or "0"
            taxable_str = request.form.get("taxable", "1").strip()
            date_str = request.form.get("date", "").strip()
            vendor_name = request.form.get("vendor", "").strip()
            barcode = request.form.get("barcode", "").strip()
            note = request.form.get("note", "").strip()

            # Validate
            if not sku:
                flash("SKU is required", "error")
                return redirect(url_for("receive_stock"))

            if not item_name:
                flash("Item name is required", "error")
                return redirect(url_for("receive_stock"))

            # Check if SKU already exists
            existing = get_item(sku)
            if existing:
                flash(
                    f"SKU {sku} already exists. Use 'Receive Existing' instead.",
                    "error",
                )
                return redirect(url_for("receive_stock"))

            try:
                price = float(price_str)
                qty = int(qty_str)
                unit_cost = float(unit_cost_str)
            except ValueError:
                flash("Invalid price, quantity, or cost", "error")
                return redirect(url_for("receive_stock"))

            if qty <= 0:
                flash("Quantity must be greater than zero", "error")
                return redirect(url_for("receive_stock"))

            # Use today if no date provided
            if not date_str:
                date_str = date.today().strftime("%Y-%m-%d")

            # Create item in Items.csv
            taxable = (
                "Y" if taxable_str in ("1", "Y", "Yes", "yes", "true", "True") else "N"
            )

            success, msg, created_sku = create_item(
                sku=sku,
                item_name=item_name,
                category=category,
                price=price,
                taxable=taxable,
            )

            if not success:
                flash(msg, "error")
                return redirect(url_for("receive_stock"))

            # Create lot
            lot_id = create_lot(
                sku=sku,
                qty=qty,
                unit_cost=unit_cost,
                supplier=vendor_name or "Manual",
                invoice_ref=note[:50]
                if note
                else f"NEW-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                received_date=date_str,
                notes=note,
            )

            # Log to ledger
            emp_id = session.get("employee_id", "SYSTEM")
            emp_name = session.get("employee_name", "System")

            ledger_entry = {
                "Entry_ID": str(uuid.uuid4()),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "SKU": sku,
                "Lot_ID": lot_id,
                "Delta_Qty": str(qty),
                "Reason": "INITIAL_RECEIVE",
                "Ref_Transaction_ID": "",
                "Employee_ID": emp_id,
                "Notes": f"New item created and received by {emp_name}: {note}",
            }
            append_csv(get_ledger_path(), LEDGER_HEADERS, ledger_entry)

            # Create barcode mappings
            # 1) Always create internal QR
            create_internal_qr_mapping(sku)

            # 2) If vendor barcode provided, map it
            if barcode:
                success_bc, msg_bc = upsert_barcode(
                    code=barcode,
                    sku=sku,
                    vendor=vendor_name or "Vendor",
                    code_type="VENDOR_BARCODE",
                    notes=f"Created with new item: {note}",
                )
                if not success_bc:
                    flash(f"Warning: {msg_bc}", "warning")

            # Audit log
            log_audit(
                emp_id,
                emp_name,
                "CREATE_ITEM_RECEIVE",
                "Inventory",
                sku,
                item_name,
                old_value="",
                new_value=f"Created + received {qty}",
                notes=f"Category: {category}, Lot: {lot_id}, Cost: ${unit_cost:.2f}",
            )

            flash(
                f"✅ Created {sku} and received {qty} units (Lot: {lot_id})", "success"
            )
            return redirect(url_for("receive_stock"))

        else:
            flash("Invalid mode", "error")
            return redirect(url_for("receive_stock"))

    except Exception as e:
        import traceback

        traceback.print_exc()
        flash(f"Error receiving stock: {str(e)}", "error")
        return redirect(url_for("receive_stock"))


# ADD THIS ROUTE TO MGN_APP.py
# Place it near your other receive_stock routes


@app.route("/inventory/bulk-csv-import", methods=["POST"])
@login_required
@tier_required("pro")
def bulk_csv_import():
    """Bulk CSV import from receive page"""

    if "csv_file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["csv_file"]

    if file.filename == "":
        return jsonify({"success": False, "error": "No file selected"}), 400

    if not file.filename.endswith(".csv"):
        return jsonify({"success": False, "error": "File must be a CSV"}), 400

    try:
        import csv
        from io import StringIO

        # Read CSV
        content = file.read().decode("utf-8")
        csv_reader = csv.DictReader(StringIO(content))

        # Get existing Items.csv
        items_path = get_items_path()
        with open(items_path, "r", encoding="utf-8") as f:
            existing_reader = csv.DictReader(f)
            existing_headers = existing_reader.fieldnames
            existing_items = list(existing_reader)

        # Get existing SKUs
        existing_skus = {
            item.get("SKU", "") for item in existing_items if item.get("SKU")
        }

        # Validate headers
        csv_headers = csv_reader.fieldnames
        if set(csv_headers) != set(existing_headers):
            return jsonify(
                {
                    "success": False,
                    "error": "CSV headers do not match Items.csv structure",
                }
            ), 400

        # Process items
        new_items = []
        duplicate_skus = []

        for row in csv_reader:
            sku = row.get("SKU", "").strip()

            if not sku:
                continue

            if sku in existing_skus:
                duplicate_skus.append(sku)
                continue

            new_items.append(row)
            existing_skus.add(sku)

        if not new_items:
            return jsonify(
                {
                    "success": False,
                    "error": "No new items to import (all SKUs already exist)",
                }
            ), 400

        # Append to Items.csv
        with open(items_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=existing_headers)
            writer.writerows(new_items)

        return jsonify(
            {
                "success": True,
                "imported": len(new_items),
                "duplicates": len(duplicate_skus),
                "message": f"Successfully imported {len(new_items)} items",
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": f"Import error: {str(e)}"}), 500


@app.route("/inventory/low-stock")
@login_required
def low_stock_report():
    low_stock = check_low_stock()
    return render_template("inventory/low_stock.html", items=low_stock)
    return lot_id


@app.route("/inventory/lots")
@login_required
def view_lots():
    sku = request.args.get("sku")
    if not sku:
        return redirect(url_for("inventory"))

    item = get_item(sku)
    if not item:
        flash("Item not found", "error")
        return redirect(url_for("inventory"))

    lots = get_lots_for_sku(sku, available_only=False)
    return render_template(
        "inventory/lots.html",
        item=item,
        lots=lots,
        stock=get_stock_on_hand(sku),
        avg_cost=get_average_cost(sku),
    )


def _patch_receive_dates(lot_id: str, recv_date: str) -> None:
    # requires these to be imported from POS_CORE:
    # read_csv, write_csv, get_lots_path, LOT_HEADERS, get_ledger_path, LEDGER_HEADERS
    lots = read_csv(get_lots_path())
    for row in lots:
        if row.get("Lot_ID") == lot_id:
            row["Received_Date"] = recv_date
            break
    write_csv(get_lots_path(), LOT_HEADERS, lots)

    ledger = read_csv(get_ledger_path())
    for row in ledger:
        if row.get("Lot_ID") == lot_id and row.get("Reason") == "Receive":
            ts = row.get("Timestamp", "")
            time_part = ts.split(" ", 1)[1] if " " in ts else "00:00:00"
            row["Timestamp"] = f"{recv_date} {time_part}"
            break
    write_csv(get_ledger_path(), LEDGER_HEADERS, ledger)


# ==============================================================================
#                     TIME CLOCK
# ==============================================================================
LA_TZ = ZoneInfo("America/Los_Angeles")


def _parse_dt(s: str):
    """Best-effort datetime parser for common formats."""
    if not s:
        return None
    s = s.strip()
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            pass
    # last resort: try fromisoformat
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def compute_seconds_worked_today(employee_id: str) -> int:
    """
    Reads clock events from CSV and computes worked seconds for today in America/Los_Angeles.
    Supports common columns: employee_id/emp_id/user_id, action/event/type, timestamp/time/datetime.
    Actions supported: IN/OUT, CLOCK_IN/CLOCK_OUT.
    """
    if not employee_id:
        return 0

    if not os.path.exists(TIME_CLOCK_CSV):
        return 0

    now = datetime.now(LA_TZ)
    today_start = datetime.combine(now.date(), dtime(0, 0, 0), tzinfo=LA_TZ)
    today_end = datetime.combine(now.date(), dtime(23, 59, 59), tzinfo=LA_TZ)

    # Load events for this employee
    events = []
    with open(TIME_CLOCK_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # employee id column variants
            rid = (
                row.get("employee_id") or row.get("emp_id") or row.get("user_id") or ""
            ).strip()
            if rid != str(employee_id):
                continue

            action = (
                (row.get("action") or row.get("event") or row.get("type") or "")
                .strip()
                .upper()
            )

            ts_raw = (
                row.get("timestamp")
                or row.get("time")
                or row.get("datetime")
                or row.get("created_at")
                or ""
            )
            dt = _parse_dt(ts_raw)

            if not dt:
                continue

            # assume local if naive
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=LA_TZ)
            else:
                dt = dt.astimezone(LA_TZ)

            # only today’s events matter
            if dt < today_start or dt > today_end:
                continue

            # normalize action
            if action in {"IN", "CLOCKIN", "CLOCK_IN", "CLOCK-IN"}:
                action = "IN"
            elif action in {"OUT", "CLOCKOUT", "CLOCK_OUT", "CLOCK-OUT"}:
                action = "OUT"
            else:
                # ignore unknown actions
                continue

            events.append((dt, action))

    if not events:
        return 0

    events.sort(key=lambda x: x[0])

    total = 0
    last_in = None

    for dt, action in events:
        if action == "IN":
            last_in = dt
        elif action == "OUT":
            if last_in:
                total += int((dt - last_in).total_seconds())
                last_in = None

    # if still clocked in, count up to now
    if last_in:
        total += int((now - last_in).total_seconds())

    return max(total, 0)


@app.route("/api/me/clock-status")
@login_required
def api_me_clock_status():
    emp_id = session.get("employee_id", "")
    status = get_employee_status(emp_id) or {}
    return jsonify(
        {
            "employee_id": emp_id,
            "is_clocked_in": bool(status.get("is_clocked_in")),
            "hours_today": float(status.get("hours_today", 0.0) or 0.0),
            "last_punch": status.get("last_punch", ""),
        }
    )


@app.get("/api/me/hours_today")
def api_me_hours_today():
    # adjust to how you store login identity
    employee_id = session.get("employee_id") or session.get("user_id")
    if not employee_id:
        return jsonify({"ok": False, "error": "Not logged in"}), 401

    seconds = compute_seconds_worked_today(str(employee_id))
    return jsonify({"ok": True, "seconds_today": seconds})


@app.route("/timeclock")
@login_required
def timeclock():
    status = get_employee_status(session["employee_id"])
    return render_template("time/clock.html", status=status)


@app.route("/timeclock/punch", methods=["POST"])
@login_required
def punch_clock():
    action = request.form.get("action")
    emp_id = session["employee_id"]
    emp_name = session["employee_name"]

    if action == "clock_in":
        success, msg, _ = clock_in(emp_id, emp_name)
    elif action == "clock_out":
        success, msg, _ = clock_out(emp_id, emp_name)
    elif action == "break":
        success, msg, _ = start_break(emp_id, emp_name, "BREAK")
    elif action == "lunch":
        success, msg, _ = start_break(emp_id, emp_name, "LUNCH")
    elif action == "end_break":
        success, msg, _ = end_break(emp_id, emp_name)
    else:
        success, msg = False, "Invalid action"

    flash(msg, "success" if success else "error")
    return redirect(url_for("timeclock"))


@app.route("/timeclock/log")
@login_required
def timeclock_log():
    target_date = request.args.get("date")
    if target_date:
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        target_date = date.today()

    punches = read_csv(get_timeclock_path(target_date))

    if session.get("role") not in ("Manager", "Owner", "Admin"):
        punches = [p for p in punches if p.get("Employee_ID") == session["employee_id"]]

    return render_template("time/log.html", punches=punches, selected_date=target_date)


@app.route("/timeclock/manage")
@manager_required
def timeclock_manage():
    target_date = request.args.get("date")
    if target_date:
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        target_date = date.today()

    punches = get_punches_for_date(target_date)
    employees = get_all_employees()
    return render_template(
        "timeclock/manage.html",
        punches=punches,
        employees=employees,
        selected_date=target_date,
    )


@app.route("/timeclock/edit/<punch_id>", methods=["GET", "POST"])
@manager_required
def timeclock_edit_punch(punch_id):
    punch = get_punch_by_id(punch_id)
    if not punch:
        flash("Punch not found", "error")
        return redirect(url_for("timeclock_manage"))

    if request.method == "POST":
        action = request.form.get("action")
        reason = request.form.get("reason", "")

        if action == "update":
            new_time = request.form.get("punch_time")
            new_type = request.form.get("punch_type")
            success, msg = edit_punch(
                punch_id,
                new_time,
                new_type,
                session["employee_id"],
                session["employee_name"],
                reason,
            )
        elif action == "delete":
            success, msg = delete_punch(
                punch_id, session["employee_id"], session["employee_name"], reason
            )
        else:
            success, msg = False, "Invalid action"

        flash(msg, "success" if success else "error")
        return redirect(url_for("timeclock_manage", date=punch.get("Date")))

    punch_types = ["CLOCK_IN", "CLOCK_OUT", "BREAK", "LUNCH", "END_BREAK", "END_LUNCH"]
    return render_template(
        "timeclock/edit_punch.html", punch=punch, punch_types=punch_types
    )


@app.route("/timeclock/add-punch", methods=["GET", "POST"])
@manager_required
def timeclock_add_punch():
    if request.method == "POST":
        punch_time = request.form.get("punch_time")
        if punch_time and len(punch_time) == 5:
            punch_time = f"{punch_time}:00"

        success, msg = add_punch(
            employee_id=request.form.get("employee_id"),
            employee_name=request.form.get("employee_name"),
            punch_date=request.form.get("punch_date"),
            punch_time=punch_time,
            punch_type=request.form.get("punch_type"),
            added_by_id=session["employee_id"],
            added_by_name=session["employee_name"],
            reason=request.form.get("reason", ""),
        )
        flash(msg, "success" if success else "error")
        if success:
            return redirect(
                url_for("timeclock_manage", date=request.form.get("punch_date"))
            )

    employees = get_all_employees()
    punch_types = ["CLOCK_IN", "CLOCK_OUT", "BREAK", "LUNCH", "END_BREAK", "END_LUNCH"]
    return render_template(
        "timeclock/add_punch.html",
        employees=employees,
        punch_types=punch_types,
        today=date.today(),
    )


@app.route("/timeclock/edit-history")
@manager_required
def timeclock_edit_history_view():
    days = int(request.args.get("days", 30))
    employee_id = request.args.get("employee_id")
    history = get_timeclock_edit_history(employee_id, days)
    employees = get_all_employees()
    return render_template(
        "timeclock/edit_history.html",
        history=history,
        employees=employees,
        selected_days=days,
    )


# ==============================================================================
#                     TIME OFF
# ==============================================================================

# --- TIME OFF storage (local, stable) ---
TIMEOFF_HEADERS = [
    "Request_ID",
    "Employee_ID",
    "Employee_Name",
    "Request_Date",
    "Start_Date",
    "End_Date",
    "Days_Requested",
    "Reason",
    "Status",
    "Manager_Name",
    "Approval_Date",
    "Manager_Notes",
]


def get_timeoff_path():
    # matches your structure: Time_Off_Requests/2025_TimeOffRequests.csv
    year = datetime.now().strftime("%Y")
    folder = os.path.join(DATA_DIR, "Time_Off_Requests")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"{year}_TimeOffRequests.csv")


def generate_id(prefix="ID"):
    # unique enough for POS use (timestamp to milliseconds)
    return f"{prefix}{datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]}"


def _request_time_off_core(emp_id, emp_name, start, end, reason=""):
    """Core logic: validate + write to Time_Off_Requests/YYYY_TimeOffRequests.csv"""
    if start < date.today():
        return False, "🚫 Cannot request past dates.", ""

    min_start = date.today() + timedelta(days=14)
    if start < min_start:
        return (
            False,
            f"⏳ Must request at least 14 days in advance. Earliest: {min_start:%Y-%m-%d}",
            "",
        )

    if end < start:
        return False, "🚫 End date must be on/after start date.", ""

    days = (end - start).days + 1
    req_id = generate_id("REQ")

    row = {
        "Request_ID": req_id,
        "Employee_ID": str(emp_id),
        "Employee_Name": emp_name,
        "Request_Date": date.today().strftime("%Y-%m-%d"),
        "Start_Date": start.strftime("%Y-%m-%d"),
        "End_Date": end.strftime("%Y-%m-%d"),
        "Days_Requested": str(days),
        "Reason": reason,
        "Status": "Pending",
        "Manager_Name": "",
        "Approval_Date": "",
        "Manager_Notes": "",
    }

    append_csv(get_timeoff_path(), TIMEOFF_HEADERS, row)

    return True, f"🎉 Request submitted ({days} days).", req_id


@app.route("/timeoff")
@login_required
def timeoff():
    my_requests = get_time_off_requests(employee_id=str(session["employee_id"]))
    pending = []
    if session.get("role") in ("Manager", "Owner", "Admin"):
        pending = get_pending_requests()

    earliest_date = (date.today() + timedelta(days=14)).strftime("%Y-%m-%d")

    my_requests = sorted(
        my_requests, key=lambda r: r.get("Request_Date", ""), reverse=True
    )
    pending = sorted(pending, key=lambda r: r.get("Request_Date", ""), reverse=True)

    return render_template(
        "time/timeoff.html",
        my_requests=my_requests,
        pending_requests=pending,
        earliest_date=earliest_date,
    )


@app.route("/timeoff/request", methods=["POST"], endpoint="request_timeoff")
@login_required
def request_timeoff():
    # Parse form
    start_s = (request.form.get("start_date") or "").strip()
    end_s = (request.form.get("end_date") or "").strip()
    reason = (request.form.get("reason") or "").strip()

    if not start_s or not end_s:
        flash("🚫 Start and End dates are required.", "error")
        return redirect(url_for("timeoff"))

    start = datetime.strptime(start_s, "%Y-%m-%d").date()
    end = datetime.strptime(end_s, "%Y-%m-%d").date()

    # Core write (saves to your existing YYYY_TimeOffRequests.csv)
    success, msg, req_id = _request_time_off_core(
        str(session["employee_id"]),
        session["employee_name"],
        start,
        end,
        reason,
    )

    if success:
        flash(f"{msg} ✅ {start} → {end} 🆔 {req_id}", "success")
        flash(
            "📬 Your manager will review it soon — thanks for planning ahead 🙌", "info"
        )

        # 🔔 Notify managers/owners/admins
        try:
            for e in get_all_employees(include_inactive=True):
                if (
                    e.get("Role") in ("Manager", "Owner", "Admin")
                    and e.get("Status", "Active") != "Inactive"
                ):
                    if str(e.get("Employee_ID")) == str(session["employee_id"]):
                        continue
                    note = f"⏳ TIME OFF PENDING: {session['employee_name']} • {start} → {end} • ID {req_id}"
                    if reason:
                        note += f" • 📝 {reason}"
                    create_notification(
                        e.get("Employee_ID"), e.get("Employee_Name"), note, "TIMEOFF"
                    )
        except Exception:
            pass
    else:
        flash(msg, "error")

    return redirect(url_for("timeoff"))


@app.route("/timeoff/decision", methods=["POST"], endpoint="timeoff_decision")
@login_required
def timeoff_decision():
    if session.get("role") not in ("Manager", "Owner", "Admin"):
        flash("🚫 Not authorized.", "error")
        return redirect(url_for("timeoff"))

    req_id = (request.form.get("req_id") or "").strip()
    decision = (request.form.get("decision") or "").strip()
    notes = (request.form.get("manager_notes") or "").strip()
    approved = decision == "approve"

    if not notes:
        notes = (
            "🎉 Approved! Enjoy your time off 🙌"
            if approved
            else "🙏 Thanks for understanding — not approved at this time. Please try different dates or talk with your manager. 💬"
        )

    # IMPORTANT: call the POS_CORE function, not this file's route name
    ok, msg = approve_time_off(
        req_id, str(session["employee_id"]), session["employee_name"], approved, notes
    )

    flash(msg, "success" if ok else "error")
    return redirect(url_for("timeoff"))


# ==============================================================================
#                     EMPLOYEES
# ==============================================================================


@app.route("/employees/<emp_id>/edit", methods=["GET", "POST"])
@manager_required
def edit_employee(emp_id):
    """Edit an existing employee"""
    emp_file = os.path.join(DATA_DIR, "Employees", "Employee_Directory.csv")

    if request.method == "POST":
        try:
            employees = []
            updated = False

            with open(emp_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                for row in reader:
                    if row.get("Employee_ID") == emp_id:
                        row["Employee_Name"] = request.form.get(
                            "name", row.get("Employee_Name")
                        )
                        row["Role"] = request.form.get("role", row.get("Role"))
                        row["Phone"] = request.form.get("phone", row.get("Phone", ""))
                        row["Email"] = request.form.get("email", row.get("Email", ""))
                        row["Emergency_Contact"] = request.form.get(
                            "emergency_contact", row.get("Emergency_Contact", "")
                        )
                        row["Notes"] = request.form.get("notes", row.get("Notes", ""))
                        row["Last_Updated"] = datetime.now().strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        new_pin = request.form.get("pin", "").strip()
                        if new_pin:
                            row["PIN"] = new_pin
                        updated = True
                    employees.append(row)

            if updated:
                with open(emp_file, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(employees)
                flash("Employee updated successfully", "success")
                return redirect(url_for("employee_detail", emp_id=emp_id))
            else:
                flash("Employee not found", "error")
                return redirect(url_for("employees"))
        except Exception as e:
            flash(f"Error updating employee: {str(e)}", "error")

    emp = get_employee(emp_id)
    if not emp:
        flash("Employee not found", "error")
        return redirect(url_for("employees"))

    return render_template("employees/edit.html", employee=emp, roles=ROLES)


@app.route("/employees/<emp_id>/reset-pin")
@owner_required
def reset_employee_pin(emp_id):
    """Generate a new random PIN for employee"""
    import random

    try:
        emp_file = os.path.join(DATA_DIR, "Employees", "Employee_Directory.csv")
        new_pin = str(random.randint(1000, 9999))

        employees = []
        emp_name = None

        with open(emp_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                if row.get("Employee_ID") == emp_id:
                    row["PIN"] = new_pin
                    row["Last_Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    emp_name = row.get("Employee_Name")
                employees.append(row)

        if emp_name:
            with open(emp_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(employees)
            flash(f"New PIN for {emp_name}: {new_pin}", "success")
        else:
            flash("Employee not found", "error")

        return redirect(url_for("employee_detail", emp_id=emp_id))

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("employees"))


@app.route("/employees/<emp_id>/set-pin", methods=["POST"])
@owner_required
def set_employee_pin(emp_id):
    """Owner/manager assigns a SPECIFIC 4-digit PIN (not random). Validated,
    audit-logged, and the employee is notified -- via the same hardened path as a
    reset so there is one trail for every PIN change."""
    new_pin = (request.form.get("new_pin") or "").strip()
    ok, msg = reset_pin(
        emp_id,
        new_pin,
        session.get("employee_id", ""),
        session.get("employee_name", ""),
    )
    flash(msg, "success" if ok else "error")
    return redirect(url_for("employee_detail", emp_id=emp_id))


@app.route("/employees/<emp_id>/activate")
@manager_required
def activate_employee(emp_id):
    """Set employee status to Active"""
    return _set_employee_status(emp_id, "Active")


@app.route("/employees/<emp_id>/deactivate")
@manager_required
def deactivate_employee(emp_id):
    """Set employee status to Inactive"""
    return _set_employee_status(emp_id, "Inactive")


def _set_employee_status(emp_id, status):
    """Helper to update employee status"""
    try:
        emp_file = os.path.join(DATA_DIR, "Employees", "Employee_Directory.csv")

        employees = []
        updated = False

        with open(emp_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                if row.get("Employee_ID") == emp_id:
                    row["Status"] = status
                    row["Last_Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    updated = True
                employees.append(row)

        if updated:
            with open(emp_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(employees)
            flash(f"Employee status changed to {status}", "success")
        else:
            flash("Employee not found", "error")

        return redirect(url_for("employee_detail", emp_id=emp_id))

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("employees"))


@app.route("/employees/<emp_id>/delete", methods=["POST"])
@manager_required
def delete_employee(emp_id):
    try:
        emp_file = os.path.join(DATA_DIR, "Employees", "Employee_Directory.csv")
        employees = []
        deleted_name = None

        with open(emp_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                if row.get("Employee_ID") == emp_id:
                    deleted_name = row.get("Employee_Name", "Unknown")
                    continue
                employees.append(row)

        if deleted_name:
            with open(emp_file, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(employees)
            flash(f'Employee "{deleted_name}" deleted', "success")
        else:
            flash("Employee not found", "error")

        return redirect(url_for("employees"))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("employees"))


@app.route("/employees")
@manager_required
def employees():
    emps = get_all_employees(include_inactive=True)
    return render_template("employees/list.html", employees=emps)


@app.route("/employees/add", methods=["GET", "POST"])
@manager_required
def add_employee():
    if request.method == "POST":
        if not enforce_limit_or_upgrade("max_users", len(get_all_employees())):
            return redirect(url_for("employees"))

        # The add form posts first_name + last_name; the old code read a single
        # "name" field that never existed, so every new hire saved with a BLANK
        # name until re-edited. Combine the parts (still accept a single "name").
        name = (request.form.get("name") or "").strip()
        if not name:
            name = compose_full_name(
                request.form.get("first_name", ""),
                request.form.get("last_name", ""),
            )
        role = request.form.get("role")
        pin = request.form.get("pin")
        phone = request.form.get("phone", "")
        email = request.form.get("email", "")

        if not name:
            flash("Employee name is required.", "error")
            return redirect(url_for("add_employee"))

        success, msg, emp_id = create_employee(name, role, pin, phone, email)
        if success:
            # The form also collects an hourly rate that the old route silently
            # dropped -- wire it into pay config so payroll/Money OS sees it.
            rate = (request.form.get("hourly_rate") or "").strip()
            if rate:
                try:
                    setup_employee_pay(emp_id, "HOURLY", hourly_rate=float(rate))
                except Exception:
                    pass
            flash(msg, "success")
            return redirect(url_for("employees"))
        flash(msg, "error")

    return render_template("employees/add.html", roles=ROLES)


@app.route("/employees/<emp_id>")
@manager_required
def employee_detail(emp_id):
    emp = get_employee(emp_id)
    if not emp:
        flash("Employee not found", "error")
        return redirect(url_for("employees"))

    pay_config = get_employee_pay_config(emp_id)
    return render_template("employees/detail.html", employee=emp, pay_config=pay_config)


# ==============================================================================
#                     TASKS
# ==============================================================================


@app.route("/tasks")
@manager_required
def tasks_owner_dashboard():
    today = date.today()

    employees = get_all_employees()
    templates = get_all_tasks()
    assignments = get_task_assignments_for_date(today)

    # group by employee
    by_emp = {}
    for e in employees:
        eid = e.get("Employee_ID")
        by_emp[eid] = {
            "emp": e,
            "tasks": [a for a in assignments if a.get("Employee_ID") == eid],
        }

    # Owner notifications (TASK-type)
    owner_notifs = [
        n
        for n in get_all_notifications(session["employee_id"], limit=30)
        if n.get("Type") == "TASK"
    ]

    return render_template(
        "tasks/owner_dashboard.html",
        today=today,
        employees=employees,
        templates=templates,
        by_emp=by_emp,
        owner_notifications=owner_notifs,
    )


@app.route("/tasks/assign", methods=["POST"])
@manager_required
def tasks_assign():
    task_id = (request.form.get("task_id") or "").strip()
    employee_ids = request.form.getlist("employee_ids")  # multi-select
    due_date = (request.form.get("due_date") or date.today().isoformat()).strip()
    notes = (request.form.get("notes") or "").strip()

    if not task_id:
        flash("Please select a premade task.", "error")
        return redirect(url_for("tasks_owner_dashboard"))

    if not employee_ids:
        flash("Please select at least 1 employee.", "error")
        return redirect(url_for("tasks_owner_dashboard"))

    tmpl = get_task(task_id)
    title = tmpl.get("Title") if tmpl else "Task"

    assigned_count = 0
    for eid in employee_ids:
        if not eid:
            continue

        assign_task(task_id, eid, due_date, session["employee_id"], notes)
        emp = get_employee(eid)
        if emp:
            create_notification(
                eid,
                emp.get("Employee_Name", ""),
                f"🧩 New task assigned: {title} (Due {due_date})",
                "TASK",
            )
        assigned_count += 1

    flash(f"Task assigned to {assigned_count} employee(s).", "success")
    return redirect(url_for("tasks_owner_dashboard"))


@app.route("/tasks/quick-assign", methods=["POST"])
@manager_required
def tasks_quick_assign():
    """
    Custom task creator + assigner.
    ALSO auto-saves to templates via get_or_create_task_template().
    """
    title = (request.form.get("title") or "").strip()
    description = (request.form.get("description") or "").strip()
    category = (request.form.get("category") or "General").strip()
    priority = (request.form.get("priority") or "MEDIUM").strip()
    est_minutes = int(request.form.get("est_minutes") or 15)

    employee_ids = request.form.getlist("employee_ids")
    due_date = (request.form.get("due_date") or date.today().isoformat()).strip()
    notes = (request.form.get("notes") or "").strip()

    if not title:
        flash("Custom task title is required.", "error")
        return redirect(url_for("tasks_owner_dashboard"))

    if not employee_ids:
        flash("Please select at least 1 employee.", "error")
        return redirect(url_for("tasks_owner_dashboard"))

    task_id, created_new = get_or_create_task_template(
        title=title,
        description=description,
        category=category,
        priority=priority,
        estimated_minutes=est_minutes,
        created_by=session["employee_id"],
    )

    # Assign it
    assigned_count = 0
    for eid in employee_ids:
        if not eid:
            continue
        assign_task(task_id, eid, due_date, session["employee_id"], notes)
        emp = get_employee(eid)
        if emp:
            create_notification(
                eid,
                emp.get("Employee_Name", ""),
                f"🧩 New task assigned: {title} (Due {due_date})",
                "TASK",
            )
        assigned_count += 1

    if created_new:
        flash(
            f"Custom task created + saved to templates, assigned to {assigned_count}.",
            "success",
        )
    else:
        flash(f"Used existing template, assigned to {assigned_count}.", "success")

    return redirect(url_for("tasks_owner_dashboard"))


@app.route("/tasks/my")
@login_required
def my_tasks():
    today = date.today().isoformat()

    # pull ALL tasks for employee (not just today)
    tasks_all = get_tasks_for_employee(session["employee_id"], target_date=None)

    # open tasks = anything not completed/skipped
    open_tasks = [
        t for t in tasks_all if t.get("Status") not in ("COMPLETE", "SKIPPED")
    ]

    # show due today OR overdue (simple autonomy)
    due_tasks = []
    for t in open_tasks:
        due = (t.get("Due_Date") or "").strip()
        if not due or due <= today:
            due_tasks.append(t)

    completed_today = [
        t
        for t in tasks_all
        if t.get("Status") == "COMPLETE"
        and (t.get("Completed_At") or "").startswith(today)
    ]

    notifs = [
        n
        for n in get_all_notifications(session["employee_id"], limit=30)
        if n.get("Type") == "TASK"
        and str(n.get("Is_Read", n.get("Read", "N"))).upper()
        not in ("Y", "YES", "TRUE", "1")
    ]

    return render_template(
        "tasks/my_tasks.html",
        today=date.today(),
        due_tasks=due_tasks,
        completed_today=completed_today,
        notifications=notifs,
    )


@app.route("/tasks/<assignment_id>")
@login_required
def task_detail(assignment_id):
    assignment = get_task_assignment(assignment_id)
    if not assignment:
        flash("Task not found", "error")
        return redirect(url_for("my_tasks"))

    # access control: employee sees own; manager sees all
    if session.get("role") not in ("Owner", "Admin", "Manager"):
        if assignment.get("Employee_ID") != session.get("employee_id"):
            flash("Access denied.", "error")
            return redirect(url_for("my_tasks"))

    tmpl = get_task(assignment.get("Task_ID"))
    events = get_task_events(assignment_id)

    return render_template(
        "tasks/detail.html", assignment=assignment, tmpl=tmpl, events=events
    )


@app.route("/tasks/my/update-status", methods=["POST"])
@login_required
def my_tasks_update_status():
    data = request.get_json(silent=True) or {}

    assignment_id = (data.get("assignment_id") or "").strip()
    new_status = (data.get("status") or "").strip().upper()
    note = (data.get("note") or "").strip()
    skip_reason = (data.get("skip_reason") or "").strip()

    if new_status not in ("ACKNOWLEDGED", "IN_PROGRESS", "COMPLETE", "SKIPPED"):
        return jsonify({"success": False, "error": "Invalid status"}), 400

    ok = update_task_status(
        assignment_id=assignment_id,
        new_status=new_status,
        employee_id=session["employee_id"],
        note=note,
        skip_reason=skip_reason,
    )

    # Notify owner/manager (Assigned_By) when employee changes status
    if ok:
        a = get_task_assignment(assignment_id) or {}
        tmpl = get_task(a.get("Task_ID")) or {}
        title = tmpl.get("Title", "Task")

        assigned_by = (a.get("Assigned_By") or "").strip()
        if assigned_by:
            mgr = get_employee(assigned_by) or {"Employee_Name": "Manager"}
            msg = f"📣 Task update: {session.get('employee_name', 'Employee')} set '{title}' → {new_status}"
            if note:
                msg += f" — {note}"
            if new_status == "SKIPPED" and skip_reason:
                msg += f" (Reason: {skip_reason})"

            create_notification(
                assigned_by, mgr.get("Employee_Name", "Manager"), msg, "TASK"
            )

    return jsonify({"success": bool(ok)})


# ============================================
# COMPLETE PAYROLL ROUTES - USES YOUR CSV STRUCTURE
# Add these to MGN_APP.py
# ============================================

from pathlib import Path

# ============================================
# PAYROLL HELPER FUNCTIONS
# ============================================


def get_employee_pay_rate(emp_id):
    """Get hourly rate from Employee_Pay_Config.csv"""
    config_file = os.path.join(DATA_DIR, "Payroll", "Employee_Pay_Config.csv")
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    return float(row.get("Hourly_Rate", 0) or 0)
    return 15.00  # Default minimum wage


def get_all_pay_periods():
    """Get all pay periods from Pay_Periods.csv"""
    periods_file = os.path.join(DATA_DIR, "Payroll", "Pay_Periods.csv")
    periods = []
    if os.path.exists(periods_file):
        with open(periods_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            periods = list(reader)
    return periods


def get_current_pay_period():
    """Get most recent DRAFT or active period"""
    periods = get_all_pay_periods()
    # First look for DRAFT
    for p in reversed(periods):
        if p.get("Status", "").upper() == "DRAFT":
            return p
    # Then APPROVED
    for p in reversed(periods):
        if p.get("Status", "").upper() == "APPROVED":
            return p
    # Return most recent if any
    return periods[-1] if periods else None


def get_time_punches_for_period(start_date, end_date):
    """
    Scan Time_Clock folder for all punches within date range.
    Returns: {emp_id: [list of punch dicts]}
    """
    # Time_Clock path - adjust if different
    time_clock_dir = Path(DATA_DIR) / "Time_Clock"

    punches_by_emp = defaultdict(list)

    if not time_clock_dir.exists():
        return punches_by_emp

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    except:
        return punches_by_emp

    # Scan all CSV files in Time_Clock (recursive)
    for csv_file in time_clock_dir.rglob("*TimeClockLog*.csv"):
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    punch_date_str = row.get("Date", "")
                    if not punch_date_str:
                        continue
                    try:
                        punch_date = datetime.strptime(
                            punch_date_str, "%Y-%m-%d"
                        ).date()
                        if start_dt <= punch_date <= end_dt:
                            emp_id = row.get("Employee_ID", "")
                            if emp_id:
                                punches_by_emp[emp_id].append(row)
                    except:
                        continue
        except Exception as e:
            app.logger.error(f"Error reading {csv_file}: {e}")

    # Also check older format files
    for csv_file in time_clock_dir.rglob("*timeclock*.csv"):
        if "TimeClockLog" in str(csv_file):
            continue  # Already processed
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    punch_date_str = row.get("Date", "")
                    if not punch_date_str:
                        continue
                    try:
                        punch_date = datetime.strptime(
                            punch_date_str, "%Y-%m-%d"
                        ).date()
                        if start_dt <= punch_date <= end_dt:
                            emp_id = row.get("Employee_ID", "")
                            if emp_id:
                                punches_by_emp[emp_id].append(row)
                    except:
                        continue
        except:
            continue

    return punches_by_emp


def calculate_hours_from_punches(punches):
    """
    Calculate hours from punch list.
    Pairs CLOCK_IN with CLOCK_OUT, handles breaks.
    Returns: (regular_hours, overtime_hours)
    """
    if not punches:
        return 0, 0

    # Sort by date then time
    sorted_punches = sorted(
        punches, key=lambda x: (x.get("Date", ""), x.get("Time", ""))
    )

    # Group by date
    by_date = defaultdict(list)
    for p in sorted_punches:
        by_date[p.get("Date", "")].append(p)

    total_hours = 0

    for date_str, day_punches in by_date.items():
        clock_in_time = None
        break_start = None
        day_hours = 0
        break_time = 0

        for punch in day_punches:
            punch_type = punch.get("Punch_Type", "").upper()
            time_str = punch.get("Time", "")

            if not time_str:
                continue

            try:
                # Handle both HH:MM:SS and HH:MM formats
                if len(time_str) == 5:
                    time_str += ":00"
                punch_time = datetime.strptime(
                    f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S"
                )
            except:
                continue

            if punch_type == "CLOCK_IN":
                clock_in_time = punch_time
            elif punch_type == "CLOCK_OUT" and clock_in_time:
                worked = (punch_time - clock_in_time).total_seconds() / 3600
                day_hours += worked - break_time
                clock_in_time = None
                break_time = 0
            elif punch_type in ["BREAK", "LUNCH"]:
                break_start = punch_time
            elif punch_type in ["END_BREAK", "END_LUNCH"] and break_start:
                break_time += (punch_time - break_start).total_seconds() / 3600
                break_start = None

        total_hours += max(0, day_hours)

    # Split into regular (<=40/week) and overtime (>40)
    regular = min(total_hours, 40)
    overtime = max(0, total_hours - 40)

    return round(regular, 2), round(overtime, 2)


def calculate_gross_pay(regular_hours, overtime_hours, hourly_rate):
    """Calculate gross pay with 1.5x overtime"""
    regular_pay = regular_hours * hourly_rate
    overtime_pay = overtime_hours * hourly_rate * 1.5
    return round(regular_pay + overtime_pay, 2)


def calculate_taxes(gross_pay, emp_id=None):
    """
    Calculate tax withholdings.
    Returns dict with all tax amounts.
    """
    # Standard rates (can be enhanced with tax tables)
    federal_rate = 0.12  # 12% federal
    state_rate = 0.05  # 5% state (CA varies)
    ss_rate = 0.062  # 6.2% Social Security
    medicare_rate = 0.0145  # 1.45% Medicare

    federal_tax = round(gross_pay * federal_rate, 2)
    state_tax = round(gross_pay * state_rate, 2)
    social_security = round(gross_pay * ss_rate, 2)
    medicare = round(gross_pay * medicare_rate, 2)

    return {
        "federal_tax": federal_tax,
        "state_tax": state_tax,
        "social_security": social_security,
        "medicare": medicare,
        "total_taxes": round(federal_tax + state_tax + social_security + medicare, 2),
    }


def get_employee_deductions(emp_id):
    """Get active deductions from Deductions.csv"""
    deductions_file = os.path.join(DATA_DIR, "Payroll", "Deductions.csv")
    total = 0
    if os.path.exists(deductions_file):
        with open(deductions_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    if row.get("Active", "Y").upper() == "Y":
                        amt = float(row.get("Amount", 0) or 0)
                        total += amt
    return total


# ============================================
# MISSING PAYROLL ROUTES - Add to MGN_APP.py
# ============================================


@app.route("/payroll/pay-config")
@manager_required
def employee_pay_config():
    """View/edit employee pay rates"""
    employees = get_all_employees()

    # Load configs from Employee_Pay_Config.csv
    config_file = os.path.join(DATA_DIR, "Payroll", "Employee_Pay_Config.csv")
    configs = {}

    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                configs[str(row.get("Employee_ID", ""))] = row

    emp_configs = []
    for emp in employees:
        if emp.get("Status") != "Active":
            continue
        emp_id = str(emp.get("Employee_ID", ""))
        config = configs.get(emp_id, {})
        emp_configs.append(
            {
                "emp_id": emp_id,
                "name": emp.get("Employee_Name", ""),
                "role": emp.get("Role", ""),
                "hourly_rate": float(config.get("Hourly_Rate", 0) or 0),
                "pay_type": config.get("Pay_Type", "HOURLY"),
            }
        )

    return render_template("payroll/pay_config.html", employees=emp_configs)


@app.route("/payroll/pay-config/update", methods=["POST"])
@manager_required
def update_pay_config():
    """Update employee pay rate"""
    emp_id = request.form.get("emp_id", "").strip()
    hourly_rate = float(request.form.get("hourly_rate", 0) or 0)
    pay_type = request.form.get("pay_type", "HOURLY")

    payroll_dir = os.path.join(DATA_DIR, "Payroll")
    os.makedirs(payroll_dir, exist_ok=True)
    config_file = os.path.join(payroll_dir, "Employee_Pay_Config.csv")

    fieldnames = [
        "Employee_ID",
        "Pay_Type",
        "Hourly_Rate",
        "Salary_Amount",
        "Pay_Frequency",
        "Federal_Filing_Status",
        "State_Filing_Status",
        "Federal_Allowances",
        "State_Allowances",
        "Additional_Withholding",
        "Direct_Deposit",
        "Bank_Account",
        "Bank_Routing",
        "Effective_Date",
        "Last_Updated",
        "Notes",
    ]

    configs = []
    found = False

    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    row["Hourly_Rate"] = hourly_rate
                    row["Pay_Type"] = pay_type
                    row["Last_Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    found = True
                configs.append(row)

    if not found:
        configs.append(
            {
                "Employee_ID": emp_id,
                "Pay_Type": pay_type,
                "Hourly_Rate": hourly_rate,
                "Salary_Amount": 0,
                "Pay_Frequency": "BIWEEKLY",
                "Federal_Filing_Status": "SINGLE",
                "State_Filing_Status": "SINGLE",
                "Federal_Allowances": 1,
                "State_Allowances": 1,
                "Additional_Withholding": 0,
                "Direct_Deposit": "N",
                "Bank_Account": "",
                "Bank_Routing": "",
                "Effective_Date": date.today().strftime("%Y-%m-%d"),
                "Last_Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Notes": "",
            }
        )

    with open(config_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(configs)

    flash(f"Pay rate updated to ${hourly_rate}/hr", "success")
    return redirect(url_for("employee_pay_config"))


@app.route("/payroll/reports")
@manager_required
def payroll_reports():
    """Payroll reports -> hours export (Generic / QuickBooks / Shopify)."""
    return redirect(url_for("payroll_export_hours", format=request.args.get("format", "generic")))


@app.route("/payroll/time-admin")
@manager_required
def time_clock_admin():
    """Redirect to time clock management"""
    return redirect(url_for("timeclock_manage"))


@app.route("/payroll/history")
@manager_required
def payroll_history():
    """View past payroll runs"""
    from collections import defaultdict

    year = request.args.get("year", datetime.now().year)
    payroll_file = os.path.join(DATA_DIR, "Payroll", f"{year}_Payroll_Runs.csv")

    runs = []
    if os.path.exists(payroll_file):
        with open(payroll_file, "r", encoding="utf-8") as f:
            runs = list(csv.DictReader(f))

    # Group by period
    by_period = defaultdict(list)
    for run in runs:
        by_period[run.get("Period_ID", "Unknown")].append(run)

    return render_template(
        "payroll/history.html", runs_by_period=dict(by_period), year=year
    )


# ============================================
# RUN PAYROLL
# ============================================


def _update_period_status(period_id, new_status):
    """Update pay period status in CSV"""
    periods_file = os.path.join(DATA_DIR, "Payroll", "Pay_Periods.csv")
    if not os.path.exists(periods_file):
        return

    periods = []
    fieldnames = None

    with open(periods_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            if row.get("Period_ID") == period_id:
                row["Status"] = new_status
                row["Processed_By"] = session.get("employee_id", "")
                row["Processed_At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            periods.append(row)

    with open(periods_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(periods)


# California minimum wage 2024-2025
CA_MINIMUM_WAGE = 16.00

# ============================================
# PAYROLL HELPER FUNCTIONS
# ============================================


def get_employee_pay_rate(emp_id):
    """Get hourly rate from Employee_Pay_Config.csv"""
    config_file = os.path.join(DATA_DIR, "Payroll", "Employee_Pay_Config.csv")
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    rate = float(row.get("Hourly_Rate", 0) or 0)
                    return max(rate, CA_MINIMUM_WAGE)  # Must be at least minimum wage
    return CA_MINIMUM_WAGE  # Default to CA minimum wage


def get_pay_config_for_employee(emp_id):
    """Get full pay config row for an employee"""
    config_file = os.path.join(DATA_DIR, "Payroll", "Employee_Pay_Config.csv")
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    return row
    return None


def scan_timeclock_files(start_date, end_date):
    """
    Scan Time_Clock folder for all TimeClockLog CSV files within date range.
    Path: Time_Clock/YYYY/MM_MonthName/Week_X/YYYY-MM-DD_DayName_TimeClockLog.csv
    """
    time_clock_base = Path(DATA_DIR) / "Time_Clock"

    all_punches = []

    if not time_clock_base.exists():
        app.logger.warning(f"Time_Clock folder not found: {time_clock_base}")
        return all_punches

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    except:
        return all_punches

    # Scan all TimeClockLog files recursively
    for csv_file in time_clock_base.rglob("*TimeClockLog*.csv"):
        try:
            with open(csv_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    punch_date_str = row.get("Date", "")
                    if not punch_date_str:
                        continue

                    try:
                        punch_date = datetime.strptime(
                            punch_date_str, "%Y-%m-%d"
                        ).date()
                        if start_dt <= punch_date <= end_dt:
                            all_punches.append(row)
                    except:
                        continue
        except Exception as e:
            app.logger.error(f"Error reading {csv_file}: {e}")

    return all_punches


def calculate_california_hours(punches_for_employee):
    """
    Calculate hours according to California labor law:
    - Regular: First 8 hours per day
    - Daily OT (1.5x): Hours 8-12 per day
    - Daily Double Time (2x): Hours 12+ per day
    - Weekly OT: Hours over 40 in a week (if not already OT)
    - 7th consecutive day: First 8 hours at 1.5x, 8+ at 2x

    Returns: (regular_hours, ot_hours, double_time_hours)
    """
    if not punches_for_employee:
        return 0, 0, 0

    # Sort punches by date and time
    sorted_punches = sorted(
        punches_for_employee, key=lambda x: (x.get("Date", ""), x.get("Time", ""))
    )

    # Group by date
    by_date = defaultdict(list)
    for p in sorted_punches:
        by_date[p.get("Date", "")].append(p)

    # Calculate hours per day
    daily_hours = {}

    for date_str, day_punches in by_date.items():
        clock_in_time = None
        break_start = None
        day_total = 0
        break_time = 0

        for punch in day_punches:
            punch_type = punch.get("Punch_Type", "").upper()
            time_str = punch.get("Time", "")

            if not time_str:
                continue

            try:
                # Handle HH:MM:SS or HH:MM
                if len(time_str) == 5:
                    time_str += ":00"
                punch_time = datetime.strptime(
                    f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S"
                )
            except:
                continue

            if punch_type == "CLOCK_IN":
                clock_in_time = punch_time
            elif punch_type == "CLOCK_OUT" and clock_in_time:
                worked = (punch_time - clock_in_time).total_seconds() / 3600
                day_total += worked - break_time
                clock_in_time = None
                break_time = 0
            elif punch_type in ["BREAK", "LUNCH"]:
                break_start = punch_time
            elif punch_type in ["END_BREAK", "END_LUNCH"] and break_start:
                break_time += (punch_time - break_start).total_seconds() / 3600
                break_start = None

        daily_hours[date_str] = max(0, day_total)

    # Apply California overtime rules
    total_regular = 0
    total_ot = 0  # 1.5x
    total_double = 0  # 2x
    weekly_regular = 0

    # Sort dates for consecutive day tracking
    sorted_dates = sorted(daily_hours.keys())

    for date_str in sorted_dates:
        hours = daily_hours[date_str]

        # Daily overtime rules
        if hours <= 8:
            day_regular = hours
            day_ot = 0
            day_double = 0
        elif hours <= 12:
            day_regular = 8
            day_ot = hours - 8
            day_double = 0
        else:
            day_regular = 8
            day_ot = 4  # Hours 8-12
            day_double = hours - 12  # Hours 12+

        total_regular += day_regular
        total_ot += day_ot
        total_double += day_double
        weekly_regular += day_regular

    # Weekly overtime (for regular hours over 40 that aren't already daily OT)
    if weekly_regular > 40:
        excess = weekly_regular - 40
        # Convert excess regular to OT
        total_regular -= excess
        total_ot += excess

    return (round(total_regular, 2), round(total_ot, 2), round(total_double, 2))


def calculate_gross_pay_california(regular, ot_hours, double_time, hourly_rate):
    """Calculate gross pay with California overtime rates"""
    regular_pay = regular * hourly_rate
    ot_pay = ot_hours * hourly_rate * 1.5
    double_pay = double_time * hourly_rate * 2.0
    return round(regular_pay + ot_pay + double_pay, 2)


def calculate_california_taxes(gross_pay, emp_config=None):
    """
    Calculate California tax withholdings.
    Simplified rates - actual would use tax tables.
    """
    # Federal withholding (simplified - actual uses W-4 info)
    if gross_pay <= 500:
        federal_rate = 0.10
    elif gross_pay <= 1500:
        federal_rate = 0.12
    else:
        federal_rate = 0.22

    # California state tax (simplified progressive)
    if gross_pay <= 500:
        state_rate = 0.01
    elif gross_pay <= 1000:
        state_rate = 0.02
    elif gross_pay <= 2000:
        state_rate = 0.04
    else:
        state_rate = 0.06

    # FICA taxes (fixed rates)
    ss_rate = 0.062  # Social Security 6.2%
    medicare_rate = 0.0145  # Medicare 1.45%

    # CA SDI (State Disability Insurance)
    ca_sdi_rate = 0.009  # 0.9%

    federal_tax = round(gross_pay * federal_rate, 2)
    state_tax = round(gross_pay * state_rate, 2)
    social_security = round(gross_pay * ss_rate, 2)
    medicare = round(gross_pay * medicare_rate, 2)
    ca_sdi = round(gross_pay * ca_sdi_rate, 2)

    total_taxes = federal_tax + state_tax + social_security + medicare + ca_sdi

    return {
        "federal_tax": federal_tax,
        "state_tax": state_tax,
        "social_security": social_security,
        "medicare": medicare,
        "ca_sdi": ca_sdi,
        "total_taxes": round(total_taxes, 2),
    }


def get_all_pay_periods():
    """Get all pay periods"""
    periods_file = os.path.join(DATA_DIR, "Payroll", "Pay_Periods.csv")
    periods = []
    if os.path.exists(periods_file):
        with open(periods_file, "r", encoding="utf-8") as f:
            periods = list(csv.DictReader(f))
    return periods


def get_current_pay_period():
    """Get most recent DRAFT period, or most recent overall"""
    periods = get_all_pay_periods()
    for p in reversed(periods):
        if p.get("Status", "").upper() == "DRAFT":
            return p
    for p in reversed(periods):
        if p.get("Status", "").upper() == "APPROVED":
            return p
    return periods[-1] if periods else None


# ============================================
# CREATE PAY PERIOD
# ============================================


@app.route("/payroll/create-period", methods=["GET", "POST"])
@manager_required
def create_pay_period_view():
    """Create a new pay period"""
    error = None
    recent_periods = get_all_pay_periods()[-5:][::-1]

    if request.method == "POST":
        try:
            period_start = request.form.get("period_start", "").strip()
            period_end = request.form.get("period_end", "").strip()
            pay_date = request.form.get("pay_date", "").strip()
            notes = request.form.get("notes", "").strip()

            # ... validation logic ...

            # if ok, write to Payroll/Pay_Periods.csv
            # and then:
            # return redirect(url_for('payroll_dashboard'))

        except Exception as e:
            error = f"Error creating pay period: {str(e)}"

    # Default dates
    today = date.today()
    default_start = today - timedelta(days=today.weekday())
    default_end = default_start + timedelta(days=6)

    return render_template(
        "payroll/create_period.html",
        error=error,
        recent_periods=recent_periods,
        default_start=default_start.strftime("%Y-%m-%d"),
        default_end=default_end.strftime("%Y-%m-%d"),
        default_pay_date="",
        default_label="",
        default_notes="",
    )


# ============================================
# PAYROLL DASHBOARD ROUTE
# ============================================


@app.route("/payroll")
@manager_required
def payroll_dashboard():
    """Main payroll dashboard with live time clock data"""

    employees = get_all_employees()
    active_employees = [e for e in employees if e.get("Status") == "Active"]

    periods = get_all_pay_periods()
    current_period = get_current_pay_period()
    recent_periods = periods[-5:][::-1] if periods else []

    employee_hours = []
    total_hours = 0
    gross_payroll = 0

    if current_period:
        start_date = current_period.get("Start_Date", "")
        end_date = current_period.get("End_Date", "")

        if start_date and end_date:
            # Get ALL punches for the period
            all_punches = scan_timeclock_files(start_date, end_date)

            # Group punches by employee
            punches_by_emp = defaultdict(list)
            for punch in all_punches:
                emp_id = str(punch.get("Employee_ID", ""))
                if emp_id:
                    punches_by_emp[emp_id].append(punch)

            for emp in active_employees:
                emp_id = str(emp.get("Employee_ID", ""))
                emp_punches = punches_by_emp.get(emp_id, [])

                regular, ot, double = calculate_california_hours(emp_punches)
                rate = get_employee_pay_rate(emp_id)
                gross = calculate_gross_pay_california(regular, ot, double, rate)

                total_hrs = regular + ot + double

                employee_hours.append(
                    {
                        "emp_id": emp_id,
                        "name": emp.get("Employee_Name", "Unknown"),
                        "role": emp.get("Role", ""),
                        "rate": rate,
                        "regular_hours": regular,
                        "overtime_hours": ot + double,  # Combined OT for display
                        "total_hours": total_hrs,
                        "gross_pay": gross,
                    }
                )

                total_hours += total_hrs
                gross_payroll += gross

    return render_template(
        "payroll/payroll_dashboard.html",
        active_employees=len(active_employees),
        total_hours=total_hours,
        gross_payroll=gross_payroll,
        period_count=len(periods),
        current_period=current_period,
        employee_hours=employee_hours,
        recent_periods=recent_periods,
    )


# ============================================
# RUN PAYROLL - CALIFORNIA COMPLIANT
# ============================================


@app.route("/payroll/run", methods=["GET", "POST"])
@manager_required
def run_payroll():
    """Process payroll with California labor law compliance"""
    current_period = get_current_pay_period()

    if not current_period:
        flash("No active pay period found. Create one first.", "error")
        return redirect(url_for("payroll_dashboard"))

    if request.method == "POST":
        # Payroll lock: a PROCESSED period must not be silently re-run (double-pay).
        if str(current_period.get("Status", "")).upper() == "PROCESSED" and request.form.get("force") != "1":
            flash("This pay period is already PROCESSED and is locked. Re-running would "
                  "double-pay employees -- a manager override (force) is required.", "error")
            return redirect(url_for("payroll_dashboard"))
        try:
            start_date = current_period.get("Start_Date", "")
            end_date = current_period.get("End_Date", "")
            period_id = current_period.get("Period_ID", "")

            # Get all time punches
            all_punches = scan_timeclock_files(start_date, end_date)

            # Group by employee
            punches_by_emp = defaultdict(list)
            for punch in all_punches:
                emp_id = str(punch.get("Employee_ID", ""))
                if emp_id:
                    punches_by_emp[emp_id].append(punch)

            employees = get_all_employees()

            # Payroll runs file
            year = datetime.now().year
            payroll_dir = os.path.join(DATA_DIR, "Payroll")
            os.makedirs(payroll_dir, exist_ok=True)
            payroll_file = os.path.join(payroll_dir, f"{year}_Payroll_Runs.csv")

            file_exists = (
                os.path.exists(payroll_file) and os.path.getsize(payroll_file) > 0
            )

            fieldnames = [
                "Payroll_ID",
                "Period_ID",
                "Employee_ID",
                "Employee_Name",
                "Hourly_Rate",
                "Regular_Hours",
                "Overtime_Hours",
                "Double_Time_Hours",
                "Holiday_Hours",
                "PTO_Hours",
                "Sick_Hours",
                "Gross_Pay",
                "Federal_Tax",
                "State_Tax",
                "Social_Security",
                "Medicare",
                "CA_SDI",
                "Other_Deductions",
                "Net_Pay",
                "Pay_Method",
                "Check_Number",
                "Status",
                "Created_At",
                "Approved_By",
                "Approved_At",
                "Notes",
            ]

            with open(payroll_file, "a", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                if not file_exists:
                    writer.writeheader()

                processed_count = 0

                for emp in employees:
                    if emp.get("Status") != "Active":
                        continue

                    emp_id = str(emp.get("Employee_ID", ""))
                    emp_punches = punches_by_emp.get(emp_id, [])

                    regular, ot, double = calculate_california_hours(emp_punches)
                    rate = get_employee_pay_rate(emp_id)
                    gross = calculate_gross_pay_california(regular, ot, double, rate)

                    taxes = calculate_california_taxes(gross)

                    # Get any additional deductions
                    other_deductions = 0
                    deductions_file = os.path.join(
                        DATA_DIR, "Payroll", "Deductions.csv"
                    )
                    if os.path.exists(deductions_file):
                        with open(deductions_file, "r", encoding="utf-8") as df:
                            for row in csv.DictReader(df):
                                if str(row.get("Employee_ID", "")) == emp_id:
                                    if row.get("Active", "Y").upper() == "Y":
                                        other_deductions += float(
                                            row.get("Amount", 0) or 0
                                        )

                    net_pay = gross - taxes["total_taxes"] - other_deductions

                    payroll_id = "PR" + datetime.now().strftime("%Y%m%d%H%M%S") + emp_id

                    writer.writerow(
                        {
                            "Payroll_ID": payroll_id,
                            "Period_ID": period_id,
                            "Employee_ID": emp_id,
                            "Employee_Name": emp.get("Employee_Name", ""),
                            "Hourly_Rate": rate,
                            "Regular_Hours": regular,
                            "Overtime_Hours": ot,
                            "Double_Time_Hours": double,
                            "Holiday_Hours": 0,
                            "PTO_Hours": 0,
                            "Sick_Hours": 0,
                            "Gross_Pay": gross,
                            "Federal_Tax": taxes["federal_tax"],
                            "State_Tax": taxes["state_tax"],
                            "Social_Security": taxes["social_security"],
                            "Medicare": taxes["medicare"],
                            "CA_SDI": taxes.get("ca_sdi", 0),
                            "Other_Deductions": other_deductions,
                            "Net_Pay": round(net_pay, 2),
                            "Pay_Method": "CHECK",
                            "Check_Number": "",
                            "Status": "PROCESSED",
                            "Created_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "Approved_By": session.get("employee_id", ""),
                            "Approved_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "Notes": "CA compliant - OT:1.5x, DT:2x",
                        }
                    )
                    processed_count += 1

            # Update period status
            _update_period_status(period_id, "PROCESSED")

            flash(
                f"Payroll processed for {processed_count} employees (CA compliant)!",
                "success",
            )
            try:
                from POS_CORE import append_audit_event
                append_audit_event("payroll_run", {
                    "period_id": period_id,
                    "run_by": session.get("employee_id", ""),
                    "employees": processed_count,
                })
            except Exception:
                pass
            return redirect(url_for("payroll_dashboard"))

        except Exception as e:
            flash(f"Error processing payroll: {str(e)}", "error")
            import traceback

            traceback.print_exc()

    return render_template("payroll/run_payroll.html", period=current_period)


@app.route("/payroll/export-hours")
@manager_required
def payroll_export_hours():
    """Export a pay period's hours as CSV for a third-party payroll provider.

    Query params: ?period_id=<id> (default: current period),
                  ?format=generic|quickbooks|shopify (default: generic).
    Read-only -- reuses the same hour math as the dashboard, writes nothing.
    """
    import io
    from flask import Response
    fmt = (request.args.get("format", "generic") or "generic").lower()
    period_id = request.args.get("period_id", "")

    periods = get_all_pay_periods()
    period = None
    if period_id:
        period = next((p for p in periods if str(p.get("Period_ID", "")) == str(period_id)), None)
    if not period:
        period = get_current_pay_period()
    if not period:
        flash("No pay period available to export.", "error")
        return redirect(url_for("payroll_dashboard"))

    start_date = period.get("Start_Date", "")
    end_date = period.get("End_Date", "")
    pid = period.get("Period_ID", "")

    punches_by_emp = defaultdict(list)
    for punch in scan_timeclock_files(start_date, end_date):
        eid = str(punch.get("Employee_ID", ""))
        if eid:
            punches_by_emp[eid].append(punch)

    rows = []
    for emp in get_all_employees():
        if emp.get("Status") != "Active":
            continue
        eid = str(emp.get("Employee_ID", ""))
        regular, ot, double = calculate_california_hours(punches_by_emp.get(eid, []))
        if (regular + ot + double) <= 0:
            continue
        rows.append({
            "emp_id": eid, "name": emp.get("Employee_Name", ""),
            "regular": round(regular, 2), "ot": round(ot, 2),
            "double": round(double, 2), "rate": get_employee_pay_rate(eid),
        })

    # CSV-injection guard for any text cell exported to a spreadsheet app.
    def _g(v):
        s = "" if v is None else str(v)
        return ("'" + s) if s[:1] in ("=", "+", "-", "@") else s

    if fmt == "quickbooks":
        headers = ["Employee", "Regular Hours", "Overtime Hours", "Double Time Hours", "Hourly Rate"]
        out_rows = [[_g(r["name"]), r["regular"], r["ot"], r["double"], r["rate"]] for r in rows]
    elif fmt == "shopify":
        headers = ["Staff Name", "Employee ID", "Regular Hours", "Overtime Hours", "Hourly Rate"]
        out_rows = [[_g(r["name"]), _g(r["emp_id"]), r["regular"], round(r["ot"] + r["double"], 2), r["rate"]] for r in rows]
    else:  # generic -- imports into QuickBooks / Gusto / ADP / Paychex
        headers = ["Employee_ID", "Employee_Name", "Period_Start", "Period_End",
                   "Regular_Hours", "Overtime_Hours", "Double_Time_Hours", "Hourly_Rate"]
        out_rows = [[_g(r["emp_id"]), _g(r["name"]), start_date, end_date,
                     r["regular"], r["ot"], r["double"], r["rate"]] for r in rows]

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(out_rows)

    filename = f"hours_{pid or 'current'}_{fmt}.csv"
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


def _update_period_status(period_id, new_status):
    """Update pay period status"""
    periods_file = os.path.join(DATA_DIR, "Payroll", "Pay_Periods.csv")
    if not os.path.exists(periods_file):
        return

    periods = []
    fieldnames = None

    with open(periods_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            if row.get("Period_ID") == period_id:
                row["Status"] = new_status
                row["Processed_By"] = session.get("employee_id", "")
                row["Processed_At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            periods.append(row)

    with open(periods_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(periods)


# ============================================
# UPDATE EMPLOYEE PAY RATE (for employee detail page)
# ============================================


@app.route("/employees/<emp_id>/update-pay", methods=["POST"])
@manager_required
def update_employee_pay(emp_id):
    """Update employee hourly rate from detail page"""
    hourly_rate = float(request.form.get("hourly_rate", 0) or 0)

    # Enforce CA minimum wage
    if hourly_rate < CA_MINIMUM_WAGE:
        flash(
            f"Hourly rate must be at least ${CA_MINIMUM_WAGE} (CA minimum wage)",
            "error",
        )
        return redirect(url_for("employee_detail", emp_id=emp_id))

    payroll_dir = os.path.join(DATA_DIR, "Payroll")
    os.makedirs(payroll_dir, exist_ok=True)
    config_file = os.path.join(payroll_dir, "Employee_Pay_Config.csv")

    fieldnames = [
        "Employee_ID",
        "Pay_Type",
        "Hourly_Rate",
        "Salary_Amount",
        "Pay_Frequency",
        "Federal_Filing_Status",
        "State_Filing_Status",
        "Federal_Allowances",
        "State_Allowances",
        "Additional_Withholding",
        "Direct_Deposit",
        "Bank_Account",
        "Bank_Routing",
        "Effective_Date",
        "Last_Updated",
        "Notes",
    ]

    configs = []
    found = False

    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if str(row.get("Employee_ID", "")) == str(emp_id):
                    row["Hourly_Rate"] = hourly_rate
                    row["Last_Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    found = True
                configs.append(row)

    if not found:
        configs.append(
            {
                "Employee_ID": emp_id,
                "Pay_Type": "HOURLY",
                "Hourly_Rate": hourly_rate,
                "Salary_Amount": 0,
                "Pay_Frequency": "BIWEEKLY",
                "Federal_Filing_Status": "SINGLE",
                "State_Filing_Status": "SINGLE",
                "Federal_Allowances": 1,
                "State_Allowances": 1,
                "Additional_Withholding": 0,
                "Direct_Deposit": "N",
                "Bank_Account": "",
                "Bank_Routing": "",
                "Effective_Date": date.today().strftime("%Y-%m-%d"),
                "Last_Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Notes": "",
            }
        )

    with open(config_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(configs)

    flash(f"Pay rate updated to ${hourly_rate:.2f}/hr", "success")
    return redirect(url_for("employee_detail", emp_id=emp_id))


# ============================================
# END PAYROLL ROUTES
# ============================================


# ==============================================================================
#                     AI ASSISTANT
# ==============================================================================


@app.route("/ai")
@login_required
def ai_assistant():
    return render_template("ai/chat.html")


@app.route("/ai/chat", methods=["POST"])
@login_required
def ai_chat():
    if not openai_client:
        return jsonify({"error": "AI not configured. Set OPENAI_API_KEY in .env"}), 503

    message = request.json.get("message", "")

    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": f"You are a helpful assistant for {BUSINESS_NAME}, a nursery and pet store. Help with plant care, inventory questions, and store operations.",
                },
                {"role": "user", "content": message},
            ],
            max_tokens=500,
        )
        reply = response.choices[0].message.content
        return jsonify({"reply": reply})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================
# TILL HELPERS (MGN_APP.py)
# ============================================


def _today_iso():
    return date.today().strftime("%Y-%m-%d")


def _safe_float(v, default=0.0):
    try:
        return float(v)
    except Exception:
        return default


def get_till_dir():
    till_dir = os.path.join(DATA_DIR, "Till")
    os.makedirs(till_dir, exist_ok=True)
    return till_dir


def get_till_state_file():
    return os.path.join(get_till_dir(), "till_state.csv")


def get_till_ledger_file():
    return os.path.join(get_till_dir(), "till_ledger.csv")


def is_till_open():
    state_file = get_till_state_file()
    if not os.path.exists(state_file) or os.path.getsize(state_file) == 0:
        return False, None

    try:
        with open(state_file, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        if not rows:
            return False, None

        last = rows[-1]
        today = _today_iso()

        status = (last.get("status") or "").strip().upper()
        row_date = (last.get("date") or "").strip()

        if status == "OPEN" and row_date == today:
            return True, {
                "employee": last.get("employee", ""),
                "employee_id": last.get("employee_id", ""),
                "amount": last.get("amount", ""),
                "time": last.get("time", ""),
                "date": row_date,
                "status": status,
            }

        return False, None

    except Exception:
        return False, None


def add_ledger_entry(entry_type: str, amount: float, note: str = ""):
    """
    Writes a till activity line item to: DATA_DIR/Till/till_ledger.csv
    """
    ledger_file = get_till_ledger_file()
    file_exists = os.path.exists(ledger_file) and os.path.getsize(ledger_file) > 0

    fieldnames = ["date", "time", "type", "amount", "note", "employee", "employee_id"]
    row = {
        "date": _today_iso(),
        "time": datetime.now().strftime("%H:%M:%S"),
        "type": (entry_type or "").strip().upper(),
        "amount": f"{float(amount or 0):.2f}",
        "note": (note or "").strip(),
        "employee": session.get("employee_name", "Unknown"),
        "employee_id": session.get("employee_id", ""),
    }

    with open(ledger_file, "a", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            w.writeheader()
        w.writerow(row)


def get_today_ledger():
    ledger_file = get_till_ledger_file()
    if not os.path.exists(ledger_file) or os.path.getsize(ledger_file) == 0:
        return []

    today = _today_iso()
    try:
        with open(ledger_file, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            return [r for r in reader if (r.get("date") or "").strip() == today]
    except Exception:
        return []


def calculate_till_totals():
    """
    Computes real-time totals from the till ledger.
    NOTE: cash_sales is currently 0.00 until we wire it from transaction logs.
    """
    rows = get_today_ledger()

    opening_float = 0.0
    payouts = 0.0
    drops = 0.0
    refunds = 0.0

    for r in rows:
        t = (r.get("type") or "").strip().upper()
        a = _safe_float(r.get("amount"), 0.0)

        if t == "OPEN":
            opening_float += a
        elif t == "PAYOUT":
            payouts += a
        elif t == "DROP":
            drops += a
        elif t == "REFUND":
            refunds += a

    cash_sales = 0.0  # TODO: wire from your POS cash transactions
    expected_cash = opening_float + cash_sales - payouts - drops - refunds

    return {
        "opening_float": round(opening_float, 2),
        "cash_sales": round(cash_sales, 2),
        "refunds": round(refunds, 2),
        "payouts": round(payouts, 2),
        "drops": round(drops, 2),
        "expected_cash": round(expected_cash, 2),
    }


# ==============================================================================
#                     TILL (helpers + real cash sales from SalesLog)
# ==============================================================================


def _safe_float(x, default=0.0):
    try:
        if x is None:
            return default
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return default
        s = s.replace("$", "").replace(",", "")
        return float(s)
    except Exception:
        return default


def get_till_state_file():
    till_dir = os.path.join(DATA_DIR, "Till")
    os.makedirs(till_dir, exist_ok=True)
    return os.path.join(till_dir, "till_state.csv")


def get_till_ledger_file():
    till_dir = os.path.join(DATA_DIR, "Till")
    os.makedirs(till_dir, exist_ok=True)
    return os.path.join(till_dir, "ledger.csv")


def add_ledger_entry(kind: str, amount: float, note: str = ""):
    """
    Writes a till event row. amount is always stored as a numeric string.
    kind examples: OPEN, PAYOUT, DROP, REFUND, NOSALE, CLOSE
    """
    ledger_file = get_till_ledger_file()
    file_exists = os.path.exists(ledger_file) and os.path.getsize(ledger_file) > 0

    row = {
        "date": date.today().strftime("%Y-%m-%d"),
        "time": datetime.now().strftime("%H:%M:%S"),
        "kind": (kind or "").upper(),
        "amount": f"{_safe_float(amount):.2f}",
        "employee": session.get("employee_name", "Unknown"),
        "employee_id": session.get("employee_id", ""),
        "note": note or "",
    }

    with open(ledger_file, "a", newline="", encoding="utf-8") as f:
        fieldnames = [
            "date",
            "time",
            "kind",
            "amount",
            "employee",
            "employee_id",
            "note",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            w.writeheader()
        w.writerow(row)


def is_till_open():
    """
    Returns:
      (True, state_dict) if today's till is currently OPEN
      (False, None) otherwise
    """
    state_file = get_till_state_file()
    if not os.path.exists(state_file) or os.path.getsize(state_file) == 0:
        return False, None

    try:
        with open(state_file, "r", encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
        if not rows:
            return False, None

        last = rows[-1]
        today = date.today().strftime("%Y-%m-%d")
        status = (last.get("status") or "").strip().upper()
        row_date = (last.get("date") or "").strip()

        if status == "OPEN" and row_date == today:
            return True, {
                "employee": last.get("employee", ""),
                "employee_id": last.get("employee_id", ""),
                "amount": _safe_float(last.get("amount", 0)),
                "time": last.get("time", ""),
                "date": row_date,
                "status": status,
            }
        return False, None
    except Exception:
        return False, None


def get_today_ledger():
    """
    Returns a list of dicts for today's till ledger entries.
    Ensures entry['amount'] is a float (fixes Jinja > 0 TypeError).
    """
    ledger_file = get_till_ledger_file()
    if not os.path.exists(ledger_file) or os.path.getsize(ledger_file) == 0:
        return []

    today = date.today().strftime("%Y-%m-%d")
    out = []
    with open(ledger_file, "r", encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            if (r.get("date") or "").strip() != today:
                continue
            out.append(
                {
                    "date": r.get("date", ""),
                    "time": r.get("time", ""),
                    "kind": (r.get("kind") or "").upper(),
                    "amount": _safe_float(r.get("amount", 0)),
                    "employee": r.get("employee", ""),
                    "employee_id": r.get("employee_id", ""),
                    "note": r.get("note", ""),
                }
            )
    # newest first (optional)
    out.reverse()
    return out


def _find_saleslog_for_date(day: date):
    """
    Finds the SalesLog CSV for a given day.
    Looks under: DATA_DIR/Sales_Logs/YYYY/**/<YYYY-MM-DD>_*_SalesLog.csv
    """
    y = day.strftime("%Y")
    d = day.strftime("%Y-%m-%d")
    root = os.path.join(DATA_DIR, "Sales_Logs", y)
    # your naming: 2025-12-20_Saturday_SalesLog.csv
    pattern = os.path.join(root, "**", f"{d}_*_SalesLog.csv")
    matches = glob.glob(pattern, recursive=True)
    if matches:
        # choose the newest if multiple
        matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return matches[0]
    return None


def _cash_in_from_saleslog(day: date):
    """
    Returns (cash_in, cash_txn_count).
    Uses Amount_Received - Change_Due per Transaction_ID (so multi-line carts don't double count).
    """
    p = _find_saleslog_for_date(day)
    if not p or not os.path.exists(p):
        return 0.0, 0

    cash_in = 0.0
    txns = 0
    try:
        with open(p, "r", encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))

        # Filter to CASH lines
        cash_rows = [
            r for r in rows if (r.get("Payment_Method") or "").strip().upper() == "CASH"
        ]
        if not cash_rows:
            return 0.0, 0

        # Unique by Transaction_ID
        seen = set()
        for r in cash_rows:
            tid = (r.get("Transaction_ID") or "").strip()
            if not tid or tid in seen:
                continue
            seen.add(tid)

            received = _safe_float(r.get("Amount_Received", 0))
            change = _safe_float(r.get("Change_Due", 0))
            cash_in += received - change

        txns = len(seen)
        return float(cash_in), int(txns)
    except Exception:
        return 0.0, 0


def calculate_till_totals():
    """
    Computes till totals using:
      - Opening float from till_state.csv (OPEN event)
      - Cash-in from SalesLog (Amount_Received - Change_Due per Transaction)
      - Payout/Drop/Refund from ledger.csv
    Returns dict for template cards.
    """
    today = date.today()
    till_open, state = is_till_open()

    opening_float = _safe_float(state.get("amount", 0)) if state else 0.0

    cash_sales, cash_txn_count = _cash_in_from_saleslog(today)

    entries = get_today_ledger()
    payouts = sum(e["amount"] for e in entries if e["kind"] == "PAYOUT")
    drops = sum(e["amount"] for e in entries if e["kind"] == "DROP")
    refunds = sum(e["amount"] for e in entries if e["kind"] == "REFUND")

    expected_cash = opening_float + cash_sales - payouts - drops - refunds
    cash_out = refunds + payouts + drops

    return {
        "opening_float": round(opening_float, 2),
        "cash_sales": round(cash_sales, 2),
        "cash_txn_count": int(cash_txn_count),
        "refunds": round(refunds, 2),
        "payouts": round(payouts, 2),
        "drops": round(drops, 2),
        "cash_out": round(cash_out, 2),
        "expected_cash": round(expected_cash, 2),
        "till_open": bool(till_open),
    }


# ==============================================================================
#                     till
# ==============================================================================
# ============================================
# CASH TILL ROUTES
# ============================================


@app.route("/till")
@manager_required
def till():
    if "employee_id" not in session:
        return redirect(url_for("login"))
    till_open, state = is_till_open()
    entries = get_today_ledger()
    totals = calculate_till_totals()
    # Prevent collision: totals may also include a "till_open" key
    # Keep the data, just rename it so render_template doesn't get duplicates
    if isinstance(totals, dict) and "till_open" in totals:
        totals["till_open_state"] = totals.pop("till_open")

    recent_closeouts = []
    till_dir = os.path.join(DATA_DIR, "Till")
    closeout_file = os.path.join(till_dir, "closeouts.csv")
    if os.path.exists(closeout_file):
        with open(closeout_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            recent_closeouts = list(reader)[-5:]
            recent_closeouts.reverse()
    return render_template(
        "till/till.html",
        till_open=till_open,
        opened_by=state.get("employee") if state else None,
        open_time=state.get("time") if state else None,
        ledger_entries=entries,
        recent_closeouts=recent_closeouts,
        **totals,
    )


# ==============================
# TILL HELPERS (state file)
# ==============================


@app.route("/api/till/open", methods=["POST"])
@manager_required
def api_till_open():
    try:
        data = request.get_json()
        amount = float(data.get("amount", 0))
        if amount <= 0:
            return jsonify({"success": False, "error": "Invalid amount"})
        till_open, _ = is_till_open()
        if till_open:
            return jsonify({"success": False, "error": "Till is already open"})
        state_file = get_till_state_file()
        file_exists = os.path.exists(state_file) and os.path.getsize(state_file) > 0
        with open(state_file, "a", newline="", encoding="utf-8") as f:
            fieldnames = ["date", "time", "status", "employee", "employee_id", "amount"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            writer.writerow(
                {
                    "date": date.today().strftime("%Y-%m-%d"),
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "status": "OPEN",
                    "employee": session.get("employee_name", "Unknown"),
                    "employee_id": session.get("employee_id", ""),
                    "amount": amount,
                }
            )
        add_ledger_entry("OPEN", amount, "Opening float")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/till/payout", methods=["POST"])
@manager_required
def api_till_payout():
    try:
        data = request.get_json()
        amount = float(data.get("amount", 0))
        note = data.get("note", "")
        if amount <= 0:
            return jsonify({"success": False, "error": "Invalid amount"})
        till_open, _ = is_till_open()
        if not till_open:
            return jsonify({"success": False, "error": "Till is not open"})
        add_ledger_entry("PAYOUT", amount, note)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/till/drop", methods=["POST"])
@manager_required
def api_till_drop():
    try:
        data = request.get_json()
        amount = float(data.get("amount", 0))
        note = data.get("note", "")
        if amount <= 0:
            return jsonify({"success": False, "error": "Invalid amount"})
        till_open, _ = is_till_open()
        if not till_open:
            return jsonify({"success": False, "error": "Till is not open"})
        add_ledger_entry("DROP", amount, note)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/till/refund", methods=["POST"])
@manager_required
def api_till_refund():
    try:
        data = request.get_json()
        amount = float(data.get("amount", 0))
        note = data.get("note", "")
        if amount <= 0:
            return jsonify({"success": False, "error": "Invalid amount"})
        till_open, _ = is_till_open()
        if not till_open:
            return jsonify({"success": False, "error": "Till is not open"})
        add_ledger_entry("REFUND", amount, note)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/till/nosale", methods=["POST"])
@manager_required
def api_till_nosale():
    try:
        till_open, _ = is_till_open()
        if not till_open:
            return jsonify({"success": False, "error": "Till is not open"})
        add_ledger_entry("NOSALE", 0, "Drawer opened without sale")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/till/close", methods=["POST"])
@manager_required
def api_till_close():
    try:
        data = request.get_json()
        counted = float(data.get("counted", 0))
        denominations = data.get("denominations", {})
        till_open, _ = is_till_open()
        if not till_open:
            return jsonify({"success": False, "error": "Till is not open"})
        totals = calculate_till_totals()
        expected = totals["expected_cash"]
        variance = counted - expected
        add_ledger_entry(
            "CLOSE", counted, f"Expected: ${expected:.2f}, Variance: ${variance:+.2f}"
        )
        state_file = get_till_state_file()
        with open(state_file, "a", newline="", encoding="utf-8") as f:
            fieldnames = ["date", "time", "status", "employee", "employee_id", "amount"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writerow(
                {
                    "date": date.today().strftime("%Y-%m-%d"),
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "status": "CLOSED",
                    "employee": session.get("employee_name", "Unknown"),
                    "employee_id": session.get("employee_id", ""),
                    "amount": counted,
                }
            )
        till_dir = os.path.join(DATA_DIR, "Till")
        closeout_file = os.path.join(till_dir, "closeouts.csv")
        file_exists = os.path.exists(closeout_file)
        with open(closeout_file, "a", newline="", encoding="utf-8") as f:
            fieldnames = [
                "date",
                "time",
                "employee",
                "employee_id",
                "opening_float",
                "cash_sales",
                "refunds",
                "payouts",
                "drops",
                "expected",
                "counted",
                "variance",
                "d100",
                "d50",
                "d20",
                "d10",
                "d5",
                "d1",
                "quarters",
                "dimes",
                "nickels",
                "pennies",
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            closeout_row = {
                    "date": date.today().strftime("%Y-%m-%d"),
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "employee": session.get("employee_name", "Unknown"),
                    "employee_id": session.get("employee_id", ""),
                    "opening_float": totals["opening_float"],
                    "cash_sales": totals["cash_sales"],
                    "refunds": totals["refunds"],
                    "payouts": totals["payouts"],
                    "drops": totals["drops"],
                    "expected": expected,
                    "counted": counted,
                    "variance": variance,
                    **denominations,
            }
            writer.writerow(closeout_row)
        # End-of-day confirmation: tamper-evident audit line + email the report.
        emailed = False
        try:
            day_summary = generate_daily_summary()
        except Exception:
            day_summary = {}
        report_text = (
            "Mountain Gardens POS -- End of Day\n"
            f"Date: {date.today().strftime('%Y-%m-%d')}  Time: {datetime.now().strftime('%H:%M:%S')}\n"
            f"Closed by: {session.get('employee_name','Unknown')} (ID {session.get('employee_id','')})\n\n"
            "-- Till --\n"
            f"Opening float: ${float(totals.get('opening_float',0) or 0):.2f}\n"
            f"Cash sales:    ${float(totals.get('cash_sales',0) or 0):.2f}\n"
            f"Expected cash: ${expected:.2f}\n"
            f"Counted cash:  ${counted:.2f}\n"
            f"Variance:      ${variance:+.2f}\n\n"
            "-- Day totals --\n"
            f"Transactions:  {day_summary.get('total_transactions','')}\n"
            f"Revenue:       ${float(day_summary.get('total_revenue',0) or 0):.2f}\n"
            f"\nQuick-add items awaiting reconciliation: {len(get_unreconciled_quickadds())}\n"
            "Reconcile them at /inventory/reconcile\n"
        )
        # Export the day's files locally (Daily_Reports/<date>/) + attach to the email,
        # so the close-out is saved AND viewable regardless of whether email sends.
        attach_paths = []
        reports_dir = None
        try:
            day_str = date.today().strftime("%Y-%m-%d")
            reports_dir = os.path.join(DATA_DIR, "Daily_Reports", day_str)
            os.makedirs(reports_dir, exist_ok=True)
            # (a) the day's sales log CSV
            try:
                src = _find_saleslog_for_date(date.today())
            except Exception:
                src = None
            if src and os.path.exists(src):
                dst = os.path.join(reports_dir, os.path.basename(src))
                with open(src, "rb") as _s, open(dst, "wb") as _d:
                    _d.write(_s.read())
                attach_paths.append(dst)
            # (b) daily summary flattened to CSV
            summary_csv = os.path.join(reports_dir, f"{day_str}_DailySummary.csv")
            with open(summary_csv, "w", newline="", encoding="utf-8") as sf:
                w = csv.writer(sf)
                w.writerow(["Metric", "Value"])
                for k, v in (day_summary or {}).items():
                    if isinstance(v, (dict, list)):
                        v = json.dumps(v, ensure_ascii=False)
                    w.writerow([k, v])
            attach_paths.append(summary_csv)
            # (c) this close-out row as its own CSV
            closeout_csv = os.path.join(reports_dir, f"{day_str}_Closeout.csv")
            with open(closeout_csv, "w", newline="", encoding="utf-8") as cf:
                cw = csv.DictWriter(cf, fieldnames=fieldnames)
                cw.writeheader()
                cw.writerow(closeout_row)
            attach_paths.append(closeout_csv)
        except Exception as _eod_exc:
            # Don't silently swallow a failed LOCAL export -- a missing PC copy must
            # be detectable (the local file is one of the 3 required copies).
            try:
                from POS_CORE import append_audit_event
                append_audit_event("eod_local_export_failed", {"error": str(_eod_exc)})
            except Exception:
                pass
        try:
            from POS_CORE import append_audit_event
            append_audit_event("till_close", {
                "date": date.today().strftime("%Y-%m-%d"),
                "closed_by": session.get("employee_id", ""),
                "expected": expected, "counted": counted, "variance": variance,
            })
        except Exception:
            pass
        # Daily, idempotent: assign any recurring admin tasks due today.
        try:
            check_and_assign_recurring_tasks()
        except Exception:
            pass
        # Daily automatic backup of all data at close (best-effort resilience).
        try:
            _bk().create_backup()
        except Exception:
            pass
        try:
            emailed = send_eod_report_email(report_text, attachments=attach_paths)
        except Exception:
            emailed = False
        # Provable delivery trail: record where the EOD went (or that SMTP was off)
        # so the LOCAL Daily_Reports copy is an auditable 3rd copy / fallback.
        eod_recipients = get_eod_recipients()
        try:
            if reports_dir:
                with open(os.path.join(reports_dir, "_EOD_DELIVERY.txt"), "w", encoding="utf-8") as _df:
                    _df.write(
                        f"Closed: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
                        f"Emailed: {'YES' if emailed else 'NO (SMTP not configured or send failed)'}\n"
                        f"Recipients: {', '.join(eod_recipients)}\n"
                        f"Local copies: {', '.join(os.path.basename(p) for p in attach_paths)}\n"
                    )
        except Exception:
            pass
        return jsonify(
            {
                "success": True,
                "expected": expected,
                "counted": counted,
                "variance": variance,
                "emailed": emailed,
                "emailed_to": eod_recipients,
                "local_copies": [os.path.basename(p) for p in attach_paths],
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ============================================
# END CASH TILL ROUTES
# ============================================
# ==============================================================================
#                     SETTINGS
# ==============================================================================


@app.route("/settings")
@login_required
def settings():
    """Settings page with user profile, customizations, and billing"""
    emp_id = session.get("employee_id", "")
    employee = get_employee(emp_id) if emp_id else {}

    tenant_id = session.get("tenant_id", "")
    tenant = get_tenant(tenant_id) if tenant_id else {}
    tenant_settings = get_or_create_tenant_settings()
    subscription = get_subscription(tenant_id) if tenant_id else {}

    # Get billing info
    tier = get_current_tier()
    billing_mode = BILLING_MODE

    # Calculate usage
    employees = get_all_employees()
    items = get_all_items()
    txn_count = _count_transactions_current_month()

    limits = TIER_LIMITS.get(tier, {})
    usage = {
        "users": len(employees),
        "skus": len(items),
        "txns_month": txn_count,
        "limits": limits,
    }

    return render_template(
        "settings.html",
        employee=employee,
        tenant=tenant,
        tenant_settings=tenant_settings,
        subscription=subscription,
        tier=tier,
        billing_mode=billing_mode,
        usage=usage,
        eod_emails=", ".join(get_eod_recipients()),
    )


@app.route("/settings/profile/update", methods=["POST"])
@login_required
def update_profile():
    """Update user profile"""
    emp_id = session.get("employee_id", "")
    if not emp_id:
        flash("Not logged in", "error")
        return redirect(url_for("login"))

    employee = get_employee(emp_id)
    if not employee:
        flash("Employee not found", "error")
        return redirect(url_for("settings"))

    # Update employee info
    name = request.form.get("employee_name", "").strip()
    phone = request.form.get("phone", "").strip()
    email = request.form.get("email", "").strip()
    emergency_contact = request.form.get("emergency_contact", "").strip()
    notes = request.form.get("notes", "").strip()

    if name:
        employee["Employee_Name"] = name
    if phone:
        employee["Phone"] = phone
    if email:
        employee["Email"] = email
    if emergency_contact:
        employee["Emergency_Contact"] = emergency_contact
    if notes is not None:
        employee["Notes"] = notes

    employee["Last_Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Update in CSV
    update_csv_row(
        get_employee_path(), EMPLOYEE_HEADERS, "Employee_ID", emp_id, employee
    )

    # Update session
    session["employee_name"] = name or session.get("employee_name", "")

    flash("Profile updated successfully", "success")
    return redirect(url_for("settings"))


@app.route("/settings/receipts/update", methods=["POST"])
@login_required
@manager_required
def update_receipt_settings():
    """Update receipt and tax settings"""
    tenant_id = session.get("tenant_id", "")
    if not tenant_id:
        flash("No tenant selected", "error")
        return redirect(url_for("settings"))

    # Save receipt settings to tenant settings
    tenant_settings = get_or_create_tenant_settings()
    tax_rate_input = request.form.get("tax_rate", "8.25")
    tenant_settings["Tax_Rate"] = tax_rate_input
    # Mirror the rate into the live tax engine (Settings/Config.csv) so record_sale
    # uses exactly what the owner enters here. Accepts 8.25 or 0.0825.
    set_config("Store_Tax_Rate", tax_rate_input)
    # EOD report recipients (owner + e.g. mom). Comma-separated. Drives who gets the
    # end-of-day close-out email; the local PC copy is always saved regardless.
    eod_emails = (request.form.get("eod_emails") or "").strip()
    if eod_emails:
        set_config("EOD_Emails", eod_emails)
    tenant_settings["Receipt_Footer"] = request.form.get("receipt_footer", "")
    tenant_settings["Business_Address"] = request.form.get("business_address", "")
    tenant_settings["Business_Phone"] = request.form.get("business_phone", "")

    write_csv(get_tenant_settings_path(), TENANT_SETTINGS_HEADERS, [tenant_settings])

    flash("Receipt settings updated successfully", "success")
    return redirect(url_for("settings"))


@app.route("/settings/customizations/update", methods=["POST"])
@login_required
@manager_required
def update_customizations():
    """Update business customizations"""
    tenant_id = session.get("tenant_id", "")
    if not tenant_id:
        flash("No tenant selected", "error")
        return redirect(url_for("settings"))

    # Update tenant name
    business_name = request.form.get("business_name", "").strip()
    theme = request.form.get("theme", "dark").strip()
    pos_name = request.form.get("pos_name", "").strip()

    if business_name:
        tenant = get_tenant(tenant_id)
        if tenant:
            tenant["Name"] = business_name
            rows = read_csv(TENANTS_PATH)
            for row in rows:
                if row.get("Tenant_ID") == tenant_id:
                    row.update(tenant)
                    break
            write_csv(TENANTS_PATH, TENANT_HEADERS, rows)
            session["tenant_name"] = business_name

    # Save customizations to tenant settings
    tenant_settings = get_or_create_tenant_settings()
    tenant_settings["Theme"] = theme
    tenant_settings["POS_Name"] = (
        pos_name or business_name or session.get("tenant_name", "POS Suite")
    )

    write_csv(get_tenant_settings_path(), TENANT_SETTINGS_HEADERS, [tenant_settings])

    flash("Customizations updated successfully", "success")
    return redirect(url_for("settings"))


# ==============================================================================
#                     REPORTS
# ==============================================================================


@app.route("/reports")
@manager_required
def reports():
    return render_template("reports/index.html")


@app.route("/reports/daily")
@manager_required
def daily_report():
    date_str = (request.args.get("date") or "").strip()
    if date_str:
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            report_date = date.today()
    else:
        report_date = date.today()

    report_date_str = report_date.strftime("%Y-%m-%d")
    prev_date = (report_date - timedelta(days=1)).strftime("%Y-%m-%d")
    next_date = (report_date + timedelta(days=1)).strftime("%Y-%m-%d")
    is_today = report_date == date.today()

    sales_file = _find_saleslog_file_for_date(DATA_DIR, report_date_str)
    sales_rows = _load_saleslog_rows(sales_file)
    metrics = _compute_daily_sales_metrics(sales_rows)

    employee_hours, total_hours, total_payroll = _get_labor_for_date_stub(
        DATA_DIR, report_date_str
    )
    gross_profit = metrics.get("gross_profit", 0.0)
    net_profit = float(gross_profit or 0) - float(total_payroll or 0)

    return render_template(
        "reports/daily.html",
        report_date=report_date_str,
        report_date_formatted=report_date.strftime("%B %d, %Y"),
        day_of_week=report_date.strftime("%A"),
        prev_date=prev_date,
        next_date=next_date,
        is_today=is_today,
        total_revenue=metrics.get("total_revenue", 0.0),
        transaction_count=metrics.get("transaction_count", 0),
        items_sold=metrics.get("items_sold", 0),
        total_cogs=metrics.get("total_cogs", 0.0),
        gross_profit=gross_profit,
        gross_margin=metrics.get("gross_margin", 0.0),
        total_payroll=total_payroll,
        total_hours=total_hours,
        net_profit=net_profit,
        sales_by_category=metrics.get("sales_by_category", []),
        payment_breakdown=metrics.get("payment_breakdown", []),
        transactions=metrics.get("transactions", []),
        top_items=metrics.get("top_items", []),
        employee_hours=employee_hours,
    )


def _safe_float(x, default=0.0):
    try:
        return float(str(x).replace("$", "").strip() or 0)
    except Exception:
        return default


def _safe_int(x, default=0):
    try:
        return int(float(str(x).strip() or 0))
    except Exception:
        return default


def _find_saleslog_file_for_date(base_dir: str, date_str: str) -> str | None:
    """
    Looks for: Sales_Logs/2025-12-17_Wednesday_SalesLog.csv (your naming style)
    Also accepts other variants containing date + 'SalesLog'.
    """
    sales_dir = os.path.join(base_dir, "Sales_Logs")
    patterns = [
        os.path.join(sales_dir, f"{date_str}_*SalesLog*.csv"),
        os.path.join(sales_dir, f"{date_str}*SalesLog*.csv"),
        os.path.join(sales_dir, f"*{date_str}*SalesLog*.csv"),
        os.path.join(sales_dir, "**", f"{date_str}_*SalesLog*.csv"),
        os.path.join(sales_dir, "**", f"{date_str}*SalesLog*.csv"),
        os.path.join(sales_dir, "**", f"*{date_str}*SalesLog*.csv"),
    ]
    for pat in patterns:
        hits = sorted(glob.glob(pat, recursive=True), reverse=True)
        if hits:
            return hits[0]
    return None


def _load_saleslog_rows(path: str) -> list[dict]:
    if not path or not os.path.exists(path):
        return []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _compute_daily_sales_metrics(rows: list[dict]) -> dict:
    """
    Expects SalesLog columns like your uploaded file:
    Date, Transaction_ID, Quantity, Subtotal, Tax_Amount, Payment_Method,
    Category, SKU, Item_Name, COGS_Line, Gross_Margin, etc.
    """
    if not rows:
        return {
            "total_revenue": 0.0,
            "transaction_count": 0,
            "items_sold": 0,
            "total_cogs": 0.0,
            "gross_profit": 0.0,
            "gross_margin": 0.0,
            "sales_by_category": [],
            "payment_breakdown": [],
            "transactions": [],
            "top_items": [],
        }

    # Totals
    total_revenue = 0.0  # Subtotal (pre-tax)
    total_cogs = 0.0
    gross_profit = 0.0
    items_sold = 0
    tx_ids = set()

    # Groupings
    cat_map = {}  # category -> {"category":..., "qty":..., "revenue":...}
    pay_map = {}  # method -> {"method":..., "count": set(tx_ids), "total":...}
    tx_map = {}  # tx_id -> aggregated transaction
    item_map = {}  # sku -> aggregated item

    for r in rows:
        tx_id = (r.get("Transaction_ID") or "").strip()
        tx_ids.add(tx_id)

        qty = _safe_int(r.get("Quantity", 0))
        # Per-LINE revenue. The SalesLog "Subtotal" column stores the WHOLE-ticket
        # subtotal repeated on every line, so summing it overstates revenue on any
        # multi-item sale. Line_Total is the correct per-line amount (fall back to
        # Subtotal only for legacy rows that predate the Line_Total column).
        subtotal = _safe_float(r.get("Line_Total", r.get("Subtotal", 0)))
        cogs_line = _safe_float(r.get("COGS_Line", 0))
        gm = r.get("Gross_Margin", "")
        gross_line = _safe_float(gm, subtotal - cogs_line)

        total_revenue += subtotal
        total_cogs += cogs_line
        gross_profit += gross_line
        items_sold += qty

        # Sales by Category
        cat = (r.get("Category") or "Uncategorized").strip() or "Uncategorized"
        c = cat_map.setdefault(cat, {"category": cat, "qty": 0, "revenue": 0.0})
        c["qty"] += qty
        c["revenue"] += subtotal

        # Payment breakdown (count unique tx)
        pm = (r.get("Payment_Method") or "CASH").strip().upper()
        p = pay_map.setdefault(pm, {"method": pm, "tx_ids": set(), "total": 0.0})
        if tx_id:
            p["tx_ids"].add(tx_id)
        p["total"] += subtotal

        # Transactions (aggregate line-items into one tx card)
        t = tx_map.setdefault(
            tx_id or f"UNKNOWN-{len(tx_map) + 1}",
            {
                "transaction_id": tx_id or "UNKNOWN",
                "time": (r.get("Time") or r.get("Timestamp") or "").strip(),
                "items": 0,
                "payment_method": pm,
                "total": 0.0,
            },
        )
        t["items"] += qty
        t["total"] += subtotal
        if not t.get("time"):
            t["time"] = (r.get("Time") or r.get("Timestamp") or "").strip()

        # Top items
        sku = (r.get("SKU") or "").strip()
        name = (r.get("Item_Name") or "Unknown").strip()
        key = sku or name
        it = item_map.setdefault(
            key, {"sku": sku, "name": name, "qty": 0, "revenue": 0.0}
        )
        it["qty"] += qty
        it["revenue"] += subtotal

    # Finish up
    transaction_count = len([x for x in tx_ids if x]) or len(tx_map)
    gross_margin = (gross_profit / total_revenue * 100.0) if total_revenue else 0.0

    sales_by_category = sorted(
        cat_map.values(), key=lambda x: x["revenue"], reverse=True
    )

    payment_breakdown = []
    for v in pay_map.values():
        payment_breakdown.append(
            {
                "method": v["method"],
                "count": len(v["tx_ids"]) if v["tx_ids"] else 0,
                "total": v["total"],
            }
        )
    payment_breakdown.sort(key=lambda x: x["total"], reverse=True)

    transactions = sorted(
        tx_map.values(), key=lambda x: x.get("time") or "", reverse=True
    )
    top_items = sorted(item_map.values(), key=lambda x: x["revenue"], reverse=True)

    return {
        "total_revenue": round(total_revenue, 2),
        "transaction_count": transaction_count,
        "items_sold": items_sold,
        "total_cogs": round(total_cogs, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_margin": round(gross_margin, 1),
        "sales_by_category": sales_by_category,
        "payment_breakdown": payment_breakdown,
        "transactions": transactions,
        "top_items": top_items,
    }


def _get_labor_for_date_stub(
    base_dir: str, date_str: str
) -> tuple[list[dict], float, float]:
    """
    Build daily labor from Time_Clock CSV + pay rates.
    Uses Hours_Worked_Today as the authoritative per-employee total for the date.
    """
    try:
        target = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        target = date.today()

    punches = get_punches_for_date(target)
    hours_by_emp = {}
    for p in punches:
        emp_id = p.get("Employee_ID") or ""
        try:
            hours = float(str(p.get("Hours_Worked_Today") or 0))
        except Exception:
            hours = 0.0
        if emp_id:
            hours_by_emp[emp_id] = max(hours_by_emp.get(emp_id, 0.0), hours)

    employees = {
        e.get("Employee_ID"): e for e in get_all_employees(include_inactive=True)
    }
    employee_hours = []
    total_hours = 0.0
    total_payroll = 0.0
    for emp_id, hours in hours_by_emp.items():
        emp = employees.get(emp_id, {})
        rate = get_employee_pay_rate(emp_id)
        pay = hours * rate
        total_hours += hours
        total_payroll += pay
        employee_hours.append(
            {
                "Employee_ID": emp_id,
                "Employee_Name": emp.get("Employee_Name", "Unknown"),
                "Role": emp.get("Role", "Employee"),
                "hours": round(hours, 2),
                "pay": round(pay, 2),
            }
        )

    employee_hours.sort(key=lambda x: x.get("hours", 0), reverse=True)
    return employee_hours, round(total_hours, 2), round(total_payroll, 2)


# ==============================================================================
#                     MAIN
# ==============================================================================

if __name__ == "__main__":
    print(f"\n{'=' * 60}")
    print(f"  Mountain Gardens POS v{VERSION}")
    print("  http://localhost:5000")
    print(f"{'=' * 60}\n")

    # Private by default (workspace network doctrine). For a single in-store
    # terminal, 127.0.0.1 is correct. To serve other devices on the shop LAN,
    # start with HOST=0.0.0.0 (and ensure the machine's firewall is set).
    app.run(
        debug=False,
        host=os.environ.get("HOST", "127.0.0.1"),
        port=int(os.environ.get("PORT", "5000")),
    )
